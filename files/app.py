"""
app.py — Interfaz Streamlit del CRM + asistente de prospección B2B.

Ejecutar:  streamlit run app.py

Toda la lógica vive en crm_core.py. Esta capa solo orquesta:
selección de industria, configuración de campaña, vista previa con aprobación
humana, marcado de envíos, bandeja de respuestas, follow-ups por lote, estados
y dashboard con las 8 gráficas. La app NUNCA envía sola: prepara el mensaje, la
persona aprueba y envía por su cuenta, y luego marca "Enviado".
"""

import datetime as dt
from pathlib import Path

import pandas as pd
import streamlit as st

from crm_core import (
    CRM, CHANNEL_STEPS, STAGE_COLS, INDUSTRIES, RESPONSE_CLASSES,
    RESPONSE_RESULTS, OUTCOME_VALUES, ACTIVITY_CHANNELS, IMPORT_FIELDS,
    TEMPLATE_COLUMNS, MANUAL_STATES, PROSPECT_CHANNELS, WARMTH_LEVELS,
    SCHEDULE_STATUS, FOLLOWUP_CHANNELS, COLD_CALL_RESULTS, STATE_CATEGORIES,
    NOTIF_SETTINGS_SHEET, WORKFLOW_SHEET, NOTIF_DEFAULTS, WORKFLOW_DEFAULTS,
    LINKEDIN_QUEUE_STATUS, QUICK_RESCHEDULE, ICP_FIELDS, ICP_DEFAULTS,
    FIT_LEVELS, StateStore, _is_marked, _norm,
    create_blank_crm,
)
import pandas as pd
import integrations
import gmail_service
import linkedin_inbox as li_inbox
import linkedin_session as ls_session

st.set_page_config(page_title="Prospecting CRM", layout="wide")

# --------------------------------------------------------------------------- #
# Bootstrap: la app arranca SIN necesidad de un Excel existente
# --------------------------------------------------------------------------- #

import shutil

WORKSPACE = Path("workspace")
WORKSPACE.mkdir(exist_ok=True)

DEMO_SOURCE = Path("files/crm_DEMO.xlsx")
DEFAULT_XLSX = WORKSPACE / "crm_actual.xlsx"

st.sidebar.title("⚙️ Configuración")

modo_crm = st.sidebar.radio(
    "Cómo iniciar CRM",
    ["Usar archivo demo", "Subir mi propio Excel", "Empezar CRM vacío"]
)

uploaded_excel = None

if modo_crm == "Usar archivo demo":
    if not DEFAULT_XLSX.exists():
        shutil.copy(DEMO_SOURCE, DEFAULT_XLSX)
    xlsx_path = str(DEFAULT_XLSX)

elif modo_crm == "Subir mi propio Excel":
    uploaded_excel = st.sidebar.file_uploader(
        "Sube tu Excel CRM",
        type=["xlsx"]
    )

    if uploaded_excel is not None:
        with open(DEFAULT_XLSX, "wb") as f:
            f.write(uploaded_excel.getbuffer())
        xlsx_path = str(DEFAULT_XLSX)
    else:
        st.info("Sube un Excel para iniciar.")
        st.stop()

else:
    EMPTY_XLSX = WORKSPACE / "crm_vacio.xlsx"

    if not EMPTY_XLSX.exists():
        shutil.copy(DEMO_SOURCE, EMPTY_XLSX)
        # Por ahora se usa el demo como plantilla estructural.
        # Luego puedes borrar leads desde la app o ajustar crm_core.py
        # para crear hojas vacías reales.

    xlsx_path = str(EMPTY_XLSX)


@st.cache_resource(show_spinner="Cargando CRM…")
def get_crm(path, mtime):
    return CRM(path)


def _looks_like_crm(data: bytes) -> bool:
    """True si el xlsx subido ya es un CRM (algún encabezado con Outcome
    Status); False si parece una lista de leads para importar."""
    import io
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True)
        for ws in wb.worksheets:
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True),
                          ())
            if any(str(h).strip() == "Outcome Status" for h in header if h):
                return True
    except Exception:
        pass
    return False


if "crm_path" not in st.session_state:
    st.title("👋 Bienvenido a tu Prospecting CRM")
    st.write("Elige cómo quieres empezar. No necesitas tener un Excel: "
             "la app puede crear uno por ti.")

    c1, c2, c3 = st.columns(3)

    with c1:
        st.subheader("🎬 Archivo demo")
        st.caption("Un CRM de ejemplo con leads, campañas y notificaciones "
                   "para explorar la app.")
        if st.button("Usar archivo demo", type="primary",
                     use_container_width=True):
            dest = WORKSPACE / "demo_crm.xlsx"
            if not dest.exists():
                if DEMO_SOURCE.exists():
                    import shutil
                    shutil.copy(DEMO_SOURCE, dest)
                else:   # sin demo empaquetado: crear uno con las 9 industrias
                    create_blank_crm(str(dest), industries=INDUSTRIES)
            st.session_state["crm_path"] = str(dest)
            st.rerun()

    with c2:
        st.subheader("📤 Subir mi Excel")
        st.caption("Si ya es un CRM de esta app, se usa directo. Si es una "
                   "lista de leads, se crea un CRM y se abre el asistente de "
                   "importación (preview + mapeo de columnas).")
        up = st.file_uploader("Archivo .xlsx / .csv", type=["xlsx", "xls", "csv"],
                              key="boot_up")
        if up is not None and st.button("Continuar con este archivo",
                                        type="primary",
                                        use_container_width=True):
            data = up.getvalue()
            if up.name.lower().endswith((".xlsx", ".xls")) and \
                    _looks_like_crm(data):
                dest = WORKSPACE / f"subido_{up.name}"
                dest.write_bytes(data)
                st.session_state["crm_path"] = str(dest)
            else:
                dest = WORKSPACE / "mi_crm.xlsx"
                if not dest.exists():
                    create_blank_crm(str(dest))
                st.session_state["crm_path"] = str(dest)
                st.session_state["pending_import"] = (up.name, data)
                st.session_state["goto_import"] = True
            st.rerun()

    with c3:
        st.subheader("🆕 CRM vacío")
        st.caption("Estructura base lista para capturar o importar leads: "
                   "Leads, Activity_Log, Notifications, Campaigns, "
                   "Follow_Ups y Scheduled_Messages.")
        if st.button("Empezar CRM vacío", type="primary",
                     use_container_width=True):
            dest = WORKSPACE / "crm_vacio.xlsx"
            if not dest.exists():
                create_blank_crm(str(dest))
            st.session_state["crm_path"] = str(dest)
            st.rerun()

    if Path(DEFAULT_XLSX).exists():
        st.divider()
        if st.button(f"📂 Abrir archivo detectado: {DEFAULT_XLSX}"):
            st.session_state["crm_path"] = DEFAULT_XLSX
            st.rerun()

    st.stop()

xlsx_path = st.session_state["crm_path"]
if not Path(xlsx_path).exists():
    st.error(f"El archivo de trabajo desapareció: {xlsx_path}")
    if st.button("Volver a la pantalla inicial"):
        st.session_state.pop("crm_path", None)
        st.rerun()
    st.stop()

st.sidebar.title("⚙️ Configuración")
st.sidebar.caption(f"📄 {Path(xlsx_path).name}")
with st.sidebar.expander("⚙️ Configuración avanzada"):
    st.caption(f"Ruta actual: `{xlsx_path}`")
    manual_path = st.text_input("Ruta manual del Excel", xlsx_path)
    if st.button("Usar esta ruta"):
        if Path(manual_path).exists():
            st.session_state["crm_path"] = manual_path
            st.rerun()
        else:
            st.error("Esa ruta no existe.")
    if st.button("🔁 Cambiar de archivo (pantalla inicial)"):
        st.session_state.pop("crm_path", None)
        st.rerun()

crm = get_crm(xlsx_path, Path(xlsx_path).stat().st_mtime)
IND = crm.industries()   # hojas de leads dinámicas (funciona con CRM vacío)

def persist():
    """Guarda con backup previo y avisa claramente si el Excel está abierto."""
    crm.ensure_formatting()
    out, bkp, err = crm.save()
    if err:
        st.error("⚠️ " + err)
        return False
    msg = "Excel y estado guardados ✅"
    if bkp:
        msg += f"  ·  backup: {Path(bkp).name}"
    st.toast(msg)
    return True

# --------------------------------------------------------------------------- #
# Navegación
# --------------------------------------------------------------------------- #

_nav = ["📣 Campaña", "📮 Gmail Campaigns", "📧 Email Campaigns",
        "💼 LinkedIn Manager", "📤 Importar leads", "🔁 Follow-ups",
        "📥 Respuestas", "🔔 Notifications", "🏷️ Estados",
        "🎨 Configuración de Estados", "⚙️ Workflow Config", "📊 Dashboard"]
_default_page = _nav.index("📤 Importar leads") \
    if st.session_state.pop("goto_import", False) else 0
page = st.sidebar.radio(
    "Sección", _nav, index=_default_page,
)


def _render_inbox():
    st.caption("Sin scraping ni automatización del navegador (política de "
               "LinkedIn y regla del proyecto). Fuentes: registro manual y "
               "la **exportación oficial** de LinkedIn (Configuración → "
               "Obtener una copia de tus datos → messages.csv). La capa de "
               "proveedores permite conectar en el futuro una API oficial "
               "sin reescribir el CRM.")
    prov_name = st.radio("Fuente", ["Exportación oficial (messages.csv)",
                                    "Registro manual"], horizontal=True)
    if prov_name.startswith("Exportación"):
        own = st.text_input("Tu nombre en LinkedIn (para filtrar tus "
                            "propios mensajes)", key="li_own")
        up = st.file_uploader("messages.csv de tu exportación oficial",
                              type=["csv"], key="li_csv")
        if up is not None and st.button("🔄 Sync Inbox", type="primary"):
            prov = li_inbox.DataExportProvider(up.getvalue(), own_name=own)
            res = li_inbox.sync_inbox(crm, prov)
            if persist():
                st.success(f"Procesados: {res.get('processed', 0)} · "
                           f"duplicados omitidos: {res.get('duplicates', 0)} · "
                           f"asociados a lead: {res.get('matched', 0)} · "
                           f"a revisión: {res.get('review', 0)}")
    else:
        m1, m2 = st.columns(2)
        s_name = m1.text_input("Remitente (nombre)", key="li_ms")
        s_url = m2.text_input("LinkedIn URL del remitente", key="li_mu")
        s_txt = st.text_area("Mensaje recibido (pégalo aquí)", key="li_mt")
        if st.button("Registrar y procesar", type="primary"):
            prov = li_inbox.ManualProvider()
            prov.add(s_name, s_txt, profile_url=s_url)
            res = li_inbox.sync_inbox(crm, prov)
            if persist():
                r0 = res
                st.success(f"Procesado. Asociado a lead: "
                           f"{'sí' if r0.get('matched') else 'no'} · "
                           f"a revisión: {r0.get('review', 0)}")

    # bandeja + cola de revisión humana
    if "LinkedIn_Inbox" in crm.wb.sheetnames:
        ws = crm.wb["LinkedIn_Inbox"]
        rows_ib = [[ws.cell(r, c).value for c in
                    range(1, len(li_inbox.LI_INBOX_HEADERS) + 1)]
                   for r in range(2, ws.max_row + 1)]
        if rows_ib:
            df_ib = pd.DataFrame(rows_ib,
                                 columns=li_inbox.LI_INBOX_HEADERS)
            rev = df_ib[df_ib["Requires Review"] == "Sí"]
            if len(rev):
                st.subheader(f"🧑‍⚖️ Revisión humana pendiente ({len(rev)})")
                st.dataframe(rev, use_container_width=True,
                             hide_index=True)
                st.caption("Clasifica estos casos manualmente en "
                           "Notifications → Registrar respuesta, o en "
                           "Estados.")
            st.subheader("Bandeja (últimos 50)")
            st.dataframe(df_ib.tail(50), use_container_width=True,
                         hide_index=True)

    # ---------- Centro de notificaciones LinkedIn ----------

def _render_li_notifications():
    unread = li_inbox.read_li_notifications(crm, unread_only=True)
    st.metric("🔔 No leídas", len(unread))
    only_unread = st.checkbox("Ver solo no leídas", value=True)
    notis = unread if only_unread else             li_inbox.read_li_notifications(crm)
    if not notis:
        st.info("Sin notificaciones LinkedIn por ahora.")
    for n in reversed(notis):
        pri = {"Alta": "🔴", "Media": "🟡", "Baja": "⚪"}.get(
            str(n.get("priority")), "🟡")
        with st.expander(f"{pri} {n.get('type')} · {n.get('title')} · "
                         f"{str(n.get('created_at'))[:16]}"):
            st.write(n.get("message") or "—")
            st.caption(f"id: {n.get('id')} · lead: "
                       f"{n.get('lead_id') or '—'} · ref: "
                       f"{n.get('action_url_or_reference') or '—'}")
            c1, c2 = st.columns(2)
            if not n.get("read_at"):
                if c1.button("Mark as Read", key=f"nc_r{n['_row']}"):
                    li_inbox.mark_li_notification_read(crm, n["_row"])
                    if persist():
                        st.rerun()
            if n.get("lead_id"):
                if c2.button("Open Lead", key=f"nc_o{n['_row']}"):
                    ld = crm._find_lead_by_key(n["lead_id"])
                    if ld:
                        st.session_state["resp_lead"] = (ld.sheet, ld.row)
                        st.info(f"Lead cargado: {ld.full_name}. Ábrelo en "
                                "la sección 📥 Respuestas o 🏷️ Estados.")
                    else:
                        st.warning("No encontré ese lead en el CRM.")
    if notis and st.button("Marcar todas como leídas"):
        for n in li_inbox.read_li_notifications(crm, unread_only=True):
            li_inbox.mark_li_notification_read(crm, n["_row"])
        if persist():
            st.rerun()



def render_reschedule(lead, prefix):
    """Formulario reutilizable de reprogramación de follow-up (Follow-ups,
    Gmail, Cold Call, Estados y Notifications). Guarda fila + historial +
    Activity_Log + notificación."""
    with st.expander("🗓️ Reprogramar follow-up"):
        quick = st.radio("Rápido", [q[0] for q in QUICK_RESCHEDULE],
                         horizontal=True, key=f"{prefix}_q")
        qh = dict(QUICK_RESCHEDULE)[quick]
        base = dt.datetime.now() + dt.timedelta(hours=qh) if qh else None
        r1, r2, r3 = st.columns(3)
        nd = r1.date_input("Nueva fecha",
                           base.date() if base else
                           dt.date.today() + dt.timedelta(days=3),
                           key=f"{prefix}_d")
        nt = r2.time_input("Hora", dt.time(9, 0), key=f"{prefix}_t")
        ch = r3.selectbox("Canal", FOLLOWUP_CHANNELS, key=f"{prefix}_c")
        r4, r5 = st.columns(2)
        reason = r4.text_input("Motivo de reprogramación", key=f"{prefix}_r")
        user = r5.text_input("Owner/User", key=f"{prefix}_u")
        note = st.text_input("Nota", key=f"{prefix}_n")
        if st.button("Guardar reprogramación", key=f"{prefix}_go", type="primary"):
            step, when = crm.reschedule_followup(
                lead, nd.isoformat(), new_time=nt.strftime("%H:%M"), channel=ch,
                reason=reason, notes=note or None, user=user)
            if persist():
                st.success(f"✅ Reprogramado {step or 'próximo follow-up'} → {when}. "
                           "Guardado en Follow_Up_History, Activity_Log y Notifications.")
                return True
    return False


def gmail_oauth_panel(cfg):
    """Panel único de conexión Gmail con PKCE persistido.
    Devuelve el dict de credenciales si hay sesión conectada; si no, dibuja el
    botón de conexión / maneja el callback y devuelve None."""
    creds = st.session_state.get("gmail_creds")
    if creds is not None:
        return creds

    qp = st.query_params
    if "code" in qp:
        # recuperar el verifier: session_state primero; si el redirect creó una
        # sesión nueva (típico en Streamlit), lo recuperamos del 'state'.
        verifier = st.session_state.get("gmail_code_verifier") or             gmail_service.recover_verifier(qp.get("state", ""))
        if not verifier:
            st.error("Se perdió la sesión de autorización (falta code_verifier).")
            if st.button("🔄 Reintentar conexión Gmail", key="gm_retry_missing"):
                for k in ("gmail_auth_url", "gmail_oauth_state",
                          "gmail_code_verifier", "gmail_creds"):
                    st.session_state.pop(k, None)
                st.query_params.clear()
                st.rerun()
            return None
        try:
            creds = gmail_service.exchange_code(cfg, qp["code"], verifier)
            st.session_state["gmail_creds"] = creds
            for k in ("gmail_auth_url", "gmail_oauth_state",
                      "gmail_code_verifier"):
                st.session_state.pop(k, None)
            st.query_params.clear()   # borra code/state/verifier de la URL
            st.success("✅ Gmail conectado")
            st.rerun()
        except Exception as e:
            st.error(f"No pude completar la autorización: {e}")
            if st.button("🔄 Reintentar conexión Gmail", key="gm_retry_exc"):
                for k in ("gmail_auth_url", "gmail_oauth_state",
                          "gmail_code_verifier", "gmail_creds"):
                    st.session_state.pop(k, None)
                st.query_params.clear()
                st.rerun()
        return None

    # generar la URL UNA sola vez por sesión (el verifier debe ser el mismo
    # que viaje en la URL que el usuario abre)
    if "gmail_auth_url" not in st.session_state:
        try:
            auth_url, state, verifier = gmail_service.build_auth_url(cfg)
            st.session_state["gmail_auth_url"] = auth_url
            st.session_state["gmail_oauth_state"] = state
            st.session_state["gmail_code_verifier"] = verifier
        except Exception as e:
            st.error(f"No pude generar la URL de autorización: {e}")
            return None
    st.link_button("🔗 Conectar Gmail", st.session_state["gmail_auth_url"],
                   type="primary")
    st.caption("Se abrirá el consentimiento de Google. Al autorizar, volverás "
               "a la app.")
    return None

st.sidebar.caption("Aprobación humana obligatoria antes de cada envío. "
                   "La app no automatiza LinkedIn ni evade límites. "
                   "Backup automático antes de cada guardado.")

# =========================================================================== #
# 1) CAMPAÑA
# =========================================================================== #
if page == "📣 Campaña":
    st.header("📣 Nueva campaña")

    # --- Tabla de leads disponibles (con filtros) ANTES de prospectar ---
    st.subheader("Leads disponibles")

    @st.cache_data(show_spinner="Cargando leads disponibles…")
    def _available_table(ind, mtime):
        rows = []
        for l in crm.all_leads(ind):
            stt = crm.state.get(l.key)
            last_contact = crm._cell(crm.wb[l.sheet], crm.maps[l.sheet],
                                     l.row, "Last Contact")
            rows.append({
                "Nombre": l.full_name, "Empresa": l.company, "Puesto": l.job_title,
                "Seniority": l.seniority_level or "—", "Industria": l.industry,
                "LinkedIn": l.linkedin, "Email": l.email,
                "Prioridad": crm.lead_priority(l),
                "Estado": crm.lead_status(l, stt),
                "Último contacto": str(last_contact or ""),
                "Siguiente acción": crm.next_action(l, stt),
            })
        return pd.DataFrame(rows)

    industry = st.selectbox("Industria", IND)
    avail = _available_table(industry, Path(xlsx_path).stat().st_mtime)

    fc = st.columns(4)
    f_sen = fc[0].multiselect("Seniority", sorted(avail["Seniority"].unique()))
    f_comp = fc[1].multiselect("Empresa", sorted(avail["Empresa"].dropna().unique())[:500])
    f_prio = fc[2].multiselect("Prioridad", ["Alta", "Media-Alta", "Media", "Baja"])
    f_est = fc[3].multiselect("Estado", sorted(avail["Estado"].unique()))
    view = avail.copy()
    if f_sen: view = view[view["Seniority"].isin(f_sen)]
    if f_comp: view = view[view["Empresa"].isin(f_comp)]
    if f_prio: view = view[view["Prioridad"].isin(f_prio)]
    if f_est: view = view[view["Estado"].isin(f_est)]
    st.dataframe(view, use_container_width=True, hide_index=True)
    st.caption(f"{len(view)} leads en la vista (de {len(avail)} en {industry}).")

    # ======================================================================= #
    # Crear campaña filtrada (selección + programación de mensajes)
    # ======================================================================= #
    st.divider()
    st.subheader("🎯 Crear campaña filtrada")

    cf = st.columns(4)
    cf_channel = cf[0].selectbox("Canal de prospección", PROSPECT_CHANNELS)
    cf_country = cf[1].text_input("País/ubicación contiene")
    cf_warm = cf[2].multiselect("Nivel de calentamiento", WARMTH_LEVELS)
    cf_n = cf[3].number_input("Nº de leads", 1, 200, 30)

    # Construir el universo filtrado (reutiliza los filtros de arriba + estos)
    filtered = []
    for l in crm.all_leads(industry):
        stt = crm.state.get(l.key)
        if f_sen and (l.seniority_level or "—") not in f_sen: continue
        if f_comp and l.company not in f_comp: continue
        if f_prio and crm.lead_priority(l) not in f_prio: continue
        if f_est and crm.lead_status(l, stt) not in f_est: continue
        if cf_warm and crm.classify_warmth(l, stt) not in cf_warm: continue
        loc = crm._cell(crm.wb[l.sheet], crm.maps[l.sheet], l.row, "Location") or ""
        if cf_country and _norm(cf_country) not in _norm(str(loc)): continue
        filtered.append((l, stt, loc))
    filtered = filtered[:cf_n]

    if not filtered:
        st.info("Ningún lead cumple esos filtros. Ajusta los parámetros.")
    else:
        st.write(f"**{len(filtered)} leads** cumplen los filtros. "
                 "Selecciona/deselecciona y prepara el mensaje de cada uno.")
        msg_tmpl = st.text_area(
            "Plantilla de mensaje sugerido (usa {first}, {company}, {title})",
            "Hola {first}, vi tu rol en {company} y me gustaría conectar.")
        c_i1, c_i2 = st.columns(2)
        start = c_i1.text_input("Fecha/hora inicio (YYYY-MM-DD HH:MM)",
                                dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
        interval = c_i2.number_input("Intervalo entre leads (min)", 1, 1440, 45)
        try:
            t0 = dt.datetime.strptime(start, "%Y-%m-%d %H:%M")
        except ValueError:
            t0 = dt.datetime.now()

        selected = []
        for i, (l, stt, loc) in enumerate(filtered):
            first = (l.full_name or "").split(" ")[0]
            last_msg = ""
            for e in reversed(stt.log):
                if e.get("action") == "sent":
                    last_msg = (e.get("message") or "")[:50]; break
            last_contact = crm._cell(crm.wb[l.sheet], crm.maps[l.sheet],
                                     l.row, "Last Contact")
            send_at = t0 + dt.timedelta(minutes=i * interval)
            with st.expander(
                f"{i+1:>2}. {l.full_name or '(sin nombre)'} · {l.company} · "
                f"{crm.lead_priority(l)} · {crm.classify_warmth(l, stt)} · ⏱ {send_at:%m-%d %H:%M}"):
                cc = st.columns(2)
                cc[0].write(f"**Puesto:** {l.job_title or '—'}  \n"
                            f"**Seniority:** {l.seniority_level or '—'}  \n"
                            f"**Industria:** {l.industry}  \n"
                            f"**Ubicación:** {loc or '—'}")
                cc[1].write(f"**LinkedIn:** {l.linkedin or '—'}  \n"
                            f"**Email:** {l.email or '—'}  \n"
                            f"**Teléfono:** {l.phone or '—'}  \n"
                            f"**Canal recomendado:** {crm.recommend_channel(l)}")
                st.caption(f"Follow-up actual: {crm.lead_stage(l, stt)} · "
                           f"Último contacto: {last_contact or '—'} · "
                           f"Último mensaje: {last_msg or '—'}")
                default_msg = msg_tmpl.format(first=first, company=l.company,
                                              title=l.job_title)
                msg = st.text_area("Mensaje sugerido (editable)", default_msg,
                                   key=f"cf_msg{i}", height=80)
                pick = st.checkbox("Incluir en la campaña", value=True, key=f"cf_pick{i}")
                if pick:
                    selected.append((l, msg, send_at))

        st.write(f"Seleccionados: **{len(selected)}**")
        cbtn = st.columns(2)
        if cbtn[0].button("📝 Guardar como BORRADORES (Draft)", key="cf_draft"):
            for l, msg, send_at in selected:
                crm.schedule_message(l, cf_channel, msg, send_at.isoformat(timespec="minutes"),
                                     status="Draft")
            if persist():
                st.success(f"{len(selected)} mensajes guardados como Draft en "
                           "Scheduled_Messages.")
        if cbtn[1].button("📅 Programar (Scheduled)", type="primary", key="cf_sched"):
            for l, msg, send_at in selected:
                crm.schedule_message(l, cf_channel, msg, send_at.isoformat(timespec="minutes"),
                                     status="Scheduled")
            if persist():
                st.success(f"{len(selected)} mensajes programados. Revísalos y márcalos "
                           "como enviados manualmente en 'Agenda de mensajes'.")

    # --- Agenda de mensajes (revisar y marcar enviado manualmente) ---
    st.divider()
    st.subheader("🗓️ Agenda de mensajes (Scheduled_Messages)")
    st.caption("La app NO envía nada sola. Aquí revisas la cola y marcas cada mensaje "
               "como enviado tú mismo cuando lo mandaste.")
    sched = crm.read_scheduled()
    if not sched:
        st.info("No hay mensajes en la agenda todavía.")
    else:
        fstat = st.multiselect("Filtrar por estado", SCHEDULE_STATUS,
                               default=["Draft", "Scheduled"])
        shown = [s for s in sched if not fstat or s.get("Status") in fstat]
        st.dataframe(pd.DataFrame([{k: v for k, v in s.items() if k != "_row"}
                                   for s in shown]),
                     use_container_width=True, hide_index=True)
        ids = {f"{s['Schedule ID']} · {s['Full Name']} · {s['Status']}": s for s in shown}
        if ids:
            pick = st.selectbox("Actualizar mensaje", list(ids.keys()))
            row = ids[pick]["_row"]
            a1, a2, a3 = st.columns(3)
            if a1.button("Marcar ENVIADO (registro real)", type="primary"):
                crm.update_schedule_status(row, "Sent")
                if persist():
                    st.success("Marcado como enviado y registrada la fecha real.")
            if a2.button("Marcar Scheduled"):
                crm.update_schedule_status(row, "Scheduled")
                if persist(): st.success("Marcado como Scheduled.")
            if a3.button("Cancelar mensaje"):
                crm.update_schedule_status(row, "Cancelled")
                if persist(): st.success("Cancelado.")

    st.divider()
    st.subheader("Configurar campaña")
    c2, c3 = st.columns(2)
    channel = c2.selectbox("Canal", list(CHANNEL_STEPS.keys()))
    step_label = c3.selectbox("Paso", CHANNEL_STEPS[channel])
    step_index = CHANNEL_STEPS[channel].index(step_label)

    # --- Apartado Seniority Level (lee el del Excel y filtra la campaña) ---
    st.subheader("Seniority Level")
    sen_values = sorted({(l.seniority_level or "—") for l in crm.all_leads(industry)})
    sel_sen = st.multiselect(
        "Filtrar por Seniority Level (vacío = todos). Se usa para calcular prioridad.",
        sen_values, default=f_sen)
    st.caption("Prioridad: Owner/Founder/CEO/President = **Alta** · "
               "VP/Director/Head/Manager = **Media-Alta** · "
               "Buyer/Purchasing/Sourcing/Procurement/Engineer = **Media** · "
               "Assistant/Coordinator/Intern = **Baja**.")

    c4, c5, c6 = st.columns(3)
    n_leads = c4.number_input("Leads por campaña", 1, 200, 30)
    interval = c5.number_input("Intervalo entre leads (min)", 1, 1440, 45)
    start = c6.text_input("Inicio (YYYY-MM-DD HH:MM)",
                          dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
    gap_days = st.number_input("Programar siguiente follow-up en (días)", 1, 60, 3)

    tmpl = st.text_area(
        "Plantilla del mensaje (usa {first}, {company}, {title})",
        "Hola {first}, vi tu rol en {company} y me gustaría conectar.")

    if st.button("Construir cola", type="primary"):
        q = crm.build_queue(industry, channel, step_index, n_leads * 3)
        if sel_sen:
            q = [l for l in q if (l.seniority_level or "—") in sel_sen]
        st.session_state.queue = q[:n_leads]
        st.session_state.qmeta = dict(industry=industry, channel=channel,
                                      step=step_label, step_index=step_index,
                                      interval=interval, start=start, gap_days=gap_days,
                                      tmpl=tmpl)

    queue = st.session_state.get("queue", [])
    if queue:
        meta = st.session_state.qmeta
        st.success(f"{len(queue)} leads en cola · {meta['channel']} · {meta['step']} · "
                   f"1 cada {meta['interval']} min")
        # Distribución de prioridad en la cola
        prio_counts = {}
        for l in queue:
            p = crm.lead_priority(l)
            prio_counts[p] = prio_counts.get(p, 0) + 1
        st.caption("Prioridad en la cola: " +
                   " · ".join(f"{k}: {v}" for k, v in prio_counts.items()))
        try:
            t0 = dt.datetime.strptime(meta["start"], "%Y-%m-%d %H:%M")
        except ValueError:
            t0 = dt.datetime.now()

        st.subheader("Vista previa — aprueba lead por lead")
        approvals = {}
        for i, lead in enumerate(queue):
            send_at = t0 + dt.timedelta(minutes=i * meta["interval"])
            first = (lead.full_name or "").split(" ")[0]
            msg = meta["tmpl"].format(first=first, company=lead.company,
                                      title=lead.job_title)
            prio = crm.lead_priority(lead)
            with st.expander(
                f"{i+1:>2}. {lead.full_name} · {lead.company} · "
                f"{lead.seniority_level or '—'} · 🎯 {prio} · ⏱ {send_at:%H:%M}"):
                msg = st.text_area("Mensaje", msg, key=f"msg{i}", height=90)
                approvals[i] = st.checkbox("✅ Aprobar este mensaje", key=f"ap{i}")
                st.session_state[f"finalmsg{i}"] = msg

        st.divider()
        if st.button("Marcar aprobados como ENVIADOS", type="primary"):
            sent = 0
            for i, lead in enumerate(queue):
                if approvals.get(i):
                    msg = st.session_state.get(f"finalmsg{i}", "")
                    crm.mark_sent(lead, meta["channel"], meta["step"], message=msg)
                    crm.schedule_next(lead, meta["channel"], meta["step_index"],
                                      meta["gap_days"])
                    sent += 1
            if persist():
                st.success(f"{sent} marcados como enviados y siguiente follow-up programado.")
                st.session_state.queue = crm.build_queue(
                    meta["industry"], meta["channel"], meta["step_index"], len(queue))

# =========================================================================== #
# 2) IMPORTAR LEADS
# =========================================================================== #
elif page == "📤 Importar leads":
    st.header("📤 Importar leads desde Excel/CSV")

    with st.expander("ℹ️ Formato esperado del archivo", expanded=True):
        st.markdown(
            "El archivo puede ser **.xlsx** o **.csv**. No necesita todas las columnas: "
            "se importa cada lead **con lo que tenga**. Columnas recomendadas (los nombres "
            "pueden variar; la app las detecta y puedes corregir el mapeo):")
        st.dataframe(pd.DataFrame([
            {"First Name": "Jane", "Last Name": "Doe", "Full Name": "Jane Doe",
             "Job Title": "VP Procurement", "Company Name": "Acme Manufacturing",
             "Industry": "Automotive", "Seniority Level": "VP",
             "LinkedIn URL": "linkedin.com/in/janedoe", "Email": "jane@acme.com",
             "Phone": "+1 555 0100", "Location": "Detroit, MI",
             "Value Proposition": "Reducimos 15% tooling", "Pain Point": "Lead times",
             "Notes": "Feria 2025"},
            {"First Name": "", "Last Name": "", "Full Name": "John Smith",
             "Job Title": "Buyer", "Company Name": "Globex", "Industry": "Packaging",
             "Seniority Level": "", "LinkedIn URL": "linkedin.com/in/johnsmith",
             "Email": "john@globex.com", "Phone": "", "Location": "",
             "Value Proposition": "", "Pain Point": "Empaque sustentable", "Notes": ""},
        ]), use_container_width=True, hide_index=True)
        tmpl_path = Path(xlsx_path).resolve().parent / "plantilla_importacion_leads.xlsx"
        crm.make_import_template(str(tmpl_path.parent))
        with open(tmpl_path, "rb") as fh:
            st.download_button("⬇️ Descargar plantilla_importacion_leads.xlsx", fh.read(),
                               file_name="plantilla_importacion_leads.xlsx",
                               mime="application/vnd.openxmlformats-officedocument."
                                    "spreadsheetml.sheet")

    # --- Alta manual de leads (funciona también con CRM vacío) ---
    with st.expander("➕ Agregar lead manualmente"):
        m1, m2, m3 = st.columns(3)
        mf_name = m1.text_input("Nombre completo", key="man_name")
        mf_comp = m2.text_input("Empresa", key="man_comp")
        mf_title = m3.text_input("Puesto", key="man_title")
        m4, m5, m6 = st.columns(3)
        mf_email = m4.text_input("Email", key="man_email")
        mf_li = m5.text_input("LinkedIn URL", key="man_li")
        mf_phone = m6.text_input("Teléfono", key="man_phone")
        m7, m8, m9 = st.columns(3)
        mf_loc = m7.text_input("Ubicación", key="man_loc")
        mf_sen = m8.text_input("Seniority Level", key="man_sen")
        mf_sheet = m9.selectbox("Hoja destino", IND, key="man_sheet")
        mf_notes = st.text_input("Notas", key="man_notes")
        if st.button("Agregar lead", type="primary", key="man_go"):
            if not (mf_name or mf_comp or mf_email or mf_li):
                st.error("Captura al menos nombre, empresa, email o LinkedIn.")
            else:
                added, dups = crm.add_lead_manual(
                    mf_sheet, **{"Full Name": mf_name, "Company Name": mf_comp,
                                 "Job Title": mf_title, "Email/Gmail": mf_email,
                                 "LinkedIn Profile": mf_li, "Phone": mf_phone,
                                 "Location": mf_loc, "Seniority Level": mf_sen,
                                 "Notes": mf_notes, "Industry": mf_sheet})
                if persist():
                    if added:
                        st.success(f"✅ Lead agregado a '{mf_sheet}'.")
                    else:
                        st.warning("Ese lead ya existía (duplicado).")

    ups = st.file_uploader("Sube uno o varios archivos .xlsx/.csv",
                           type=["xlsx", "xls", "csv"], accept_multiple_files=True)
    pending = st.session_state.get("pending_import")
    if pending and not ups:
        import io
        st.info(f"📎 Archivo de la pantalla inicial listo para importar: "
                f"**{pending[0]}**")
        _bio = io.BytesIO(pending[1]); _bio.name = pending[0]
        ups = [_bio]
    if ups:
        frames = []
        for up in ups:
            try:
                d = pd.read_csv(up) if up.name.lower().endswith(".csv") else pd.read_excel(up)
                d["__archivo__"] = up.name
                frames.append(d)
            except Exception as e:
                st.error(f"No pude leer {up.name}: {e}")
        if frames:
            df = pd.concat(frames, ignore_index=True)
            st.subheader("Vista previa consolidada")
            st.caption(f"{len(ups)} archivo(s) · {len(df)} filas combinadas")
            st.dataframe(df.head(30), use_container_width=True)

            src_cols = [c for c in df.columns if c != "__archivo__"]
            auto = crm.detect_import_mapping(src_cols)
            st.subheader("Mapeo de columnas")
            st.caption("Detecté estas correspondencias. Ajusta manualmente si hace falta.")
            none_opt = "—(ninguna)—"
            options = [none_opt] + src_cols
            mapping = {}
            cols = st.columns(2)
            for i, field in enumerate(IMPORT_FIELDS):
                default = auto.get(field)
                idx = options.index(default) if default in options else 0
                with cols[i % 2]:
                    sel = st.selectbox(field, options, index=idx, key=f"map_{field}")
                if sel != none_opt:
                    mapping[field] = sel

            # Pre-cálculo de duplicados en el consolidado (informativo)
            def row_key(r):
                g = lambda f: (str(r[mapping[f]]).strip()
                               if mapping.get(f) and pd.notna(r.get(mapping[f])) else "")
                full = g("Full Name") or (g("First Name") + " " + g("Last Name")).strip()
                return StateStore.key(g("LinkedIn Profile") or None, g("Email/Gmail") or None,
                                      full or None, g("Company Name") or None,
                                      g("Phone") or None)
            if mapping:
                keys = [row_key(r) for _, r in df.iterrows()]
                dup_in_file = len(keys) - len(set(keys))
                st.caption(f"Duplicados detectados dentro del archivo: {dup_in_file} "
                           "(se omitirán; también se omiten los que ya existan en el CRM).")

            c1, c2 = st.columns(2)
            route = c1.checkbox("Enrutar por columna de industria", value=True)
            default_ind = c2.selectbox("Industria por defecto (si falta o no enruta)",
                                       IND,
                                       index=IND.index("Other") if "Other" in IND else 0)
            st.caption("Se agregan los leads con lo disponible; solo se omiten filas "
                       "totalmente vacías y duplicados (LinkedIn / email / nombre+empresa).")

            if st.button("Importar al Excel principal", type="primary"):
                if not mapping:
                    st.error("Mapea al menos una columna para poder importar.")
                else:
                    if not (mapping.get("LinkedIn Profile") or mapping.get("Email/Gmail")
                            or mapping.get("Full Name") or mapping.get("First Name")
                            or mapping.get("Company Name") or mapping.get("Phone")):
                        st.warning("Sin columna de identidad: los leads se agregarán pero "
                                   "no podrán deduplicarse ni seguirse individualmente.")
                    added, dups, per = crm.import_leads(
                        df.drop(columns=["__archivo__"], errors="ignore"),
                        mapping, route_by_industry=route, default_industry=default_ind)
                    if persist():
                        st.session_state.pop("pending_import", None)
                        st.success(f"✅ {added} leads importados · {dups} duplicados omitidos.")
                        if per:
                            st.table(pd.DataFrame(
                                [{"Industria": k, "Agregados": v} for k, v in per.items()]))

# =========================================================================== #
# 3) FOLLOW-UPS
# =========================================================================== #
elif page == "🔁 Follow-ups":
    st.header("🔁 Follow-ups")
    industry = st.selectbox("Industria", IND)
    channel = st.selectbox("Canal", ACTIVITY_CHANNELS)
    seq_channel = "Cold Call" if channel in ("Cold Call", "Llamada") else \
                  ("Email" if channel == "Email" else "LinkedIn")
    today = dt.date.today()

    # --- Vista general con todas las columnas + filtros por fase/estado ---
    st.subheader("Vista general")
    rows = []
    for lead in crm.all_leads(industry):
        stt = crm.state.get(lead.key)
        steps = CHANNEL_STEPS[seq_channel]
        current = "—"
        for s in steps:
            if _is_marked(lead.stage_values.get(s)):
                current = s
        nxt_step, nxt_date = "—", "—"
        if stt.scheduled:
            k = sorted(stt.scheduled.items(), key=lambda kv: kv[1])[0]
            nxt_step = k[0]
            nxt_date = k[1][:10]
        last_msg = "—"
        for e in reversed(stt.log):
            if e.get("action") == "sent":
                last_msg = (e.get("message") or "—")[:60]
                break
        last_contact = crm._cell(crm.wb[lead.sheet], crm.maps[lead.sheet],
                                 lead.row, "Last Contact")
        rows.append({
            "Nombre": lead.full_name, "Empresa": lead.company, "Puesto": lead.job_title,
            "Seniority": lead.seniority_level or "—", "Industria": lead.industry,
            "Follow up actual": current,
            "Último contacto": str(last_contact or "—"),
            "Canal último contacto": crm.lead_channel(lead, stt),
            "Mensaje usado": last_msg,
            "Siguiente follow-up": f"{nxt_step} ({nxt_date})" if nxt_step != "—" else
                                   crm.next_action(lead, stt),
            "Estado": crm.lead_status(lead, stt),
        })
    fdf = pd.DataFrame(rows)

    # Filtros por fase / estado
    phase_filters = ["Follow Up 1 pendiente", "Follow Up 1 enviado",
                     "Follow Up 2 pendiente", "Follow Up 2 enviado",
                     "Follow Up 3 pendiente", "Follow Up 3 enviado",
                     "respondieron", "prospectar después"]
    fp = st.multiselect("Ver", phase_filters)

    def _match(row):
        if not fp:
            return True
        st_ = row["Estado"]; cur = row["Follow up actual"]
        for f in fp:
            if f == "respondieron" and st_ == "respondió":
                return True
            if f == "prospectar después" and st_ == "prospectar después":
                return True
            if f.startswith("Follow Up"):
                n = f.split(" ")[2]
                enviado = cur == f"Follow Up {n}"
                if "enviado" in f and enviado:
                    return True
                if "pendiente" in f and not enviado and st_ == "pendiente":
                    return True
        return False

    if not fdf.empty:
        view = fdf[fdf.apply(_match, axis=1)] if fp else fdf
        st.dataframe(view, use_container_width=True, hide_index=True)
        st.caption(f"{len(view)} de {len(fdf)} leads.")
    else:
        st.info("Aún no hay leads en esta industria.")

    # --- Editor por lead: marcar enviado / cambiar canal / estado / nota ---
    st.divider()
    st.subheader("Actualizar un lead desde aquí")
    leads = crm.all_leads(industry)
    if leads:
        labels = [f"{l.full_name} · {l.company} · [{crm.lead_stage(l)}]" for l in leads]
        si = st.selectbox("Lead", range(len(leads)), format_func=lambda i: labels[i],
                          key="fu_lead")
        lead = leads[si]
        e1, e2 = st.columns(2)
        upd_channel = e1.selectbox("Canal", ACTIVITY_CHANNELS, key="fu_ch")
        # próximo paso a marcar en el canal de secuencia elegido
        seq2 = "Cold Call" if upd_channel in ("Cold Call", "Llamada") else \
               ("Email" if upd_channel == "Email" else "LinkedIn")
        nxt = None
        for s in CHANNEL_STEPS[seq2]:
            if not _is_marked(lead.stage_values.get(s)):
                nxt = s
                break
        e2.write(f"Próximo paso a marcar: **{nxt or '— (secuencia completa)'}**")
        note = st.text_input("Nota / mensaje usado", key="fu_note")
        b1, b2 = st.columns(2)
        if b1.button("Marcar follow-up como ENVIADO", type="primary") and nxt:
            crm.mark_sent(lead, upd_channel, nxt, message=note or "")
            idx = CHANNEL_STEPS[seq2].index(nxt)
            crm.schedule_next(lead, seq2, idx, gap_days=3)
            crm.add_notification(crm.read_lead(lead.sheet, lead.row),
                                 "Follow-up enviado", f"{nxt} por {upd_channel}", upd_channel)
            if persist():
                st.success(f"{nxt} marcado como enviado por {upd_channel}.")
        new_state = b2.selectbox("Cambiar estado a", ["—"] + MANUAL_STATES, key="fu_state")
        if st.button("Aplicar cambio de estado") and new_state != "—":
            crm.set_manual_state(lead, new_state, channel=upd_channel, note=note or None)
            if persist():
                st.success(f"Estado → {new_state}.")

        # --- Reprogramar follow-up (fecha/hora/canal/motivo) ---
        render_reschedule(lead, "fu_resch")

        # --- Formulario de Cold Call (registro manual de llamada) ---
        if channel in ("Cold Call", "Llamada"):
            with st.expander("📞 Registrar llamada (Cold Call)", expanded=True):
                c1, c2 = st.columns(2)
                call_date = c1.date_input("Fecha de llamada", dt.date.today(),
                                          key="cc_date")
                result = c2.selectbox("Resultado", COLD_CALL_RESULTS, key="cc_res")
                cc_note = st.text_input("Nota", key="cc_note")
                c3, c4 = st.columns(2)
                nxt_action = c3.text_input("Próxima acción", key="cc_next")
                nxt_date = c4.date_input("Próxima fecha de contacto",
                                         dt.date.today() + dt.timedelta(days=7),
                                         key="cc_nd")
                if st.button("Guardar llamada", type="primary", key="cc_go"):
                    crm.register_cold_call(
                        lead, call_date=call_date.isoformat(), result=result,
                        note=cc_note, next_action=nxt_action,
                        next_date=nxt_date.isoformat())
                    if persist():
                        st.success(f"📞 Llamada registrada: {result}. Actualizado "
                                   "en tabla, Activity_Log y Notifications.")

    # --- Aprobación por lote de follow-ups programados ---
    st.divider()
    st.subheader("Aprobar y marcar enviados (por lote)")
    due = []
    for lead in crm.all_leads(industry):
        stt = crm.state.get(lead.key)
        for step, when_iso in stt.scheduled.items():
            if step not in CHANNEL_STEPS[seq_channel]:
                continue
            if crm._is_blocked(lead, stt):
                continue
            when = dt.datetime.fromisoformat(when_iso)
            due.append((lead, step, when, when.date() <= today))

    if not due:
        st.info("No hay follow-ups programados para este canal.")
    else:
        st.write(f"{len(due)} programados ({sum(d[3] for d in due)} vencidos):")
        tmpl = st.text_area("Plantilla follow-up", "Hola {first}, retomo mi mensaje anterior.")
        chosen = {}
        for i, (lead, step, when, overdue) in enumerate(due):
            flag = "🔴 vencido" if overdue else f"🗓 {when:%Y-%m-%d}"
            chosen[i] = st.checkbox(
                f"{flag} · {step} · {lead.full_name} · {lead.company}", key=f"fu{i}")
        if st.button("Aprobar lote y marcar ENVIADOS", type="primary"):
            sent = 0
            for i, (lead, step, when, _o) in enumerate(due):
                if chosen.get(i):
                    first = (lead.full_name or "").split(" ")[0]
                    msg = tmpl.format(first=first, company=lead.company)
                    crm.mark_sent(lead, channel, step, message=msg)
                    idx = CHANNEL_STEPS[seq_channel].index(step)
                    crm.schedule_next(lead, seq_channel, idx, gap_days=3)
                    stt = crm.state.get(lead.key); stt.scheduled.pop(step, None)
                    crm.state.put(lead.key, stt)
                    sent += 1
            if persist():
                st.success(f"{sent} follow-ups marcados como enviados.")

    st.divider()
    st.subheader("Sugerencias de escalamiento de canal")
    any_s = False
    for lead in crm.all_leads(industry):
        s = crm.escalation_suggestion(lead)
        if s:
            any_s = True
            st.write(f"• **{lead.full_name}** ({lead.company}): {s}")
    if not any_s:
        st.caption("Sin sugerencias de escalamiento por ahora.")

# =========================================================================== #
# 3) RESPUESTAS
# =========================================================================== #
elif page == "📥 Respuestas":
    st.header("📥 Registrar respuesta de un lead")

    # --- Buscador / autocompletado por nombre / email / LinkedIn / empresa ---
    term = st.text_input("Buscar lead (nombre, email, LinkedIn o empresa)")
    if term:
        t = _norm(term)
        matches = []
        for sheet in crm.maps:
            for l in crm.all_leads(sheet):
                blob = _norm(" ".join([l.full_name, l.email, l.linkedin, l.company]))
                if t in blob:
                    matches.append(l)
                    if len(matches) >= 50:
                        break
            if len(matches) >= 50:
                break
        if not matches:
            st.warning("Sin coincidencias.")
        else:
            labels = [f"{l.full_name or '(sin nombre)'} · {l.company or '—'} · "
                      f"{l.email or l.linkedin or '—'}" for l in matches]
            sel = st.selectbox("Resultados", range(len(matches)),
                               format_func=lambda i: labels[i])
            lead = matches[sel]
            st.session_state.resp_lead = (lead.sheet, lead.row)

    # --- Autocompletado de datos del lead seleccionado ---
    if "resp_lead" in st.session_state:
        sheet, row = st.session_state.resp_lead
        lead = crm.read_lead(sheet, row)
        st_state = crm.state.get(lead.key)
        fu_actual = crm.lead_stage(lead, st_state)
        st.subheader("Lead seleccionado")
        st.dataframe(pd.DataFrame([{
            "Nombre": lead.full_name, "Empresa": lead.company, "Puesto": lead.job_title,
            "Industria": lead.industry, "Seniority": lead.seniority_level,
            "Email": lead.email, "LinkedIn": lead.linkedin, "Follow-up actual": fu_actual,
        }]).T.rename(columns={0: "Valor"}), use_container_width=True)

        st.subheader("Captura de la respuesta")
        message = st.text_area("Mensaje recibido")
        c1, c2 = st.columns(2)
        channel = c1.selectbox("Canal de respuesta", ACTIVITY_CHANNELS)
        result = c2.selectbox("Resultado", RESPONSE_RESULTS)
        recontact = None
        if result == "Prospectar después":
            recontact = st.date_input(
                "Fecha de recontacto (la decides tú)",
                dt.date.today() + dt.timedelta(days=30)).isoformat()
        note = st.text_input("Nota adicional")

        if st.button("Guardar respuesta", type="primary"):
            crm.apply_response_result(lead, result, channel=channel, message=message,
                                      note=note or None, recontact_date=recontact)
            if persist():
                st.success(f"✅ Respuesta guardada · resultado: {result} · "
                           "actualizado en la tabla principal, dashboard y Activity_Log.")
                del st.session_state.resp_lead

# =========================================================================== #
# NOTIFICATIONS
# =========================================================================== #
elif page == "🔔 Notifications":
    st.header("🔔 Notifications")

    tabA, tabB, tabC, tabG = st.tabs(["📝 Registrar respuesta (manual)",
                                      "🔔 Alertas", "🔌 Integraciones (Modo B)",
                                      "📧 Gmail"])

    # ---- Modo A: registro manual de respuesta ---- #
    with tabA:
        st.caption("Registro manual y seguro. Busca el lead, pega el mensaje recibido "
                   "y clasifícalo. No se conecta a ningún canal automáticamente.")
        term = st.text_input("Buscar lead (nombre, email, LinkedIn o empresa)",
                             key="noti_search")
        lead = None
        if term:
            t = _norm(term)
            matches = []
            for sheet in crm.maps:
                for l in crm.all_leads(sheet):
                    blob = _norm(" ".join([l.full_name, l.email, l.linkedin, l.company]))
                    if t in blob:
                        matches.append(l)
                        if len(matches) >= 50: break
                if len(matches) >= 50: break
            if matches:
                labels = [f"{l.full_name or '(sin nombre)'} · {l.company or '—'} · "
                          f"{l.email or l.linkedin or '—'}" for l in matches]
                si = st.selectbox("Resultados", range(len(matches)),
                                  format_func=lambda i: labels[i], key="noti_pick")
                lead = matches[si]
            else:
                st.warning("Sin coincidencias.")

        if lead is not None:
            stt = crm.state.get(lead.key)
            st.info(f"**{lead.full_name}** · {lead.company} · {lead.job_title or '—'} · "
                    f"{lead.industry} · {lead.seniority_level or '—'}  \n"
                    f"Follow-up actual: **{crm.lead_stage(lead, stt)}**")
            msg = st.text_area("Mensaje recibido (pega aquí)")
            r1, r2 = st.columns(2)
            channel = r1.selectbox("Canal", PROSPECT_CHANNELS + ["Otro"])
            rtype = r2.selectbox("Tipo de respuesta", RESPONSE_CLASSES + [
                "Pidió cotización", "Reunión agendada", "Won", "Lost",
                "Prospectar después", "Blacklist"])
            result = st.selectbox("Resultado", RESPONSE_RESULTS)
            recontact = None
            if result == "Prospectar después":
                recontact = st.date_input("Fecha próximo contacto",
                    dt.date.today() + dt.timedelta(days=30)).isoformat()
            note = st.text_input("Nota adicional")
            if st.button("Guardar respuesta", type="primary"):
                crm.apply_response_result(lead, result, channel=channel, message=msg,
                                          note=note or None, recontact_date=recontact)
                crm.add_notification(crm.read_lead(lead.sheet, lead.row),
                                     f"Respuesta: {rtype}", (msg or result)[:120], channel)
                if persist():
                    st.success(f"✅ Guardado. Actualizado estado, Current Stage, "
                               "Activity_Log, Notifications y dashboard.")

            render_reschedule(lead, "noti_resch")

    # ---- Alertas (lista de notificaciones) ---- #
    with tabB:
        b1, b2, b3 = st.columns(3)
        if b1.button("🔄 Follow-ups vencidos"):
            n = crm.scan_due_followups()
            if persist():
                st.success(f"{n} alerta(s) creada(s).")
        if b2.button("⏱ Leads sin respuesta (config)"):
            n = crm.scan_no_response()
            if persist():
                st.success(f"{n} alerta(s) creada(s) según Notification_Settings.")
        if b3.button("📣 Campañas listas"):
            n = crm.scan_campaign_reminders()
            if persist():
                st.success(f"{n} recordatorio(s) de campaña creado(s).")
        notis = crm.read_notifications()
        if not notis:
            st.info("No hay notificaciones todavía.")
        else:
            pend = [n for n in notis if str(n.get("Visto")) != "Vista"]
            st.subheader(f"Pendientes: {len(pend)} · Total: {len(notis)}")
            only_pend = st.checkbox("Ver solo pendientes", value=True)
            shown = pend if only_pend else notis
            for n in reversed(shown):
                icon = "🟡" if n.get("Visto") != "Vista" else "✅"
                with st.expander(f"{icon} {n.get('Fecha')} · {n.get('Lead')} · "
                                 f"{n.get('Tipo de evento')}"):
                    st.write(f"**Empresa:** {n.get('Empresa') or '—'}  ·  "
                             f"**Canal:** {n.get('Canal') or '—'}  ·  "
                             f"**Estado:** {n.get('Estado actual') or '—'}")
                    st.write(f"**Resumen:** {n.get('Mensaje/Resumen') or '—'}")
                    if n.get("Visto") != "Vista":
                        if st.button("Marcar como vista", key=f"seen{n['_row']}"):
                            crm.set_notification_seen(n["_row"], True)
                            if persist(): st.rerun()
                    else:
                        if st.button("Marcar como pendiente", key=f"pend{n['_row']}"):
                            crm.set_notification_seen(n["_row"], False)
                            if persist(): st.rerun()
            st.divider()
            st.dataframe(pd.DataFrame([{k: v for k, v in n.items() if k != "_row"}
                                       for n in notis]),
                         use_container_width=True, hide_index=True)

    # ---- Modo B: estado de integraciones (sin credenciales) ---- #
    with tabC:
        st.caption("Arquitectura preparada para integraciones futuras con APIs oficiales. "
                   "No hay credenciales en el código, no se envía ni se lee nada "
                   "automáticamente, y no se hace scraping.")
        for s in integrations.channel_statuses():
            badge = "🟢 API lista" if s.configured else "⚪ Manual"
            st.markdown(f"**{s.channel}** — {badge}")
            st.caption(s.note)
        st.info("LinkedIn: solo registro manual o importación CSV (política de LinkedIn). "
                "Gmail: futura vía Gmail API/OAuth. WhatsApp: futura vía WhatsApp Business "
                "API. Las credenciales, cuando existan, vendrán de variables de entorno.")

    # ---- Gmail (API oficial vía OAuth; envío SIEMPRE manual) ---- #
    with tabG:
        st.caption("Gmail vía API oficial con OAuth 2.0. Credenciales en st.secrets "
                   "(nunca en el código ni en GitHub). Lectura de correos recientes y "
                   "respuesta manual: la app no envía nada sin que tú presiones el botón.")

        if not gmail_service.libs_available():
            st.warning(gmail_service.libs_error())
        else:
            cfg, err = gmail_service.read_google_secrets(st)
            if err:
                st.error(err)
                st.code('# .streamlit/secrets.toml (NO subir a GitHub)\n'
                        '[google]\nclient_id = "TU_CLIENT_ID"\n'
                        'client_secret = "TU_CLIENT_SECRET"\n'
                        'redirect_uri = "https://TU-APP.streamlit.app"',
                        language="toml")
            else:
                # --- Conexión / OAuth (PKCE con code_verifier persistido) ---
                creds = gmail_oauth_panel(cfg)
                if creds is not None:
                    ok = True
                    try:
                        me = gmail_service.whoami(creds)
                        st.success(f"✅ Conectado como **{me}**")
                    except Exception as e:
                        st.error(f"Sesión inválida ({e}). Vuelve a conectar.")
                        st.session_state.pop("gmail_creds", None)
                        ok = False
                    if ok and st.button("Desconectar Gmail"):
                        st.session_state.pop("gmail_creds", None)
                        st.rerun()

                    if ok:
                        st.divider()
                        # --- Lectura + detección de respuestas de leads ---
                        g1, g2 = st.columns(2)
                        days = g1.number_input("Correos de los últimos (días)", 1, 30, 7)
                        maxr = g2.number_input("Máximo de correos", 5, 100, 25)
                        if st.button("📥 Leer correos recientes", type="primary"):
                            try:
                                msgs = gmail_service.list_recent_messages(
                                    creds, days=days, max_results=maxr)
                                st.session_state["gmail_msgs"] = msgs
                                st.session_state["gmail_matches"] = \
                                    gmail_service.match_messages_to_leads(msgs, crm)
                            except Exception as e:
                                st.error(f"No pude leer el buzón: {e}")

                        msgs = st.session_state.get("gmail_msgs", [])
                        matches = st.session_state.get("gmail_matches", [])
                        if msgs:
                            st.write(f"**{len(msgs)}** correos recientes · "
                                     f"**{len(matches)}** son de leads del CRM.")

                            if matches:
                                st.subheader("📬 Respuestas de leads detectadas")
                                for j, (msg, lead) in enumerate(matches):
                                    with st.expander(
                                            f"✉️ {lead.full_name} ({lead.company}) · "
                                            f"{msg['subject'][:60]}"):
                                        st.caption(f"{msg['date']} · {msg['from_email']}")
                                        st.write(msg["snippet"])
                                        if st.button("🔔 Registrar como respuesta",
                                                     key=f"gm_reg{j}"):
                                            crm.register_response(
                                                lead, "Email", message=msg["snippet"])
                                            crm.add_notification(
                                                crm.read_lead(lead.sheet, lead.row),
                                                "Respuesta detectada (Gmail)",
                                                msg["snippet"][:120], "Email")
                                            if persist():
                                                st.success(
                                                    "Registrada: follow-ups pausados, "
                                                    "notificación creada, Activity_Log "
                                                    "actualizado.")

                                        # --- Acciones rápidas sobre la respuesta ---
                                        act = st.selectbox(
                                            "Acción", ["—", "Marcar Won", "Marcar Lost",
                                                       "Marcar RFQ", "Contactar después",
                                                       "Blacklist"], key=f"gm_act{j}")
                                        gm_note = st.text_input("Nota", key=f"gm_nt{j}")
                                        gm_rd = None
                                        if act == "Contactar después":
                                            gm_rd = st.date_input(
                                                "Fecha de recontacto",
                                                dt.date.today() + dt.timedelta(days=30),
                                                key=f"gm_rd{j}").isoformat()
                                        if st.button("Aplicar acción",
                                                     key=f"gm_do{j}") and act != "—":
                                            if act == "Marcar Won":
                                                crm.set_outcome(lead, "Won",
                                                                note=gm_note or None)
                                            elif act == "Marcar Lost":
                                                crm.set_outcome(lead, "Lost",
                                                                note=gm_note or None)
                                            elif act == "Marcar RFQ":
                                                crm.mark_stage(lead, "RFQ")
                                            elif act == "Contactar después":
                                                crm.set_outcome(
                                                    lead, "Prospectar Después",
                                                    recontact_date=gm_rd,
                                                    note=gm_note or None)
                                            elif act == "Blacklist":
                                                crm.set_outcome(lead, "Blacklist",
                                                                note=gm_note or None)
                                            crm.add_notification(
                                                crm.read_lead(lead.sheet, lead.row),
                                                f"Gmail: {act}",
                                                gm_note or msg["snippet"][:80], "Email")
                                            if persist():
                                                st.success(f"✅ {act} aplicado.")

                                        # --- Reprogramar follow-up desde Gmail ---
                                        render_reschedule(lead, f"gm_resch{j}")

                                        reply = st.text_area(
                                            "Responder (envío manual)",
                                            key=f"gm_rep{j}", height=80)
                                        if st.button("📤 Enviar respuesta",
                                                     key=f"gm_send{j}"):
                                            if not reply.strip():
                                                st.warning("Escribe la respuesta primero.")
                                            else:
                                                try:
                                                    gmail_service.send_reply(
                                                        creds, msg["from_email"],
                                                        msg["subject"], reply,
                                                        thread_id=msg["threadId"],
                                                        in_reply_to=msg["message_id_hdr"])
                                                    crm.append_activity(
                                                        lead, "Email", "-", reply,
                                                        "Respuesta enviada (Gmail)")
                                                    crm.add_notification(
                                                        crm.read_lead(lead.sheet, lead.row),
                                                        "Respuesta enviada (Gmail)",
                                                        reply[:120], "Email")
                                                    if persist():
                                                        st.success("Enviado y registrado "
                                                                   "en Activity_Log.")
                                                except Exception as e:
                                                    st.error(f"No pude enviar: {e}")

                            others = [m for m in msgs if m["from_email"] not in
                                      {mm["from_email"] for mm, _ in matches}]
                            if others:
                                with st.expander(
                                        f"Otros correos recientes ({len(others)})"):
                                    st.dataframe(pd.DataFrame(
                                        [{"De": m["from_name"], "Email": m["from_email"],
                                          "Asunto": m["subject"], "Fecha": m["date"]}
                                         for m in others]),
                                        use_container_width=True, hide_index=True)

# =========================================================================== #
# 4) ESTADOS
# =========================================================================== #
elif page == "🏷️ Estados":
    st.header("🏷️ Estados")

    # --- Tabla resumen agrupada con filtros ---
    st.subheader("Resumen agrupado")

    @st.cache_data(show_spinner="Calculando estados…")
    def _states_table(mtime):
        rows = []
        for sheet in crm.maps:
            for l in crm.all_leads(sheet):
                stt = crm.state.get(l.key)
                status = crm.lead_status(l, stt)
                nxt = ""
                if stt.scheduled:
                    nxt = min(stt.scheduled.values())[:10]
                rows.append({
                    "Industria": l.industry, "Seniority": l.seniority_level or "—",
                    "Paso actual": crm.lead_stage(l, stt),
                    "Canal actual": crm.lead_channel(l, stt),
                    "Estado": status, "Próximo follow-up": nxt,
                    "_contactado": status not in ("pendiente",),
                    "_pendiente": status == "pendiente",
                    "_respondio": status == "respondió",
                    "_won": status == "won", "_lost": status == "lost",
                    "_blacklist": status == "blacklist",
                })
        return pd.DataFrame(rows)

    df = _states_table(Path(xlsx_path).stat().st_mtime)

    f1, f2, f3 = st.columns(3)
    fi = f1.multiselect("Industria", sorted(df["Industria"].unique()))
    fs = f2.multiselect("Seniority", sorted(df["Seniority"].unique()))
    fc = f3.multiselect("Canal actual", sorted(df["Canal actual"].unique()))
    f4c, f5c, f6c = st.columns(3)
    fp = f4c.multiselect("Etapa/fase (paso actual)", sorted(df["Paso actual"].unique()))
    fe = f5c.multiselect("Estado", sorted(df["Estado"].unique()))
    finals = ["won", "lost", "blacklist", "prospectar después"]
    ff = f6c.multiselect("Resultado final", finals)
    fdf = df.copy()
    if fi: fdf = fdf[fdf["Industria"].isin(fi)]
    if fs: fdf = fdf[fdf["Seniority"].isin(fs)]
    if fc: fdf = fdf[fdf["Canal actual"].isin(fc)]
    if fp: fdf = fdf[fdf["Paso actual"].isin(fp)]
    if fe: fdf = fdf[fdf["Estado"].isin(fe)]
    if ff: fdf = fdf[fdf["Estado"].isin(ff)]

    grouped = (fdf.groupby(["Industria", "Seniority", "Paso actual", "Canal actual"])
               .agg(total_leads=("Estado", "size"),
                    contactados=("_contactado", "sum"),
                    pendientes=("_pendiente", "sum"),
                    respondieron=("_respondio", "sum"),
                    won=("_won", "sum"), lost=("_lost", "sum"),
                    blacklist=("_blacklist", "sum"),
                    proximo_follow_up=("Próximo follow-up",
                                       lambda s: min([x for x in s if x], default="")))
               .reset_index())
    st.dataframe(grouped, use_container_width=True, hide_index=True)
    st.caption(f"{len(fdf)} leads en la vista · {len(grouped)} grupos")

    st.divider()
    # --- Editor: cambiar manualmente a cualquiera de los 12 estados ---
    st.subheader("Cambiar estado de un lead")
    industry = st.selectbox("Industria", IND, key="est_ind")
    leads = crm.all_leads(industry)
    labels = [f"{l.full_name} · {l.company} · [{crm.lead_stage(l)}]" for l in leads]
    sel = st.selectbox("Lead", range(len(leads)), format_func=lambda i: labels[i])
    lead = leads[sel]

    st.write(f"**{lead.full_name}** — {lead.job_title} @ {lead.company} · "
             f"estado actual: **{crm.lead_stage(lead)}**")
    new_state = st.selectbox("Current Stage / Estado", MANUAL_STATES)
    ch = st.selectbox("Canal", ACTIVITY_CHANNELS)
    recontact = None
    if new_state == "Prospectar después":
        recontact = st.date_input("Fecha de recontacto (la decides tú)",
                                  dt.date.today() + dt.timedelta(days=30)).isoformat()
    note = st.text_input("Nota")

    if st.button("Aplicar estado", type="primary"):
        crm.set_manual_state(lead, new_state, channel=ch, note=note or None,
                             recontact_date=recontact)
        if persist():
            st.success(f"✅ Estado → {new_state}. Actualizado en tabla, Activity_Log, "
                       "Notifications y dashboard.")
            _states_table.clear()

    # --- Campos editables adicionales ---
    with st.expander("✏️ Editar más campos (warmth, prioridad, canal, próxima acción…)"):
        e1, e2, e3 = st.columns(3)
        warmth = e1.selectbox("Lead Warmth Level", ["—"] + WARMTH_LEVELS)
        priority = e2.selectbox("Priority", ["—", "Alta", "Media", "Baja"])
        rec_ch = e3.selectbox("Recommended Channel", ["—"] + PROSPECT_CHANNELS)
        e4, e5 = st.columns(2)
        fu_step = e4.selectbox("Follow Up Step", ["—"] + CHANNEL_STEPS["LinkedIn"])
        next_date = e5.date_input("Fecha de próximo contacto",
                                  dt.date.today() + dt.timedelta(days=3))
        next_act = st.text_input("Próxima acción")
        more_notes = st.text_input("Notas", key="est_notes")
        if st.button("Guardar campos"):
            crm.edit_lead_fields(
                lead,
                warmth=None if warmth == "—" else warmth,
                priority=None if priority == "—" else priority,
                channel=None if rec_ch == "—" else rec_ch,
                fu_step=None if fu_step == "—" else fu_step,
                next_action=next_act or None,
                next_date=next_date.isoformat(),
                notes=more_notes or None)
            if persist():
                st.success("Campos guardados y registrados en Activity_Log.")
                _states_table.clear()

    # --- Reprogramar follow-up desde Estados ---
    render_reschedule(lead, "est_resch")

    st.divider()
    st.subheader("Historial de actividad del lead")
    log = crm.state.get(lead.key).log
    if log:
        st.dataframe(pd.DataFrame(log), use_container_width=True, hide_index=True)
    else:
        st.caption("Sin actividad registrada todavía.")

    st.divider()
    st.subheader("Historial de reprogramaciones (Follow_Up_History)")
    if "Follow_Up_History" in crm.wb.sheetnames:
        hws = crm.wb["Follow_Up_History"]
        hrows = [[hws.cell(r, c).value for c in range(1, 13)]
                 for r in range(2, hws.max_row + 1)
                 if hws.cell(r, 2).value == lead.key]
        if hrows:
            from crm_core import FU_HISTORY_HEADERS
            st.dataframe(pd.DataFrame(hrows, columns=FU_HISTORY_HEADERS),
                         use_container_width=True, hide_index=True)
        else:
            st.caption("Este lead no tiene reprogramaciones registradas.")
    else:
        st.caption("Aún no hay historial de reprogramaciones.")

# =========================================================================== #
# CONFIGURACIÓN DE ESTADOS (colores)
# =========================================================================== #
elif page == "🎨 Configuración de Estados":
    st.header("🎨 Configuración de Estados")
    st.caption("Define el color de cada estado, si se pinta la celda de estado o "
               "toda la fila, y crea estados personalizados. La configuración se "
               "guarda en la hoja State_Color_Config del Excel.")

    cfg = crm.read_state_colors()

    st.subheader("Estados disponibles")
    for state, c in cfg.items():
        with st.expander(f"{'🟢' if str(c.get('Active','Sí')).startswith('S') else '⚪'} "
                         f"{state} · #{c['Color']} · {c.get('Scope','row')}"):
            e1, e2, e3 = st.columns(3)
            color = e1.color_picker("Color", f"#{str(c['Color'])[:6]}",
                                    key=f"col_{state}")
            scope = e2.selectbox("Pintar", ["row", "cell"],
                                 index=0 if c.get("Scope", "row") == "row" else 1,
                                 format_func=lambda x: "Toda la fila" if x == "row"
                                 else "Solo celda de estado", key=f"sc_{state}")
            cat = e3.selectbox("Categoría", STATE_CATEGORIES,
                               index=STATE_CATEGORIES.index(c.get("Category", "activo"))
                               if c.get("Category") in STATE_CATEGORIES else 0,
                               key=f"cat_{state}")
            e4, e5, e6 = st.columns(3)
            action = e4.text_input("Acción sugerida", c.get("Suggested Action") or "",
                                   key=f"act_{state}")
            prio = e5.selectbox("Prioridad", ["Alta", "Media", "Baja"],
                                index=["Alta", "Media", "Baja"].index(
                                    c.get("Priority", "Media"))
                                if c.get("Priority") in ("Alta", "Media", "Baja") else 1,
                                key=f"pr_{state}")
            active = e6.checkbox("Activo", str(c.get("Active", "Sí")).startswith("S"),
                                 key=f"ac_{state}")
            if st.button("Guardar estado", key=f"sv_{state}"):
                crm.upsert_state_color(state, color, scope, cat, action, prio, active)
                if persist():
                    st.success(f"'{state}' guardado en State_Color_Config.")

    st.divider()
    st.subheader("➕ Crear estado personalizado")
    n1, n2, n3 = st.columns(3)
    new_name = n1.text_input("Nombre del estado (ej. Sample Requested, Hot Lead)")
    new_color = n2.color_picker("Color", "#FF6600")
    new_scope = n3.selectbox("Pintar", ["row", "cell"],
                             format_func=lambda x: "Toda la fila" if x == "row"
                             else "Solo celda de estado")
    n4, n5, n6 = st.columns(3)
    new_cat = n4.selectbox("Categoría", STATE_CATEGORIES)
    new_action = n5.text_input("Acción sugerida")
    new_prio = n6.selectbox("Prioridad", ["Alta", "Media", "Baja"], index=1)
    if st.button("Crear estado", type="primary"):
        if not new_name.strip():
            st.error("Ponle nombre al estado.")
        else:
            crm.upsert_state_color(new_name.strip(), new_color, new_scope, new_cat,
                                   new_action, new_prio, True)
            if persist():
                st.success(f"Estado '{new_name}' creado. Asígnalo a leads escribiendo "
                           "ese valor en Current Stage (sección Estados).")

    st.divider()
    st.subheader("🖌️ Aplicar colores al Excel")
    st.caption("Pinta filas/celdas con openpyxl según la configuración de arriba.")
    remove_cf = st.checkbox(
        "Reemplazar formato condicional heredado",
        help="El Excel trae reglas de color automáticas por Outcome Status (Won verde, "
             "Lost morado, etc.) que visualmente GANAN sobre el pintado directo. "
             "Márcalo para eliminarlas y que tu configuración sea la única fuente "
             "de color. Se hace backup antes de guardar.")
    if st.button("Aplicar colores ahora", type="primary"):
        n = crm.apply_state_colors(remove_legacy_cf=remove_cf)
        if persist():
            st.success(f"🖌️ {n} leads pintados según la configuración"
                       + (" (formato condicional heredado eliminado)." if remove_cf
                          else " (el formato condicional heredado sigue activo y "
                               "prevalece en Won/Lost/Blacklist/Prospectar)."))

# =========================================================================== #
# GMAIL CAMPAIGNS (login, sender, campañas, Follow Up 1, respuestas)
# =========================================================================== #
elif page == "📮 Gmail Campaigns":
    st.header("📮 Gmail Campaigns")
    crm.ensure_gmail_sheets()
    tC, tN, tP, tF, tR = st.tabs(["🔐 Conexión", "🆕 Crear campaña",
                                  "👀 Preview / cola", "🔁 Gmail Follow Up 1",
                                  "📥 Respuestas"])
    creds = st.session_state.get("gmail_creds")

    # ---------- Conexión ----------
    with tC:
        if not gmail_service.libs_available():
            st.warning(gmail_service.libs_error())
        else:
            cfg, err = gmail_service.read_google_secrets(st)
            if err:
                st.error(err)
            else:
                creds = gmail_oauth_panel(cfg)
            if creds is not None and not err:
                try:
                    me = gmail_service.whoami(creds)
                    crm.save_gmail_account(me)
                    persist()
                    accs = crm.read_gmail_accounts()
                    acc = next((a for a in accs
                                if str(a["Email"]).lower() == me.lower()), {})
                    c1, c2, c3 = st.columns(3)
                    c1.metric("Correo conectado", me)
                    c2.metric("Estado", acc.get("Status", "Conectada"))
                    c3.metric("Última sincronización",
                              str(acc.get("Last Sync", ""))[:16].replace("T", " "))
                    if st.button("Desconectar Gmail"):
                        st.session_state.pop("gmail_creds", None)
                        st.rerun()
                except Exception as e:
                    st.error(f"Sesión inválida ({e}). Vuelve a conectar.")
                    st.session_state.pop("gmail_creds", None)
        st.dataframe(pd.DataFrame(crm.read_gmail_accounts()),
                     use_container_width=True, hide_index=True)

    # ---------- Crear campaña ----------
    with tN:
        senders = [a["Email"] for a in crm.read_gmail_accounts()]
        if not senders:
            st.info("Conecta una cuenta Gmail primero (pestaña Conexión).")
        else:
            sender = st.selectbox("Correo emisor (sender)", senders)
            f1, f2, f3 = st.columns(3)
            f_ind = f1.multiselect("Industria", IND)
            f_sen = f2.text_input("Seniority contiene")
            f_pri = f3.multiselect("Prioridad", ["Alta", "Media", "Baja"])
            f4, f5, f6 = st.columns(3)
            f_comp = f4.text_input("Empresa contiene")
            f_loc = f5.text_input("País/ubicación contiene")
            f_warm = f6.multiselect("Lead warmth", WARMTH_LEVELS)
            f7, f8, f9 = st.columns(3)
            f_stage = f7.text_input("Current stage contiene")
            f_fustep = f8.selectbox("Follow up step (sin marcar aún)",
                                    ["Email 1"] + CHANNEL_STEPS["Email"][1:])
            n_leads = f9.number_input("Cantidad de leads", 1, 300, 30)
            st.caption("Se excluyen automáticamente: Blacklist, Lost, Do Not "
                       "Contact, Won, respondidos y leads sin email válido.")

            cname = st.text_input("Nombre de campaña",
                                  f"Gmail {dt.date.today():%b %d}")
            subj_t = st.text_input("Asunto", "Propuesta para {{company}}")
            body_t = st.text_area(
                "Mensaje base — variables: {{first_name}}, {{full_name}}, "
                "{{company}}, {{job_title}}, {{industry}}, {{seniority_level}}",
                "Hola {{first_name}},\n\nVi tu rol de {{job_title}} en "
                "{{company}}...", height=140)
            g1, g2, g3 = st.columns(3)
            start = g1.text_input("Inicio (YYYY-MM-DD HH:MM)",
                                  dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
            interval = g2.number_input("Intervalo (min)", 1, 1440, 45)
            cap = g3.number_input("Límite máximo por día", 1, 500, 50)
            g4, g5 = st.columns(2)
            need_confirm = g4.checkbox("Requiere confirmación manual antes de "
                                       "enviar", value=True)
            autos = g5.toggle("Auto-send activado", value=False)

            if st.button("Generar preview de campaña", type="primary"):
                picked = []
                for ind in (f_ind or IND):
                    for l in crm.all_leads(ind):
                        stt = crm.state.get(l.key)
                        if crm._is_blocked(l, stt) or not l.email:
                            continue
                        if _is_marked(l.stage_values.get(f_fustep)):
                            continue
                        if f_sen and _norm(f_sen) not in _norm(
                                l.seniority_level or ""):
                            continue
                        if f_pri and crm.lead_priority(l) not in f_pri:
                            continue
                        if f_comp and _norm(f_comp) not in _norm(l.company):
                            continue
                        loc = crm._cell(crm.wb[l.sheet], crm.maps[l.sheet],
                                        l.row, "Location") or ""
                        if f_loc and _norm(f_loc) not in _norm(str(loc)):
                            continue
                        if f_warm and crm.classify_warmth(l, stt) not in f_warm:
                            continue
                        cur = crm._cell(crm.wb[l.sheet], crm.maps[l.sheet],
                                        l.row, "Current Stage") or ""
                        if f_stage and _norm(f_stage) not in _norm(str(cur)):
                            continue
                        picked.append(l)
                        if len(picked) >= n_leads:
                            break
                    if len(picked) >= n_leads:
                        break
                st.session_state["gc_preview"] = dict(
                    leads=[(l.sheet, l.row) for l in picked], name=cname,
                    sender=sender, subj=subj_t, body=body_t, start=start,
                    interval=int(interval), cap=int(cap),
                    confirm=need_confirm, autosend=autos,
                    step=f_fustep)
                st.success(f"{len(picked)} leads en preview. "
                           "Ve a la pestaña 👀 Preview / cola.")

    # ---------- Preview / cola ----------
    with tP:
        pv = st.session_state.get("gc_preview")
        if not pv:
            st.info("Genera primero un preview en la pestaña Crear campaña.")
        else:
            leads = [crm.read_lead(s, r) for s, r in pv["leads"]]
            try:
                t0 = dt.datetime.strptime(pv["start"], "%Y-%m-%d %H:%M")
            except ValueError:
                t0 = dt.datetime.now()
            st.write(f"**{pv['name']}** · sender: {pv['sender']} · "
                     f"{len(leads)} leads · auto-send: "
                     f"{'ON' if pv['autosend'] else 'OFF'} · confirmación: "
                     f"{'sí' if pv['confirm'] else 'no'}")
            rows, per_lead = [], []
            for i, l in enumerate(leads):
                stt = crm.state.get(l.key)
                subj = crm.personalize(pv["subj"], l)
                body = crm.personalize(pv["body"], l)
                rows.append({"Nombre": l.full_name, "Email": l.email,
                             "Empresa": l.company, "Puesto": l.job_title,
                             "Industria": l.industry,
                             "Seniority": l.seniority_level or "—",
                             "Prioridad": crm.lead_priority(l),
                             "Estado": crm.lead_status(l, stt),
                             "Follow up": crm.lead_stage(l, stt),
                             "Mensaje": body[:60] + "…"})
                per_lead.append((l, subj, body,
                                 t0 + dt.timedelta(minutes=i * pv["interval"])))
            st.dataframe(pd.DataFrame(rows), use_container_width=True,
                         hide_index=True)

            st.subheader("Editar / excluir por lead")
            final = []
            for i, (l, subj, body, when) in enumerate(per_lead):
                with st.expander(f"{i+1}. {l.full_name} · {l.email} · "
                                 f"⏱ {when:%m-%d %H:%M}"):
                    subj = st.text_input("Asunto", subj, key=f"gc_s{i}")
                    body = st.text_area("Mensaje", body, key=f"gc_b{i}",
                                        height=100)
                    inc = st.checkbox("Incluir", True, key=f"gc_i{i}")
                    snd = st.button("📤 Enviar SOLO este ahora",
                                    key=f"gc_send1_{i}",
                                    disabled=not creds)
                    if snd:
                        try:
                            gmail_service.send_email(creds, l.email, subj, body)
                            crm.mark_sent(l, "Email", pv["step"], message=body)
                            crm._set(l, "Current Stage", "Email sent")
                            if persist():
                                st.success("Enviado y registrado.")
                        except Exception as e:
                            st.error(f"No pude enviar: {e}")
                    if inc:
                        final.append((l, subj, body, when))

            st.write(f"Incluidos: **{len(final)}**")
            a1, a2, a3 = st.columns(3)
            def _register(status):
                cid = crm.create_gmail_campaign(
                    pv["name"], pv["sender"], subject=pv["subj"],
                    total=len(final), status=status)
                for l, subj, body, when in final:
                    wiso = when.isoformat(timespec="minutes")
                    crm.add_gmail_campaign_lead(cid, l, subj, body,
                                                status=status, when_iso=wiso)
                    if status == "Scheduled":
                        crm.schedule_message(
                            l, "Email", body, wiso, follow_up_step=pv["step"],
                            status="Scheduled",
                            notes=f"{pv['name']}|SUBJ|{subj}")
                crm.set_kv(WORKFLOW_SHEET, "email_daily_cap", pv["cap"],
                           WORKFLOW_DEFAULTS)
                return cid
            if a1.button("💾 Guardar como borrador"):
                cid = _register("Draft")
                if persist():
                    st.success(f"Borrador {cid} guardado en Gmail_Campaigns.")
            if a2.button("📅 Programar campaña", type="primary"):
                cid = _register("Scheduled")
                if persist():
                    st.success(f"Campaña {cid} programada: {len(final)} correos "
                               "en cola (Scheduled_Messages).")
            confirm_ok = (not pv["confirm"]) or st.checkbox(
                "Confirmo el envío de los seleccionados", key="gc_confirm")
            if a3.button("🚀 Enviar seleccionados ahora",
                         disabled=not (creds and confirm_ok)):
                cid = crm.create_gmail_campaign(pv["name"], pv["sender"],
                                                subject=pv["subj"],
                                                total=len(final), status="Sent")
                ok = fail = 0
                for l, subj, body, when in final:
                    try:
                        gmail_service.send_email(creds, l.email, subj, body)
                        crm.mark_sent(l, "Email", pv["step"], message=body)
                        crm._set(l, "Current Stage", "Email sent")
                        crm.add_gmail_campaign_lead(cid, l, subj, body,
                                                    status="Sent")
                        ok += 1
                    except Exception as e:
                        fail += 1
                if persist():
                    st.success(f"✅ {ok} enviados · {fail} fallidos · "
                               f"campaña {cid} registrada.")

    # ---------- Gmail Follow Up 1 ----------
    with tF:
        wf = crm.workflow_config()
        gap = st.number_input("Tiempo configurado para follow up (horas)", 1, 720,
                              int(float(wf["fu1_to_fu2_hours"])))
        cands = crm.gmail_followup1_candidates(min_hours=gap)
        st.write(f"**{len(cands)}** leads listos para Gmail Follow Up 1 "
                 f"(≥{gap}h sin respuesta tras el email inicial).")
        if cands:
            st.dataframe(pd.DataFrame([{
                "Nombre": l.full_name, "Email": l.email, "Empresa": l.company,
                "Puesto": l.job_title, "Industria": l.industry,
                "Seniority": l.seniority_level or "—",
                "Prioridad": crm.lead_priority(l),
                "Email inicial": info["email_inicial"],
                "Horas sin respuesta": info["horas_sin_respuesta"],
                "Último mensaje": info["ultimo_mensaje"],
                "Asunto anterior": info["asunto_anterior"],
                "Próxima acción": info["proxima_accion"],
            } for l, info in cands]), use_container_width=True, hide_index=True)

            tmpl = st.text_area("Mensaje de Follow Up 1 (variables {{}})",
                                "Hola {{first_name}}, retomo mi correo anterior "
                                "sobre {{company}}.")
            subj_fu = st.text_input("Asunto", "Re: Propuesta para {{company}}")
            h1, h2 = st.columns(2)
            when_d = h1.date_input("Fecha", dt.date.today())
            when_t = h2.time_input("Hora", dt.time(9, 0))
            interval2 = st.number_input("Intervalo (min)", 1, 1440, 45,
                                        key="gfu_int")
            note = st.text_input("Nota")
            chosen = {}
            for i, (l, info) in enumerate(cands):
                chosen[i] = st.checkbox(f"{l.full_name} · {l.company}",
                                        key=f"gfu{i}")
            sel = [cands[i] for i in chosen if chosen[i]]
            b1, b2, b3 = st.columns(3)
            if b1.button("📅 Programar Follow Up 1", type="primary") and sel:
                t0 = dt.datetime.combine(when_d, when_t)
                for i, (l, info) in enumerate(sel):
                    w = (t0 + dt.timedelta(minutes=i * interval2)) \
                        .isoformat(timespec="minutes")
                    body = crm.personalize(tmpl, l)
                    sj = crm.personalize(subj_fu, l)
                    crm.schedule_message(l, "Email", body, w,
                                         follow_up_step="Email 2",
                                         status="Scheduled",
                                         notes=f"GmailFU1|SUBJ|{sj}")
                    crm.log_gmail_followup(l, "Email 2", sj, body, "Scheduled",
                                           w, note)
                if persist():
                    st.success(f"{len(sel)} follow-ups programados.")
            confirm_fu = st.checkbox("Confirmo el envío inmediato",
                                     key="gfu_confirm")
            if b2.button("🚀 Enviar ahora (auto-send)",
                         disabled=not (creds and confirm_fu)) and sel:
                ok = 0
                for l, info in sel:
                    body = crm.personalize(tmpl, l)
                    sj = crm.personalize(subj_fu, l)
                    try:
                        gmail_service.send_email(creds, l.email, sj, body)
                        crm.mark_sent(l, "Email", "Email 2", message=body)
                        crm._set(l, "Current Stage", "Follow Up sent")
                        crm.log_gmail_followup(l, "Email 2", sj, body, "Sent",
                                               note=note)
                        ok += 1
                    except Exception as e:
                        st.error(f"{l.email}: {e}")
                if persist():
                    st.success(f"{ok} follow-ups enviados y registrados.")
            if b3.button("🗓️ Reprogramar el primero seleccionado") and sel:
                st.session_state["gfu_resch_lead"] = (sel[0][0].sheet,
                                                      sel[0][0].row)
            if "gfu_resch_lead" in st.session_state:
                s_, r_ = st.session_state["gfu_resch_lead"]
                render_reschedule(crm.read_lead(s_, r_), "gfu_resch")

    # ---------- Respuestas ----------
    with tR:
        if st.button("🔍 Buscar respuestas recientes", type="primary",
                     disabled=not creds):
            try:
                msgs = gmail_service.list_recent_messages(creds, days=7,
                                                          max_results=50)
                st.session_state["gc_matches"] = \
                    gmail_service.match_messages_to_leads(msgs, crm)
                me = gmail_service.whoami(creds)
                crm.save_gmail_account(me)   # actualiza Last Sync
                persist()
            except Exception as e:
                st.error(f"No pude leer el buzón: {e}")
        matches = st.session_state.get("gc_matches", [])
        st.write(f"**{len(matches)}** respuestas de leads detectadas.")
        CLS = ["Interesado", "No interesado", "RFQ", "Reunión agendada", "Won",
               "Lost", "Prospectar después", "Blacklist", "Do Not Contact"]
        for j, (msg, lead) in enumerate(matches):
            with st.expander(f"✉️ {lead.full_name} ({lead.company}) · "
                             f"{msg['subject'][:60]}"):
                st.caption(f"{msg['date']} · {msg['from_email']}")
                st.write(msg["snippet"])
                if st.button("Marcar Respondió (detiene follow-ups)",
                             key=f"gc_r{j}"):
                    crm.register_response(lead, "Email", msg["snippet"])
                    if persist():
                        st.success("Respondió: follow-ups detenidos, "
                                   "notificación creada.")
                cls = st.selectbox("Clasificar", CLS, key=f"gc_c{j}")
                rd = None
                if cls == "Prospectar después":
                    rd = st.date_input("Recontacto",
                                       dt.date.today() + dt.timedelta(days=30),
                                       key=f"gc_rd{j}").isoformat()
                if st.button("Aplicar clasificación", key=f"gc_a{j}"):
                    if cls == "Do Not Contact":
                        crm.set_outcome(lead, "Do Not Contact")
                        crm.cancel_scheduled_for_lead(lead)
                    elif cls == "Won":
                        crm.set_outcome(lead, "Won")
                    elif cls == "Lost":
                        crm.set_outcome(lead, "Lost")
                    elif cls == "Blacklist":
                        crm.set_outcome(lead, "Blacklist")
                        crm.cancel_scheduled_for_lead(lead)
                    elif cls == "Prospectar después":
                        crm.set_outcome(lead, "Prospectar Después",
                                        recontact_date=rd)
                        crm.cancel_scheduled_for_lead(lead)
                    elif cls == "RFQ":
                        crm.register_response(lead, "Email", msg["snippet"])
                        crm.mark_stage(lead, "RFQ")
                    elif cls == "Reunión agendada":
                        crm.register_response(lead, "Email", msg["snippet"])
                        crm.mark_stage(lead, "Meeting")
                    else:
                        crm.register_response(lead, "Email", msg["snippet"])
                    crm.add_notification(crm.read_lead(lead.sheet, lead.row),
                                         f"Gmail: {cls}",
                                         msg["snippet"][:120], "Email")
                    if persist():
                        st.success(f"✅ {cls} aplicado.")

# =========================================================================== #
# EMAIL CAMPAIGNS (Gmail API, auto-send con opt-in explícito)
# =========================================================================== #
elif page == "📧 Email Campaigns":
    st.header("📧 Email Campaigns")
    st.caption("Campañas de correo vía Gmail API oficial (OAuth, st.secrets). "
               "El envío automático SOLO ocurre si activas 'Auto-send enabled' "
               "y con la app abierta; nunca a leads en Blacklist/Do Not Contact/Lost.")

    wf = crm.workflow_config()

    # --- Crear campaña por filtros ---
    st.subheader("Crear campaña")
    c1, c2, c3 = st.columns(3)
    industry = c1.selectbox("Industria", IND)
    n_leads = c2.number_input("Nº de leads", 1, 200, 30)
    warm = c3.multiselect("Lead warmth", WARMTH_LEVELS)
    c4, c5, c6 = st.columns(3)
    country = c4.text_input("País/ubicación contiene")
    company_f = c5.text_input("Empresa contiene")
    sen_f = c6.text_input("Seniority contiene")
    subject = st.text_input("Asunto", "Propuesta para {company}")
    body = st.text_area("Mensaje base (variables: {first}, {name}, {company}, "
                        "{title}, {industry})",
                        "Hola {first},\n\nVi tu rol de {title} en {company}...")
    c7, c8, c9 = st.columns(3)
    start = c7.text_input("Inicio (YYYY-MM-DD HH:MM)",
                          dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
    interval = c8.number_input("Intervalo entre correos (min)", 1, 1440,
                               int(float(wf.get("message_interval_min", 45))))
    daily_cap = c9.number_input("Máximo envíos por día", 1, 500,
                                int(float(wf.get("email_daily_cap", 50))))
    fu_gap = st.number_input("Follow-up automático después de (horas)", 1, 720,
                             int(float(wf.get("fu1_to_fu2_hours", 72))))

    if st.button("Crear cola de correos", type="primary"):
        try:
            t0 = dt.datetime.strptime(start, "%Y-%m-%d %H:%M")
        except ValueError:
            t0 = dt.datetime.now()
        picked = []
        for l in crm.all_leads(industry):
            stt = crm.state.get(l.key)
            if crm._is_blocked(l, stt) or not l.email:
                continue
            if warm and crm.classify_warmth(l, stt) not in warm:
                continue
            loc = crm._cell(crm.wb[l.sheet], crm.maps[l.sheet], l.row, "Location") or ""
            if country and _norm(country) not in _norm(str(loc)):
                continue
            if company_f and _norm(company_f) not in _norm(l.company):
                continue
            if sen_f and _norm(sen_f) not in _norm(l.seniority_level or ""):
                continue
            if _is_marked(l.stage_values.get("Email 1")):
                continue
            picked.append(l)
            if len(picked) >= n_leads:
                break
        for i, l in enumerate(picked):
            first = (l.full_name or "").split(" ")[0]
            vals = dict(first=first, name=l.full_name, company=l.company,
                        title=l.job_title, industry=l.industry)
            msg = body.format(**vals)
            subj = subject.format(**vals)
            when = (t0 + dt.timedelta(minutes=i * interval)) \
                .isoformat(timespec="minutes")
            crm.schedule_message(l, "Email", msg, when, follow_up_step="Email 1",
                                 status="Scheduled", notes=f"EmailCamp|SUBJ|{subj}")
        crm.set_kv(WORKFLOW_SHEET, "email_daily_cap", daily_cap, WORKFLOW_DEFAULTS)
        crm.set_kv(WORKFLOW_SHEET, "fu1_to_fu2_hours", fu_gap, WORKFLOW_DEFAULTS)
        if persist():
            st.success(f"📬 {len(picked)} correos en cola (Scheduled_Messages). "
                       "Actívalos abajo o envíalos manualmente.")

    # --- Auto-send con advertencia y procesamiento de cola ---
    st.divider()
    st.subheader("Auto-send")
    st.warning("⚠️ **Antes de activar:** respeta los límites de envío de Gmail "
               "(cuentas normales ~500/día; Workspace ~2,000/día), las leyes "
               "anti-spam aplicables (CAN-SPAM, GDPR/LOPD: identifícate, ofrece "
               "opt-out y hónralo) y la reputación de tu dominio. El tope diario "
               "configurado arriba se aplica siempre. Los envíos solo ocurren "
               "mientras la app está abierta.")
    autosend = st.toggle("Auto-send enabled", value=False)
    creds = st.session_state.get("gmail_creds")
    if autosend and not creds:
        st.error("Conecta Gmail primero (Notifications → 📧 Gmail).")
    pend_emails = [s for s in crm.read_scheduled(status="Scheduled")
                   if str(s.get("Channel")) in ("Email", "Gmail")]
    st.write(f"En cola: **{len(pend_emails)}** correos programados · "
             f"enviados hoy: **{crm.emails_sent_today()}** / {daily_cap}")
    if st.button("▶️ Procesar cola ahora", type="primary",
                 disabled=not (autosend and creds)):
        send_fn = lambda to, su, bo: gmail_service.send_email(creds, to, su, bo)
        s, k, reason = crm.process_due_emails(send_fn, daily_cap=daily_cap)
        if persist():
            st.success(f"✅ {s} enviados · {k} saltados"
                       + (f" · {reason}" if reason else ""))

    # --- Leer respuestas/rebotes y aplicarlos ---
    st.divider()
    st.subheader("Respuestas y rebotes (Gmail)")
    if st.button("📥 Leer y clasificar respuestas", disabled=not creds):
        try:
            msgs = gmail_service.list_recent_messages(creds, days=7, max_results=50)
            matches = gmail_service.match_messages_to_leads(msgs, crm)
            applied = 0
            for msg, lead in matches:
                cat = crm.classify_email_reply(msg["from_email"], msg["subject"],
                                               msg["snippet"])
                crm.handle_email_event(lead, cat, message=msg["snippet"])
                applied += 1
            # rebotes vienen de mailer-daemon (no matchean lead por remitente):
            for msg in msgs:
                if crm.classify_email_reply(msg["from_email"], msg["subject"],
                                            msg["snippet"]) == "bounce":
                    st.info(f"Rebote detectado: {msg['subject'][:70]} — identifica "
                            "el lead y márcalo en Estados si aplica.")
            if persist():
                st.success(f"{applied} respuestas clasificadas y aplicadas "
                           "(follow-ups detenidos donde corresponde).")
        except Exception as e:
            st.error(f"No pude leer el buzón: {e}")

# =========================================================================== #
# LINKEDIN MANAGER (asistido: Cliente Ideal + campañas por Fit + cola)

# =========================================================================== #
elif page == "💼 LinkedIn Manager":
    st.header("💼 LinkedIn Manager")
    st.caption("Modo seguro/asistido: sin scraping, sin bots, sin Selenium/"
               "Playwright. La app organiza, sugiere mensajes y registra; "
               "tú envías manualmente.")

    tabDash, tabSess, tabProsp, tabInbox, tabNC, tabHist, tabSet = st.tabs(
        ["📊 Dashboard", "🌐 Browser / Session", "🎯 Prospecting", "📥 Inbox",
         "🔔 Notifications", "🕓 History", "⚙️ Settings"])
    icp = st.session_state.get("icp") or crm.read_icp()

    # ---------- Dashboard (resumen LinkedIn) ----------
    with tabDash:
        camps_li = [c for c in crm.read_campaigns()
                    if c.get("Channel") == "LinkedIn"]
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("Campañas LinkedIn", len(camps_li))
        d2.metric("Activas", len([c for c in camps_li
                                  if c.get("Status") == "Activa"]))
        prof = ls_session.confirmed_icp(crm)
        d3.metric("ICP confirmado", "Sí" if prof else "No")
        unread = len(li_inbox.read_li_notifications(crm, unread_only=True))
        d4.metric("Notif. no leídas", unread)
        li_pend = sum(len(crm.campaign_pending_leads(c)) for c in camps_li
                      if c.get("Status") in ("Activa", "Pendiente"))
        st.metric("Leads pendientes en cola LinkedIn", li_pend)
        if camps_li:
            st.dataframe(pd.DataFrame([
                {"Campaña": c["Name"], "Estado": c["Status"],
                 "Enviados": c.get("Sent"), "Total": c.get("Total Leads"),
                 "Fecha": c.get("Scheduled Date")} for c in camps_li]),
                use_container_width=True, hide_index=True)
        else:
            st.info("Aún no hay campañas LinkedIn. Crea el ICP en Settings y "
                    "lanza una en Prospecting.")

    # ---------- Browser / Session ----------
    with tabSess:
        st.caption("La app NO automatiza el navegador ni inicia sesión por ti "
                   "(política de LinkedIn y del proyecto). 'Open LinkedIn' abre "
                   "el sitio en tu navegador; el login es tuyo. La conexión solo "
                   "se marca tras tu confirmación real de sesión iniciada.")
        st.info(ls_session.environment_note())

        # estrategia de lanzamiento (solo manual activo; resto slots futuros)
        opts = list(ls_session.LAUNCHERS.keys())
        strat = st.selectbox(
            "Estrategia de sesión", opts,
            format_func=lambda k: ls_session.LAUNCHERS[k].name +
            ("" if ls_session.LAUNCHERS[k].enabled else " (no disponible)"))
        if not ls_session.LAUNCHERS[strat].enabled:
            st.warning(ls_session.LAUNCHERS[strat].reason_disabled)

        sess = st.session_state.setdefault("li_session",
                                           ls_session.blank_session())
        badge = {"Connected": "🟢", "Disconnected": "⚪",
                 "Verification Required": "🟠", "Session Expired": "🔴",
                 "Error": "🔴"}.get(sess["status"], "⚪")
        st.subheader(f"{badge} LinkedIn: {sess['status']}")
        if sess.get("confirmed_at"):
            st.caption(f"Confirmada: {sess['confirmed_at']} · sesión "
                       f"{sess.get('session_id')}")

        b1, b2, b3, b4, b5 = st.columns(5)
        if b1.button("Connect LinkedIn"):
            ls_session.set_session_status(
                sess, "Verification Required",
                "Abre LinkedIn, inicia sesión y confirma abajo.")
            st.rerun()
        b2.link_button("Open LinkedIn",
                       ls_session.LAUNCHERS["manual"].open_linkedin_url())
        if b3.button("Verify Session"):
            st.session_state["li_verify"] = True
        if b4.button("Reconnect"):
            ls_session.set_session_status(sess, "Verification Required",
                                          "Reintenta el inicio de sesión.")
            st.rerun()
        if b5.button("Disconnect"):
            st.session_state["li_session"] = ls_session.blank_session()
            st.rerun()

        # verificación honesta: confirmación humana (no basta cambiar URL)
        if st.session_state.get("li_verify") or \
                sess["status"] == "Verification Required":
            st.divider()
            st.write("**Verificación de sesión** — como no automatizamos el "
                     "navegador, confirma tú que ya iniciaste sesión en "
                     "LinkedIn en tu equipo.")
            chk = st.checkbox("Confirmo que inicié sesión en LinkedIn en mi "
                              "navegador (sin CAPTCHA/2FA pendiente)")
            cc1, cc2 = st.columns(2)
            if cc1.button("✅ Marcar como conectado", disabled=not chk,
                          type="primary"):
                ls_session.set_session_status(sess, "Connected",
                                              "Confirmada por el usuario.")
                st.session_state.pop("li_verify", None)
                st.success("Sesión marcada como conectada.")
                st.rerun()
            if cc2.button("⚠️ LinkedIn me pide verificación (CAPTCHA/2FA)"):
                ls_session.set_session_status(
                    sess, "Verification Required",
                    "LinkedIn requiere verificación manual.")
                st.warning("LinkedIn requiere verificación manual. Complétala "
                           "en tu navegador; la app no la resuelve por ti.")

    # ---------- Prospecting ----------
    with tabProsp:
        prof = ls_session.confirmed_icp(crm)
        if not prof:
            st.warning("⛔ Debes definir y **confirmar el ICP** en la pestaña "
                       "⚙️ Settings antes de prospectar por LinkedIn.")
        else:
            st.success(f"ICP activo: {prof['name']} (v{prof['version']})")
            icp_obj = prof
            colf = st.columns(4)
            f_ind = colf[0].multiselect("Industry", IND)
            f_comp = colf[1].text_input("Company contiene")
            f_title = colf[2].text_input("Job Title contiene")
            f_n = colf[3].number_input("Máx. leads", 1, 300, 50)
            only_elig = st.checkbox("Solo elegibles (HIGH/MEDIUM fit)",
                                    value=True)

            if st.button("🔎 Evaluar leads vs ICP", type="primary"):
                rows, cards = [], []
                for ind in (f_ind or IND):
                    for l in crm.all_leads(ind):
                        if f_comp and _norm(f_comp) not in _norm(l.company):
                            continue
                        if f_title and _norm(f_title) not in _norm(
                                l.job_title or ""):
                            continue
                        ev = ls_session.evaluate_lead_against_icp(crm, l,
                                                                  icp_obj)
                        if only_elig and not ev["eligible"]:
                            continue
                        stt = crm.state.get(l.key)
                        rows.append({
                            "Full Name": l.full_name, "Job Title": l.job_title,
                            "Company": l.company, "Industry": l.industry,
                            "Seniority": l.seniority_level or "—",
                            "LinkedIn URL": l.linkedin or "—",
                            "ICP Score": ev["score"],
                            "Fit": ev["fit_class"],
                            "CRM Status": crm.lead_status(l, stt),
                            "LinkedIn Status":
                                li_inbox.get_lead_state(crm, l),
                            "Next Action": ev["recommended_action"]})
                        cards.append((l.sheet, l.row))
                        if len(rows) >= f_n:
                            break
                    if len(rows) >= f_n:
                        break
                rows_sorted = sorted(zip(rows, cards),
                                     key=lambda t: -t[0]["ICP Score"])
                st.session_state["li_prosp"] = rows_sorted

            data = st.session_state.get("li_prosp", [])
            if data:
                st.dataframe(pd.DataFrame([r for r, _c in data]),
                             use_container_width=True, hide_index=True)
                labels = [f"{r['Full Name']} · {r['Company']} · "
                          f"{r['ICP Score']} ({r['Fit']})" for r, _c in data]
                si = st.selectbox("Lead", range(len(data)),
                                  format_func=lambda i: labels[i])
                _row, (s, rw) = data[si]
                lead = crm.read_lead(s, rw)
                ev = ls_session.evaluate_lead_against_icp(crm, lead, icp_obj)

                h1, h2, h3 = st.columns(3)
                h1.metric("ICP Score", ev["score"])
                h2.metric("Fit", ev["fit_class"])
                h3.metric("LinkedIn Status",
                          li_inbox.get_lead_state(crm, lead))
                g1, g2 = st.columns(2)
                with g1:
                    st.markdown("**✅ Criterios cumplidos**")
                    for m in ev["matched_criteria"] or ["—"]:
                        st.write("• " + m)
                with g2:
                    st.markdown("**⚠️ Fallidos / warnings**")
                    for m in (ev["failed_criteria"] + ev["warnings"]) or ["—"]:
                        st.write("• " + m)
                if lead.linkedin:
                    st.link_button("Open Profile", lead.linkedin)

                # botones de prospección
                a = st.columns(4)
                if a[0].button("Review Lead"):
                    li_inbox.transition_lead_state(crm, lead, "REVIEW_REQUIRED",
                                                   "revisión", "user")
                    persist()
                if a[1].button("Prepare Connection"):
                    st.session_state["li_prepare"] = (s, rw)
                if a[2].button("Prepare Message"):
                    prep = li_inbox.prepare_linkedin_message(crm, lead,
                                                             None, icp)
                    st.session_state["li_prepmsg"] = prep
                if a[3].button("Mark Contacted"):
                    step, nxt = crm.assisted_mark_sent(lead, channel="LinkedIn")
                    li_inbox.transition_lead_state(crm, lead, "CONTACTED",
                                                   "marcado contactado", "user")
                    if persist():
                        st.success(f"Contactado ({step}).")
                a2 = st.columns(4)
                if a2[0].button("Skip"):
                    st.session_state["li_prosp"] = [d for d in data
                                                    if d[1] != (s, rw)]
                    st.rerun()
                if a2[1].button("Blacklist"):
                    crm.set_outcome(lead, "Blacklist")
                    li_inbox.transition_lead_state(crm, lead, "BLACKLISTED",
                                                   "blacklist", "user")
                    if persist():
                        st.success("Blacklist.")
                if a2[2].button("Prospect Later"):
                    crm.set_outcome(lead, "Prospectar Después")
                    li_inbox.transition_lead_state(crm, lead, "PROSPECT_LATER",
                                                   "later", "user")
                    if persist():
                        st.success("Prospectar después.")
                with a2[3].popover("View History"):
                    log = crm.state.get(lead.key).log
                    st.dataframe(pd.DataFrame(log) if log else
                                 pd.DataFrame([{"info": "sin historial"}]),
                                 use_container_width=True, hide_index=True)

                # preview de conexión (10 pasos) + confirmación
                if st.session_state.get("li_prepare") == (s, rw):
                    st.divider()
                    st.subheader("🔗 Prepare Connection (asistido, manual)")
                    prep = ls_session.prepare_connection(crm, lead, icp_obj)
                    for c in prep["checks"]:
                        st.write(("✅ " if c["ok"] else "❌ ") + c["paso"] +
                                 (f" — {c['detalle']}" if c["detalle"] else ""))
                    st.info(prep["preview"])
                    if prep["ready"]:
                        st.caption("La conexión la envías TÚ en LinkedIn. Este "
                                   "botón solo registra que confirmaste la "
                                   "acción; no automatiza nada.")
                        if lead.linkedin:
                            st.link_button("Abrir perfil para conectar",
                                           lead.linkedin)
                        if st.button("✔️ Confirm Connection Action",
                                     type="primary"):
                            ls_session.log_connection_action(
                                crm, lead, "connection_confirmed", "ok",
                                session_id=st.session_state.get(
                                    "li_session", {}).get("session_id", ""))
                            li_inbox.transition_lead_state(
                                crm, lead, "CONNECTION_PENDING",
                                "confirmada por usuario", "user")
                            if persist():
                                st.success("Acción registrada. Envía la "
                                           "solicitud en LinkedIn.")
                    else:
                        st.error("No listo para conexión (revisa los pasos).")

                if st.session_state.get("li_prepmsg"):
                    pm = st.session_state["li_prepmsg"]
                    st.divider()
                    st.subheader("✉️ Prepare Message")
                    st.text_area("Mensaje", pm["message"], height=110,
                                 key="li_pm_txt")
                    st.caption(f"{pm['char_count']} caracteres")
                    for w in pm["warnings"]:
                        st.warning(w)

    # ---------- Inbox ----------
    with tabInbox:
        _render_inbox()

    # ---------- Notifications ----------
    with tabNC:
        _render_li_notifications()

    # ---------- History ----------
    with tabHist:
        st.subheader("Historial de acciones de conexión")
        if "LinkedIn_Connection_Log" in crm.wb.sheetnames:
            ws = crm.wb["LinkedIn_Connection_Log"]
            rows = [[ws.cell(r, c).value for c in
                     range(1, len(ls_session.CONNECTION_LOG_HEADERS) + 1)]
                    for r in range(2, ws.max_row + 1)]
            if rows:
                st.dataframe(pd.DataFrame(
                    rows, columns=ls_session.CONNECTION_LOG_HEADERS),
                    use_container_width=True, hide_index=True)
            else:
                st.info("Sin acciones registradas.")
        else:
            st.info("Sin acciones de conexión todavía.")
        st.subheader("Evaluaciones ICP guardadas")
        if "Lead_Fit_Evaluation" in crm.wb.sheetnames:
            ws = crm.wb["Lead_Fit_Evaluation"]
            from crm_core import LEAD_FIT_HEADERS
            rows = [[ws.cell(r, c).value for c in
                     range(1, len(LEAD_FIT_HEADERS) + 1)]
                    for r in range(2, ws.max_row + 1)]
            if rows:
                st.dataframe(pd.DataFrame(rows, columns=LEAD_FIT_HEADERS).tail(
                    50), use_container_width=True, hide_index=True)

    # ---------- Settings (ICP obligatorio) ----------
    with tabSet:
        st.subheader("🧭 Perfil de Cliente Ideal (obligatorio antes de campañas)")
        st.caption("Los campos multiselección se poblan con los valores reales "
                   "de tus leads. Elegir una industria filtra empresas, puestos "
                   "y seniorities sugeridos.")
        base_ind = st.multiselect("Target Industries",
                                  ls_session.suggest_icp_options(crm)
                                  ["industries"], key="icp_ind")
        sug = ls_session.suggest_icp_options(crm, industries=base_ind or None)

        def multi_custom(label, options, key):
            chosen = st.multiselect(label, options, key=key)
            extra = st.text_input(f"➕ Agregar a {label} (coma)",
                                  key=key + "_x")
            if extra:
                chosen = chosen + [e.strip() for e in extra.split(",")
                                   if e.strip()]
            return chosen

        c1, c2 = st.columns(2)
        with c1:
            sen = multi_custom("Target Seniority Levels", sug["seniorities"],
                               "icp_sen")
            titles = multi_custom("Buyer Personas / Job Titles",
                                  sug["job_titles"], "icp_titles")
            comps = multi_custom("Target Companies", sug["companies"],
                                 "icp_comps")
        with c2:
            countries = multi_custom("Target Countries / Regions",
                                     sug["countries"], "icp_countries")
            specialties = st.text_input("Specialties / Capabilities (coma)",
                                        key="icp_spec")
            keywords = st.text_input("Keywords (coma)", key="icp_kw")

        with st.expander("Campos opcionales"):
            problems = st.text_input("Problems We Solve", key="icp_prob")
            valprop = st.text_input("Value Proposition", key="icp_vp")
            avoid = st.text_input("Leads to Avoid", key="icp_avoid")
            excl_comp = multi_custom("Excluded Companies", sug["companies"],
                                     "icp_xcomp")
            excl_titles = multi_custom("Excluded Job Titles",
                                       sug["job_titles"], "icp_xtitles")
            excl_ind = st.multiselect("Excluded Industries",
                                      ls_session.suggest_icp_options(crm)
                                      ["industries"], key="icp_xind")
            notes = st.text_input("Notes", key="icp_notes")

        name = st.text_input("Nombre del ICP", "Mi Cliente Ideal",
                             key="icp_name")
        if st.button("✅ Confirmar y guardar ICP", type="primary"):
            criteria = {
                "target_industries": base_ind, "seniorities": sen,
                "job_titles": titles, "target_companies": comps,
                "countries": countries,
                "specialties": [s.strip() for s in specialties.split(",")
                                if s.strip()],
                "keywords": [k.strip() for k in keywords.split(",")
                             if k.strip()],
                "problems": problems, "value_proposition": valprop,
                "leads_to_avoid": [a.strip() for a in avoid.split(",")
                                   if a.strip()],
                "excluded_companies": excl_comp,
                "excluded_job_titles": excl_titles,
                "excluded_industries": excl_ind, "notes": notes,
            }
            icp_id, ver = ls_session.save_icp_profile(
                crm, name, criteria, status="confirmed")
            # también refleja en el ICP simple (para evaluate_lead_fit)
            crm.save_icp({
                "industrias_objetivo": ", ".join(base_ind),
                "puestos_objetivo": ", ".join(titles),
                "seniority_ideal": ", ".join(sen),
                "empresas_objetivo": ", ".join(comps),
                "ubicacion_ideal": ", ".join(countries),
                "problemas_que_resuelvo": problems,
                "beneficios": valprop,
                "keywords_positivas": ", ".join(
                    [k.strip() for k in keywords.split(",") if k.strip()]),
                "keywords_negativas": "",
                "leads_a_evitar": avoid,
                "canales_preferidos": "LinkedIn, Email, Cold Call",
                "prioridad_minima": "Baja",
                "mensaje_base": crm.read_icp().get("mensaje_base", ""),
            })
            if persist():
                st.success(f"ICP '{name}' confirmado y guardado "
                           f"({icp_id} v{ver}). Ya puedes prospectar.")
        profs = ls_session.read_icp_profiles(crm)
        if profs:
            st.caption("ICPs guardados:")
            st.dataframe(pd.DataFrame([
                {"id": p["id"], "name": p["name"], "version": p["version"],
                 "status": p["status"], "updated_at": p["updated_at"]}
                for p in profs]), use_container_width=True, hide_index=True)

# =========================================================================== #
# WORKFLOW CONFIG (+ Follow Up Notification Settings)
# =========================================================================== #
elif page == "⚙️ Workflow Config":
    st.header("⚙️ Workflow Configuration")

    st.subheader("🔔 Follow Up Notification Settings")
    ns = crm.notif_settings()
    presets = ["24", "30", "48", "72", "76", "Personalizado"]
    cur = str(int(float(ns.get("no_response_hours", 48))))
    pick = st.selectbox("Revisar follow-ups sin respuesta cada (horas)", presets,
                        index=presets.index(cur) if cur in presets else 5)
    hours = st.number_input("Horas personalizadas", 1, 720,
                            int(float(ns.get("no_response_hours", 48)))) \
        if pick == "Personalizado" else int(pick)
    if st.button("Guardar configuración de notificaciones"):
        crm.set_kv(NOTIF_SETTINGS_SHEET, "no_response_hours", hours, NOTIF_DEFAULTS)
        if persist():
            st.success(f"Guardado en Notification_Settings: {hours}h.")

    st.divider()
    st.subheader("🔁 Workflow de follow-ups")
    wf = crm.workflow_config()
    w1, w2, w3 = st.columns(3)
    h12 = w1.number_input("Horas entre Follow Up 1 y 2", 1, 720,
                          int(float(wf["fu1_to_fu2_hours"])))
    h23 = w2.number_input("Horas entre Follow Up 2 y 3", 1, 720,
                          int(float(wf["fu2_to_fu3_hours"])))
    h34 = w3.number_input("Horas entre Follow Up 3 y 4", 1, 720,
                          int(float(wf["fu3_to_fu4_hours"])))
    w4, w5, w6 = st.columns(3)
    maxfu = w4.number_input("Máximo de Follow Ups", 1, 5,
                            int(float(wf["max_followups"])))
    defch = w5.selectbox("Canal por defecto", PROSPECT_CHANNELS,
                         index=PROSPECT_CHANNELS.index(wf["default_channel"])
                         if wf["default_channel"] in PROSPECT_CHANNELS else 0)
    interval = w6.number_input("Intervalo entre mensajes (min)", 1, 1440,
                               int(float(wf["message_interval_min"])))
    w7, w8 = st.columns(2)
    hours_rng = w7.text_input("Horario permitido (HH:MM-HH:MM)",
                              wf["allowed_hours"])
    days = w8.text_input("Días permitidos", wf["allowed_days"])
    if st.button("Guardar workflow", type="primary"):
        for k, v in [("fu1_to_fu2_hours", h12), ("fu2_to_fu3_hours", h23),
                     ("fu3_to_fu4_hours", h34), ("max_followups", maxfu),
                     ("default_channel", defch), ("allowed_hours", hours_rng),
                     ("allowed_days", days), ("message_interval_min", interval)]:
            crm.set_kv(WORKFLOW_SHEET, k, v, WORKFLOW_DEFAULTS)
        if persist():
            st.success("Workflow guardado; aplica a todas las campañas futuras.")

# =========================================================================== #
# 5) DASHBOARD
# =========================================================================== #
elif page == "📊 Dashboard":
    st.header("📊 Dashboard")

    # --- Métricas globales y desgloses (lead-level) ---
    @st.cache_data(show_spinner="Calculando métricas…")
    def _dash_metrics(mtime):
        total = nuevos = msgs = fus = resp = won = lost = prosp = black = 0
        by_ind, by_sen, by_chan, by_stage = {}, {}, {}, {}
        for sheet in crm.maps:
            for l in crm.all_leads(sheet):
                stt = crm.state.get(l.key)
                total += 1
                stage = crm.lead_stage(l, stt)
                chan = crm.lead_channel(l, stt)
                by_ind[l.industry] = by_ind.get(l.industry, 0) + 1
                sen = l.seniority_level or "—"
                by_sen[sen] = by_sen.get(sen, 0) + 1
                by_chan[chan] = by_chan.get(chan, 0) + 1
                by_stage[stage] = by_stage.get(stage, 0) + 1
                sv = l.stage_values
                fu_sent = sum(_is_marked(sv.get(s)) for s in CHANNEL_STEPS["LinkedIn"])
                em_sent = sum(_is_marked(sv.get(s)) for s in CHANNEL_STEPS["Email"])
                cc_sent = sum(_is_marked(sv.get(s)) for s in CHANNEL_STEPS["Cold Call"])
                msgs += fu_sent + em_sent + cc_sent
                fus += max(0, fu_sent - 1) + max(0, em_sent - 1)  # follow-ups (no el 1°)
                if stage == "Nuevo lead":
                    nuevos += 1
                if stt.responded or l.outcome == "Respondió":
                    resp += 1
                won += l.outcome == "Won"; lost += l.outcome == "Lost"
                prosp += l.outcome == "Prospectar Después"
                black += l.outcome == "Blacklist"
        return dict(total=total, nuevos=nuevos, msgs=msgs, fus=fus, resp=resp,
                    won=won, lost=lost, prosp=prosp, black=black,
                    by_ind=by_ind, by_sen=by_sen, by_chan=by_chan, by_stage=by_stage)

    M = _dash_metrics(Path(xlsx_path).stat().st_mtime)
    a, b, c, d = st.columns(4)
    a.metric("Total leads", M["total"]); b.metric("Leads nuevos", M["nuevos"])
    c.metric("Mensajes enviados", M["msgs"]); d.metric("Follow-ups enviados", M["fus"])
    e, f, g, h = st.columns(4)
    e.metric("Respuestas", M["resp"]); f.metric("Won", M["won"])
    g.metric("Lost", M["lost"]); h.metric("Prospectar después", M["prosp"])
    st.metric("Blacklist", M["black"])

    # --- Agenda: vencidos / hoy / próximos 7 días + pendientes por canal ---
    st.subheader("📅 Agenda de follow-ups")
    today = dt.date.today()
    overdue = due_today = next7 = cc_pend = 0
    rfq_active = 0
    state_counts = {}
    color_cfg = crm.read_state_colors()
    for ind in IND:
        for l in crm.all_leads(ind):
            stt = crm.state.get(l.key)
            disp = crm.resolve_display_state(l, stt)
            state_counts[disp] = state_counts.get(disp, 0) + 1
            if l.outcome in ("RFQ", "Quote") or _is_marked(l.stage_values.get("RFQ")):
                rfq_active += 1
            if crm._is_blocked(l, stt):
                continue
            for step, iso in stt.scheduled.items():
                d = dt.datetime.fromisoformat(iso).date()
                if d < today:
                    overdue += 1
                elif d == today:
                    due_today += 1
                elif d <= today + dt.timedelta(days=7):
                    next7 += 1
                if step in CHANNEL_STEPS["Cold Call"]:
                    cc_pend += 1
    gmail_pend = len([s for s in crm.read_scheduled()
                      if s.get("Status") == "Scheduled"
                      and str(s.get("Channel")) in ("Email", "Gmail")])

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("🔴 Vencidos", overdue)
    m2.metric("🟡 Para hoy", due_today)
    m3.metric("🗓 Próximos 7 días", next7)
    m4.metric("📨 RFQ activos", rfq_active)
    m5, m6 = st.columns(2)
    m5.metric("📞 Cold calls pendientes", cc_pend)
    m6.metric("📧 Gmail pendientes (agenda)", gmail_pend)

    # --- Métricas de email, campañas y LinkedIn asistido ---
    st.subheader("📧 Email & Campañas")
    sched_all = crm.read_scheduled()
    today_s = dt.date.today().isoformat()
    em_today = crm.emails_sent_today()
    em_prog = len([s for s in sched_all if s.get("Status") == "Scheduled"
                   and str(s.get("Channel")) in ("Email", "Gmail")])
    bounced = blocked_n = dnc = stopped = later_n = 0
    responded_n = 0
    for ind in IND:
        for l in crm.all_leads(ind):
            if l.outcome == "Email Bounced":
                bounced += 1
            if l.outcome == "Blocked":
                blocked_n += 1
            if l.outcome == "Do Not Contact":
                dnc += 1
            if l.outcome == "Prospectar Después":
                later_n += 1
            sst = crm.state.get(l.key)
            if sst.responded or l.outcome == "Respondió":
                responded_n += 1
                if any(e.get("action") == "sent" for e in sst.log):
                    stopped += 1
    camps_all = crm.read_campaigns()
    li_pend = sum(len(crm.campaign_pending_leads(c)) for c in camps_all
                  if c.get("Channel") == "LinkedIn"
                  and c.get("Status") in ("Activa", "Pendiente"))
    li_sent = sum(1 for r in crm._read_activity_log()
                  if str(r[5]) == "LinkedIn" and str(r[9]) == "Enviado")
    e1, e2, e3, e4 = st.columns(4)
    e1.metric("Emails enviados hoy", em_today)
    e2.metric("Emails programados", em_prog)
    e3.metric("Emails rebotados", bounced)
    e4.metric("Leads bloqueados", blocked_n + dnc)
    e5, e6, e7, e8 = st.columns(4)
    e5.metric("Campañas activas",
              len([c for c in camps_all if c.get("Status") == "Activa"]))
    e6.metric("Campañas pendientes",
              len([c for c in camps_all if c.get("Status") == "Pendiente"]))
    e7.metric("Campañas pausadas",
              len([c for c in camps_all if c.get("Status") == "Pausada"]))
    e8.metric("Campañas terminadas",
              len([c for c in camps_all if c.get("Status") == "Terminada"]))
    e9, e10, e11, e12 = st.columns(4)
    e9.metric("Respuestas recibidas", responded_n)
    e10.metric("Follow-ups auto detenidos", stopped)
    e11.metric("Contactar después", later_n)
    e12.metric("LinkedIn pendientes", li_pend)
    st.metric("LinkedIn enviados manualmente", li_sent)
    # Follow Up N pendientes por campaña
    fu_pend = {}
    for c in camps_all:
        if c.get("Status") in ("Activa", "Pendiente"):
            stepn = int(c.get("Step") or 0) + 1
            fu_pend[f"Follow Up {stepn}"] = fu_pend.get(f"Follow Up {stepn}", 0) + \
                len(crm.campaign_pending_leads(c))
    if fu_pend:
        st.caption("Pendientes por paso de campaña: " +
                   " · ".join(f"{k}: {v}" for k, v in sorted(fu_pend.items())))

    # --- Leads por estado/color (según State_Color_Config) ---
    st.subheader("🎨 Leads por estado / color")
    if state_counts:
        rows = []
        for stn, cnt in sorted(state_counts.items(), key=lambda kv: -kv[1]):
            c = color_cfg.get(stn, {})
            rows.append({"Estado": stn, "Leads": cnt,
                         "Color": f"#{c.get('Color','')}" if c.get("Color") else "—",
                         "Categoría": c.get("Category", "—"),
                         "Acción sugerida": c.get("Suggested Action", "—")})
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        st.bar_chart(pd.Series(state_counts))

    st.subheader("Distribuciones")
    d1, d2 = st.columns(2)
    with d1:
        st.markdown("**Leads por industria**")
        st.bar_chart(pd.Series(M["by_ind"]))
        st.markdown("**Leads por canal**")
        st.bar_chart(pd.Series(M["by_chan"]))
    with d2:
        st.markdown("**Leads por seniority level**")
        st.bar_chart(pd.Series(M["by_sen"]))
        st.markdown("**Leads por etapa**")
        st.bar_chart(pd.Series(M["by_stage"]))

    st.divider()
    st.subheader("📊 Dashboard por industria")
    rows = []
    for ind in IND:
        leads = crm.all_leads(ind)
        total = len(leads)
        def cnt(step):
            return sum(_is_marked(l.stage_values.get(step)) for l in leads)
        contacted = sum(
            any(_is_marked(l.stage_values.get(s)) for s in
                ["Follow Up 1", "Email 1", "Cold Call 1"])
            for l in leads)
        outcome = lambda v: sum(l.outcome == v for l in leads)
        row = {"Industria": ind, "Total": total, "Contactados": contacted}
        for ch, steps in CHANNEL_STEPS.items():
            for s in steps:
                row[s] = cnt(s)
        for s in STAGE_COLS:
            row[s] = cnt(s)
        row["Won"] = outcome("Won"); row["Lost"] = outcome("Lost")
        row["Blacklist"] = outcome("Blacklist")
        row["Prospectar Después"] = outcome("Prospectar Después")
        row["Conversión %"] = round(100 * row["Won"] / total, 1) if total else 0.0
        row["% Contactado"] = round(100 * contacted / total, 1) if total else 0.0
        rows.append(row)
    df = pd.DataFrame(rows)

    st.subheader("Resumen (Total · % sobre leads · conversión)")
    st.dataframe(df, use_container_width=True, hide_index=True)

    # % sobre contactados para cada métrica
    st.subheader("Las 8 gráficas")
    df_i = df.set_index("Industria")
    groups = {
        "1 · Follow-ups": CHANNEL_STEPS["LinkedIn"],
        "2 · Emails": CHANNEL_STEPS["Email"],
        "3 · Cold Calls": CHANNEL_STEPS["Cold Call"],
        "4 · Meeting/RFQ/Quote": STAGE_COLS,
        "5 · Won": ["Won"], "6 · Lost": ["Lost"],
        "7 · Blacklist": ["Blacklist"], "8 · Prospectar Después": ["Prospectar Después"],
    }
    cols = st.columns(2)
    for i, (title, metrics) in enumerate(groups.items()):
        with cols[i % 2]:
            st.markdown(f"**{title}**")
            st.bar_chart(df_i[metrics])

    st.divider()
    st.subheader("📅 Reporte diario")
    st.caption("Genera un reporte con 8 hojas: Leads_Actualizados, Activity_Log, "
               "Follow_Up, Emails_Enviados, Respondieron, Blacklist, "
               "Pendientes_Mañana y Resumen_Dia.")

    fmt = st.radio("Formato", ["Excel (.xlsx)", "CSV (.zip)"], horizontal=True)

    # Vista previa (sin generar archivo): muestra cada hoja del reporte
    if st.checkbox("Previsualizar contenido"):
        data = crm.build_daily_report_data()
        for name, sh in data.items():
            with st.expander(f"{name} · {len(sh['rows'])} filas"):
                st.dataframe(pd.DataFrame(sh["rows"], columns=sh["headers"]),
                             use_container_width=True, hide_index=True)

    if st.button("Exportar reporte del día", type="primary"):
        out_dir = str(Path(xlsx_path).resolve().parent)
        kind = "csv" if fmt.startswith("CSV") else "xlsx"
        rep = crm.export_daily_report(out_dir=out_dir, fmt=kind)
        st.success(f"Reporte generado: {Path(rep).name}")
        mime = ("application/zip" if kind == "csv" else
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with open(rep, "rb") as fh:
            st.download_button("Descargar reporte", fh.read(),
                               file_name=Path(rep).name, mime=mime)

    st.divider()
    st.subheader("📊 Dashboard exportable (multi-hoja)")
    st.caption("Excel con hojas separadas por Industria, Seniority, Current Stage, "
               "Follow Up Step, Canal, Won, Lost, Prospectar después, Blacklist, "
               "Scheduled_Messages, Activity_Log y Notifications, más un Resumen. "
               "Incluye automáticamente los leads nuevos importados.")
    if st.button("Exportar dashboard completo", type="primary"):
        out_dir = str(Path(xlsx_path).resolve().parent)
        rep = crm.export_dashboard(out_dir=out_dir)
        st.success(f"Dashboard generado: {Path(rep).name}")
        with open(rep, "rb") as fh:
            st.download_button(
                "Descargar dashboard", fh.read(), file_name=Path(rep).name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
