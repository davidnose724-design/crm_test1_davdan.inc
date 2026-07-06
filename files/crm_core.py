"""
crm_core.py — Motor del CRM + asistente de prospección B2B.

Principios de diseño (acordados con el usuario):
- El Excel es la fuente de verdad para las columnas que ya tiene.
- El estado que el Excel NO puede guardar (respondió, fecha de recontacto,
  follow-ups programados, historial de mensajes, log) vive en un sidecar JSON.
- NUNCA se altera el orden de columnas ni se rompen las fórmulas del Dashboard.
- Los colores de fila salen del formato condicional sobre 'Outcome Status' (AH);
  la app NO pinta celdas a mano.
- Todo envío es asistido con aprobación humana. La app prepara/programa/registra;
  el disparo real lo hace la persona y luego marca "Enviado".
"""

from __future__ import annotations
import json, re, shutil, datetime as dt
from dataclasses import dataclass, field, asdict
from pathlib import Path
from copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------- #
# Constantes del dominio
# --------------------------------------------------------------------------- #

INDUSTRIES = [
    "Automotive", "Pharmaceutical", "Medical Devices", "Industrial Equipment",
    "HVAC Appliances", "Packaging", "Consumer Goods", "Software IT", "Other",
]
ACTIVITY_LOG_SHEET = "Activity_Log"
NOTIFICATIONS_SHEET = "Notifications"
NON_DATA_SHEETS = {"Dashboard", "Lead Queue", ACTIVITY_LOG_SHEET, NOTIFICATIONS_SHEET,
                   "Scheduled_Messages", "Follow_Up_History", "State_Color_Config",
                   "Notification_Settings", "Workflow_Config", "Campaigns",
                   "Gmail_Accounts", "Gmail_Campaigns", "Gmail_Campaign_Leads",
                   "Gmail_Follow_Ups", "Ideal_Customer_Profile"}

# Encabezados de la hoja Notifications.
NOTIFICATION_HEADERS = [
    "Timestamp", "Fecha", "Lead", "Empresa", "Canal",
    "Tipo de evento", "Mensaje/Resumen", "Estado actual", "Visto",
]

# Estados que el usuario puede fijar manualmente desde la sección Estados.
# Los que son Outcome Status pintan fila; los de embudo marcan celdas de etapa.
MANUAL_STATES = [
    "Nuevo lead", "Mensaje inicial enviado", "Follow Up 1", "Follow Up 2",
    "Follow Up 3", "Respondió", "Won", "Lost", "Prospectar después",
    "Blacklist", "Reunión agendada", "Cotización solicitada",
]

# Canales válidos para el registro de actividad (incluye WhatsApp y llamada).
ACTIVITY_CHANNELS = ["LinkedIn", "Email", "Llamada", "WhatsApp", "Cold Call"]

# Campos que el importador intenta detectar/mapear desde un archivo externo.
IMPORT_FIELDS = [
    "Full Name", "First Name", "Last Name", "Company Name", "Job Title",
    "Industry", "Seniority Level", "Location", "LinkedIn Profile", "Email/Gmail",
    "Phone", "Value Proposition", "Pain Point", "Notes",
]

# Columnas que NO existen en el archivo original y se agregan AL FINAL (posición
# >= 41) para no mover nada que el Dashboard lea por posición.
EXTRA_COLUMNS = ["Value Proposition", "Pain Point",
                 "Lead Warmth Level", "Priority", "Recommended Channel", "Current Stage"]

# Columnas de clasificación que la app calcula/persiste al importar.
CLASSIFY_COLUMNS = ["Lead Warmth Level", "Priority", "Recommended Channel", "Current Stage"]

# Hoja de programación de mensajes (agenda manual; nunca se envía solo).
SCHEDULED_SHEET = "Scheduled_Messages"

# --- Reprogramación de follow-ups -------------------------------------------
FOLLOWUP_HISTORY_SHEET = "Follow_Up_History"
FU_HISTORY_HEADERS = [
    "Timestamp", "Lead ID", "Full Name", "Company", "Previous Follow Up Date",
    "New Follow Up Date", "Previous Channel", "New Channel", "Reason", "Notes",
    "User", "Status",
]
# Columnas editables nuevas (se crean AL FINAL si faltan).
RESCHEDULE_COLUMNS = ["Follow Up Step", "Follow Up Channel", "Next Follow Up Time",
                      "Follow Up Reason", "Owner/User"]
FOLLOWUP_CHANNELS = ["Gmail", "LinkedIn", "WhatsApp", "Cold Call", "Otro"]

# --- Configuración de colores por estado -------------------------------------
STATE_COLOR_SHEET = "State_Color_Config"
STATE_COLOR_HEADERS = ["State", "Color", "Scope", "Category", "Suggested Action",
                       "Priority", "Active"]
STATE_CATEGORIES = ["activo", "cerrado", "pendiente", "bloqueo"]
# Defaults sugeridos por el usuario (editables desde la app).
DEFAULT_STATE_COLORS = [
    # (State, ColorHex, Scope, Category, Suggested Action, Priority)
    ("Won",                  "00B050", "row",  "cerrado",   "Onboarding / celebrar", "Alta"),
    ("Lost",                 "FF0000", "row",  "cerrado",   "Analizar razón",        "Baja"),
    ("RFQ",                  "9DC3E6", "row",  "activo",    "Enviar cotización",     "Alta"),
    ("Reunión agendada",     "C9A0DC", "row",  "activo",    "Preparar reunión",      "Alta"),
    ("Prospectar después",   "FFFF99", "row",  "pendiente", "Recontactar en fecha",  "Media"),
    ("Respondió interesado", "C6EFCE", "row",  "activo",    "Dar seguimiento",       "Alta"),
    ("No interesado",        "FFC000", "row",  "cerrado",   "Cierre amable",         "Baja"),
    ("Blacklist",            "808080", "row",  "bloqueo",   "No contactar",          "Baja"),
    ("Follow Up pendiente",  "FFF2CC", "cell", "pendiente", "Enviar follow-up",      "Media"),
    ("Mensaje enviado",      "DDEBF7", "cell", "activo",    "Esperar respuesta",     "Media"),
    ("Nuevo lead",           "FFFFFF", "cell", "activo",    "Mensaje inicial",       "Media"),
]

# Resultados posibles de una cold call.
COLD_CALL_RESULTS = [
    "No contestó", "Interesado", "Pidió información", "RFQ", "Reunión agendada",
    "Contactar después", "Won", "Lost", "Blacklist",
]
SCHEDULED_HEADERS = [
    "Schedule ID", "Lead ID", "Full Name", "Company", "Seniority Level", "Industry",
    "Channel", "Message", "Scheduled DateTime", "Status", "Sent DateTime",
    "Follow Up Step", "Notes",
]
SCHEDULE_STATUS = ["Draft", "Scheduled", "Sent", "Cancelled"]

# Niveles de calentamiento del lead.
WARMTH_LEVELS = ["Frío", "Tibio", "Caliente"]

# Canales de prospección soportados (para filtros y programación).
PROSPECT_CHANNELS = ["LinkedIn", "Email", "WhatsApp", "Cold Call"]

# Industrias objetivo para clasificación por keywords en el import.
TARGET_INDUSTRY_KEYWORDS = {
    "Automotive": ["automotive", "auto", "vehicle", "car", "oem", "tier 1", "tier 2"],
    "Medical Devices": ["medical", "device", "healthcare", "pharma", "bio"],
    "Industrial Equipment": ["industrial", "equipment", "machinery", "machine"],
    "HVAC Appliances": ["hvac", "appliance", "cooling", "heating", "refriger"],
    "Packaging": ["packaging", "package", "thermoforming", "molding", "injection"],
    "Consumer Goods": ["consumer", "goods", "retail", "electronics", "appliances"],
    "Software IT": ["software", "it", "saas", "tech", "cloud"],
}

# Columnas recomendadas de la plantilla de importación (orden de la plantilla).
TEMPLATE_COLUMNS = [
    "First Name", "Last Name", "Full Name", "Job Title", "Company Name",
    "Industry", "Seniority Level", "LinkedIn URL", "Email", "Phone",
    "Location", "Value Proposition", "Pain Point", "Notes",
]

# Prioridad del lead a partir de Seniority Level / rol (se calcula, no se guarda).
PRIORITY_RULES = [
    ("Alta", ["owner", "founder", "ceo", "president", "chief", "cxo", "c-level",
              "partner", "managing director", "vp", "vice president", "director", "head"]),
    ("Media", ["manager", "buyer", "purchasing", "sourcing", "procurement", "engineer",
               "supplier quality", "lead", "senior", "specialist"]),
    ("Baja", ["assistant", "coordinator", "intern", "trainee"]),
]

# Encabezados de la hoja Activity_Log.
ACTIVITY_HEADERS = [
    "Timestamp", "Persona", "Empresa", "Industria", "Fecha",
    "Canal", "Follow-up #", "Paso", "Mensaje", "Resultado",
]

# Inferencia de seniority desde el Job Title cuando no viene en el archivo.
# (rank menor = más senior; usado solo para leads importados sin Seniority Rank)
SENIORITY_RULES = [
    ("C-Level", 1, ["chief", "cxo", "ceo", "cfo", "coo", "cto", "cpo", "president",
                    "founder", "owner", "partner", "managing director"]),
    ("VP", 2, ["vp", "vice president", "vicepresident"]),
    ("Director", 3, ["director", "head of"]),
    ("Head", 4, ["head"]),
    ("Manager", 5, ["manager", "gerente"]),
    ("Lead", 6, ["lead", "team lead"]),
    ("Senior", 7, ["senior", "sr.", "sr "]),
    ("Specialist", 8, ["specialist", "engineer", "buyer", "analyst", "coordinator"]),
]

# Nombres canónicos -> sinónimos aceptados (para mapeo difuso de columnas).
CANON_SYNONYMS = {
    "Full Name": ["full name", "name", "nombre completo", "nombre"],
    "First Name": ["first name", "nombre"],
    "Last Name": ["last name", "apellido"],
    "Job Title": ["job title", "title", "cargo", "puesto"],
    "Seniority Level": ["seniority level", "seniority", "nivel"],
    "Seniority Rank": ["seniority rank", "rank", "rango"],
    "Lead Score": ["lead score", "score", "puntaje"],
    "Opportunity Score": ["opportunity score", "opp score"],
    "Company Name": ["company name", "company", "empresa", "compania", "compañia"],
    "Industry": ["industry", "industria"],
    "Location": ["location", "ubicacion", "ubicación"],
    "Company Domain": ["company domain", "domain", "dominio"],
    "LinkedIn Profile": ["linkedin profile", "linkedin", "linkedin url", "perfil linkedin"],
    "Phone": ["phone", "telefono", "teléfono", "tel"],
    "Email/Gmail": ["email/gmail", "email", "gmail", "correo", "e-mail"],
    "Follow Up 1": ["follow up 1", "followup 1", "message 1", "msg 1"],
    "Follow Up 2": ["follow up 2", "followup 2"],
    "Follow Up 3": ["follow up 3", "followup 3"],
    "Follow Up 4": ["follow up 4", "followup 4"],
    "Follow Up 5": ["follow up 5", "followup 5"],
    "Email 1": ["email 1", "correo 1"], "Email 2": ["email 2", "correo 2"],
    "Email 3": ["email 3", "correo 3"], "Email 4": ["email 4", "correo 4"],
    "Email 5": ["email 5", "correo 5"],
    "Cold Call 1": ["cold call 1", "llamada 1", "call 1"],
    "Cold Call 2": ["cold call 2", "llamada 2", "call 2"],
    "Cold Call 3": ["cold call 3", "llamada 3", "call 3"],
    "Meeting": ["meeting", "reunion", "reunión"],
    "RFQ": ["rfq", "solicitud cotizacion"],
    "Quote": ["quote", "cotizacion", "cotización"],
    "Won": ["won", "ganado"], "Lost": ["lost", "perdido"],
    "Outcome Status": ["outcome status", "lead status", "status", "estado", "estatus"],
    "Rejection Reason": ["rejection reason", "razon rechazo", "motivo rechazo"],
    "Interest Reason": ["interest reason", "razon interes", "motivo interes"],
    "Notes": ["notes", "notas", "comentarios"],
    "Value Proposition": ["value proposition", "value prop", "propuesta de valor",
                          "propuesta valor", "valueprop"],
    "Pain Point": ["pain point", "pain points", "dolor", "punto de dolor",
                   "necesidad", "painpoint"],
    "Lead Warmth Level": ["lead warmth level", "warmth", "warmth level",
                          "calentamiento", "nivel de calentamiento"],
    "Priority": ["priority", "prioridad"],
    "Recommended Channel": ["recommended channel", "canal recomendado", "channel"],
    "Current Stage": ["current stage", "stage", "etapa", "fase", "etapa actual"],
    "Follow Up Step": ["follow up step", "fu step", "paso follow up"],
    "Follow Up Channel": ["follow up channel", "canal follow up", "fu channel"],
    "Next Follow Up Time": ["next follow up time", "hora follow up", "fu time"],
    "Follow Up Reason": ["follow up reason", "motivo follow up", "razon reprogramacion"],
    "Owner/User": ["owner/user", "owner", "user", "usuario", "responsable"],
    "First Contact": ["first contact", "primer contacto"],
    "Last Contact": ["last contact", "ultimo contacto", "último contacto"],
    "Next Follow-up": ["next follow-up", "next followup", "proximo seguimiento"],
}

# Secuencias por canal -> columnas canónicas, en orden.
CHANNEL_STEPS = {
    "LinkedIn":  ["Follow Up 1", "Follow Up 2", "Follow Up 3", "Follow Up 4", "Follow Up 5"],
    "Email":     ["Email 1", "Email 2", "Email 3", "Email 4", "Email 5"],
    "Cold Call": ["Cold Call 1", "Cold Call 2", "Cold Call 3"],
}
STAGE_COLS = ["Meeting", "RFQ", "Quote"]

# Prioridad de Lead Score (categórico) para ordenar: menor = mejor.
LEAD_SCORE_PRIORITY = {"A+": 0, "A": 1, "B+": 2, "B": 3, "C+": 4, "C": 5}

SENT = "Enviado"          # marca para etapas LinkedIn/Email
DONE = "Hecho"            # marca para Cold Call (dentro de la lista de validación)

# Estados de Outcome Status que bloquean / pausan envíos.
OUTCOME_BLOCK = {"Blacklist", "Won", "Lost", "Do Not Contact", "Email Bounced",
                 "Blocked"}     # bloqueo duro (nunca / no más)
OUTCOME_PAUSE = {"Prospectar Después", "Respondió"}  # pausa hasta decisión/fecha

# Valores nuevos de Outcome Status que la app puede escribir (extienden la validación).
OUTCOME_VALUES = ["Active", "Blacklist", "Prospectar Después", "Won", "Lost",
                  "Respondió", "Meeting", "RFQ", "Quote",
                  "Do Not Contact", "Email Bounced", "Blocked"]

# --- Configuración clave-valor (Notification_Settings / Workflow_Config) ----
NOTIF_SETTINGS_SHEET = "Notification_Settings"
WORKFLOW_SHEET = "Workflow_Config"
KV_HEADERS = ["Setting", "Value"]
NOTIF_DEFAULTS = {"no_response_hours": "48"}
WORKFLOW_DEFAULTS = {
    "fu1_to_fu2_hours": "72", "fu2_to_fu3_hours": "72", "fu3_to_fu4_hours": "96",
    "max_followups": "3", "default_channel": "LinkedIn",
    "allowed_hours": "09:00-18:00", "allowed_days": "Lun,Mar,Mie,Jue,Vie",
    "message_interval_min": "45", "email_daily_cap": "50",
}

# --- Campañas (workflow asistido) --------------------------------------------
CAMPAIGNS_SHEET = "Campaigns"
CAMPAIGN_HEADERS = ["Campaign ID", "Name", "Channel", "Industry", "Step",
                    "Status", "Total Leads", "Sent", "Scheduled Date",
                    "Created", "Notes", "Lead Keys"]
CAMPAIGN_STATUS = ["Activa", "Pendiente", "Pausada", "Terminada"]

# --- Perfil de Cliente Ideal (Buyer Persona) ---------------------------------
ICP_SHEET = "Ideal_Customer_Profile"
ICP_FIELDS = [
    ("que_vendo", "Qué vende mi empresa"),
    ("tipo_comprador", "Tipo de comprador ideal"),
    ("industrias_objetivo", "Industrias objetivo (separadas por coma)"),
    ("empresas_objetivo", "Empresas objetivo (separadas por coma)"),
    ("puestos_objetivo", "Puestos objetivo (separados por coma)"),
    ("seniority_ideal", "Seniority ideal (separados por coma)"),
    ("ubicacion_ideal", "País / ubicación ideal (separados por coma)"),
    ("tamano_empresa", "Tamaño de empresa ideal"),
    ("problemas_que_resuelvo", "Problemas que resuelvo (separados por coma)"),
    ("beneficios", "Beneficios principales (separados por coma)"),
    ("keywords_positivas", "Palabras clave positivas (separadas por coma)"),
    ("keywords_negativas", "Palabras clave negativas (separadas por coma)"),
    ("leads_a_evitar", "Leads a evitar (nombres/empresas/puestos, por coma)"),
    ("canales_preferidos", "Canales preferidos (LinkedIn, Email, Cold Call)"),
    ("prioridad_minima", "Nivel mínimo de prioridad aceptado (Alta/Media/Baja)"),
    ("mensaje_base", "Mensaje base de prospección"),
]
ICP_DEFAULTS = {k: "" for k, _ in ICP_FIELDS}
ICP_DEFAULTS.update({
    "canales_preferidos": "LinkedIn, Email, Cold Call",
    "prioridad_minima": "Baja",
    "mensaje_base": ("Hola {{first_name}}, vi tu rol de {{job_title}} en "
                     "{{company}}. Ayudamos a empresas como la tuya con "
                     "{{problema}} — {{beneficio}}. ¿Te hace sentido conectar?"),
})
FIT_LEVELS = ["Excelente", "Bueno", "Medio", "Bajo", "Malo"]

# Encabezado canónico de las hojas de leads (orden del archivo original).
CANONICAL_HEADER = [
    "Full Name", "First Name", "Last Name", "Job Title", "Seniority Level",
    "Seniority Rank", "Lead Score", "Opportunity Score", "Company Name",
    "Industry", "Location", "Company Domain", "LinkedIn Profile", "Phone",
    "Email/Gmail", "Follow Up 1", "Follow Up 2", "Follow Up 3", "Follow Up 4",
    "Follow Up 5", "Email 1", "Email 2", "Email 3", "Email 4", "Email 5",
    "Cold Call 1", "Cold Call 2", "Cold Call 3", "Meeting", "RFQ", "Quote",
    "Won", "Lost", "Outcome Status", "Rejection Reason", "Interest Reason",
    "Notes", "First Contact", "Last Contact", "Next Follow-up",
]

# Perfil de prospección: dolores probables y servicios por industria.
INDUSTRY_PAINS = {
    "Automotive": ("Lead times largos y presión de costos en tooling/estampado",
                   "Manufactura de precisión con reducción de costo por pieza"),
    "Pharmaceutical": ("Cumplimiento regulatorio y trazabilidad de componentes",
                       "Componentes con documentación y calidad certificada"),
    "Medical Devices": ("Tolerancias críticas y validación de proveedores",
                        "Maquinado de precisión con procesos validados"),
    "Industrial Equipment": ("Repuestos caros y proveedores lentos",
                             "Fabricación flexible de partes y ensambles"),
    "HVAC Appliances": ("Costos de componentes metálicos y estacionalidad",
                        "Estampado y ensamble con capacidad escalable"),
    "Packaging": ("Presión de precio y tiempos de cambio de molde",
                  "Herramentales y componentes de empaque competitivos"),
    "Consumer Goods": ("Volúmenes variables y time-to-market",
                       "Producción ágil con escalamiento rápido"),
    "Software IT": ("Hardware/prototipos y cadena de suministro física",
                    "Manufactura de soporte para hardware y prototipos"),
}

# --- Gmail Campaigns ----------------------------------------------------------
GMAIL_ACCOUNTS_SHEET = "Gmail_Accounts"
GMAIL_ACCOUNTS_HEADERS = ["Email", "Connected At", "Last Sync", "Status"]
GMAIL_CAMPAIGNS_SHEET = "Gmail_Campaigns"
GMAIL_CAMPAIGNS_HEADERS = ["Campaign ID", "Name", "Sender Email", "Channel",
                           "Created", "User", "Status", "Total Leads", "Subject"]
GMAIL_CAMPAIGN_LEADS_SHEET = "Gmail_Campaign_Leads"
GMAIL_CAMPAIGN_LEADS_HEADERS = ["Campaign ID", "Lead ID", "Full Name", "Email",
                                "Company", "Subject", "Message", "Status",
                                "Scheduled DateTime", "Sent DateTime"]
GMAIL_FOLLOWUPS_SHEET = "Gmail_Follow_Ups"
GMAIL_FOLLOWUPS_HEADERS = ["Timestamp", "Lead ID", "Full Name", "Email", "Step",
                           "Subject", "Message", "Status", "Scheduled DateTime",
                           "Note"]
# Opciones rápidas de reprogramación (etiqueta -> horas; None = personalizada).
QUICK_RESCHEDULE = [("Mañana", 24), ("En 24 horas", 24), ("En 48 horas", 48),
                    ("En 72 horas", 72), ("En 7 días", 168),
                    ("Fecha personalizada", None)]

# Estados de la cola LinkedIn asistida (Scheduled_Messages, canal LinkedIn).
LINKEDIN_QUEUE_STATUS = ["Draft", "Ready to send", "Manually sent",
                         "Responded", "Follow Up needed"]

# Palabras clave para clasificar respuestas de Gmail.
REPLY_KEYWORDS = {
    "bounce": ["mailer-daemon", "postmaster", "delivery status", "undeliver",
               "delivery failed", "returned mail", "delivery incomplete"],
    "blacklist": ["unsubscribe", "no contactar", "remove me", "stop emailing",
                  "baja de la lista", "do not contact"],
    "not_interested": ["not interested", "no me interesa", "no estamos interesados",
                       "no gracias", "not a fit", "no necesitamos"],
    "rfq": ["cotiza", "quote", "rfq", "pricing", "precio", "presupuesto"],
    "meeting": ["reunion", "reunión", "meeting", "call", "llamada", "agenda", "demo"],
    "later": ["mas adelante", "más adelante", "later", "next quarter", "en unos meses",
              "contactar despues", "contactar después", "q3", "q4"],
}

# Reglas de color de fila por Outcome Status (existentes + nuevas).
# Las 4 primeras YA existen en el archivo; se conservan idénticas.
ROW_FILLS = {
    "Blacklist":          "F4CCCC",  # rojo claro  (existente)
    "Prospectar Después": "FFF2CC",  # amarillo    (existente)
    "Won":                "D9EAD3",  # verde       (existente)
    "Lost":               "EADCF8",  # morado      (existente, opción C)
    "Respondió":          "CFE2F3",  # azul claro  (NUEVO)
    "Meeting":            "EAF6E9",  # verde claro tenue (NUEVO)
    "RFQ":                "EAF6E9",  # verde claro tenue (NUEVO)
    "Quote":              "EAF6E9",  # verde claro tenue (NUEVO)
}

RESPONSE_RESULTS = [
    "Won", "Lost", "Respondió interesado", "Respondió no interesado",
    "Prospectar después", "Blacklist", "Reunión agendada", "Cotización solicitada",
]

RESPONSE_CLASSES = [
    "Interesado", "No interesado", "Pide más información", "Pide cotización",
    "Quiere reunión", "No contactar", "Otro",
]


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ó", "o").replace("í", "i").replace("á", "a").replace("é", "e").replace("ú", "u")
    return re.sub(r"[^a-z0-9 ]+", "", s)


# --------------------------------------------------------------------------- #
# Mapeo de columnas (nombre difuso + respaldo por posición)
# --------------------------------------------------------------------------- #

class ColumnMap:
    """Mapea nombre canónico -> índice de columna (1-based) de una hoja."""
    def __init__(self, header_row):
        norm_to_canon = {}
        for canon, syns in CANON_SYNONYMS.items():
            for s in syns:
                norm_to_canon[_norm(s)] = canon
        self.canon_to_idx = {}
        for idx, raw in enumerate(header_row, 1):
            if raw is None:
                continue
            canon = norm_to_canon.get(_norm(str(raw)))
            if canon and canon not in self.canon_to_idx:
                self.canon_to_idx[canon] = idx

    def idx(self, canon):
        return self.canon_to_idx.get(canon)

    def letter(self, canon):
        i = self.idx(canon)
        return get_column_letter(i) if i else None


# --------------------------------------------------------------------------- #
# Sidecar de estado de la app (lo que el Excel no puede guardar)
# --------------------------------------------------------------------------- #

@dataclass
class LeadState:
    responded: bool = False              # respondió por algún canal (pausa todo)
    responded_channel: str | None = None
    responded_at: str | None = None
    last_message: str | None = None      # último mensaje recibido del lead
    recontact_date: str | None = None    # YYYY-MM-DD, SIEMPRE la da el usuario
    scheduled: dict = field(default_factory=dict)  # paso_canonico -> ISO datetime programado
    log: list = field(default_factory=list)        # historial de actividad


class StateStore:
    """Persistencia JSON del estado lateral, indexado por llave de lead."""
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self._d: dict[str, dict] = {}
        if self.path.exists():
            self._d = json.loads(self.path.read_text(encoding="utf-8"))

    @staticmethod
    def key(linkedin=None, email=None, name=None, company=None, phone=None) -> str:
        # Prioridad de identidad: LinkedIn > email > nombre+empresa > teléfono.
        if linkedin:
            return "li:" + re.sub(r"/+$", "", str(linkedin).strip().lower())
        if email:
            return "em:" + str(email).strip().lower()
        if name or company:
            return "nm:" + _norm(name or "") + "|" + _norm(company or "")
        if phone:
            return "ph:" + re.sub(r"\D", "", str(phone))
        return "nm:|"  # sin identidad: marcador genérico

    def get(self, k) -> LeadState:
        d = self._d.get(k)
        return LeadState(**d) if d else LeadState()

    def put(self, k, st: LeadState):
        self._d[k] = asdict(st)

    def add_log(self, k, channel, step, action, result="", notes="", message="",
                ts=None):
        st = self.get(k)
        st.log.append({
            "ts": (ts or dt.datetime.now()).isoformat(timespec="seconds")
                  if not isinstance(ts, str) else ts,
            "channel": channel, "step": step, "action": action,
            "result": result, "notes": notes, "message": message,
        })
        self.put(k, st)

    def save(self):
        self.path.write_text(json.dumps(self._d, ensure_ascii=False, indent=2), encoding="utf-8")


# --------------------------------------------------------------------------- #
# Modelo de lead (vista de una fila)
# --------------------------------------------------------------------------- #

@dataclass
class Lead:
    sheet: str
    row: int
    full_name: str = ""
    company: str = ""
    job_title: str = ""
    seniority_level: str = ""
    seniority_rank: float = 999
    lead_score: str = ""
    opportunity_score: float = 0
    linkedin: str = ""
    email: str = ""
    phone: str = ""
    industry: str = ""
    outcome: str = ""           # valor de Outcome Status en el Excel
    stage_values: dict = field(default_factory=dict)  # paso_canonico -> valor de celda

    @property
    def key(self):
        return StateStore.key(self.linkedin, self.email, self.full_name,
                              self.company, self.phone)


# --------------------------------------------------------------------------- #
# CRM principal
# --------------------------------------------------------------------------- #

class CRM:
    def __init__(self, xlsx_path: str | Path, state_path: str | Path = None):
        self.xlsx_path = Path(xlsx_path)
        self.state = StateStore(state_path or self.xlsx_path.with_suffix(".state.json"))
        self.wb = load_workbook(self.xlsx_path)
        self.maps: dict[str, ColumnMap] = {}
        for s in self.wb.sheetnames:
            if s in NON_DATA_SHEETS:
                continue
            ws = self.wb[s]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            self.maps[s] = ColumnMap(header)

    def industries(self):
        """Hojas de datos (leads) presentes en el libro, en orden."""
        return [s for s in self.wb.sheetnames if s not in NON_DATA_SHEETS]

    # ---- lectura -------------------------------------------------------- #

    def _cell(self, ws, cmap, row, canon):
        i = cmap.idx(canon)
        return ws.cell(row=row, column=i).value if i else None

    def read_lead(self, sheet, row) -> Lead:
        ws, cmap = self.wb[sheet], self.maps[sheet]
        g = lambda c: self._cell(ws, cmap, row, c)
        lead = Lead(
            sheet=sheet, row=row,
            full_name=g("Full Name") or "", company=g("Company Name") or "",
            job_title=g("Job Title") or "", seniority_level=g("Seniority Level") or "",
            seniority_rank=_num(g("Seniority Rank"), 999),
            lead_score=str(g("Lead Score") or ""),
            opportunity_score=_num(g("Opportunity Score"), 0),
            linkedin=g("LinkedIn Profile") or "", email=g("Email/Gmail") or "",
            phone=str(g("Phone") or ""), industry=g("Industry") or sheet,
            outcome=str(g("Outcome Status") or "").strip(),
        )
        for steps in CHANNEL_STEPS.values():
            for st in steps:
                lead.stage_values[st] = g(st)
        for st in STAGE_COLS:
            lead.stage_values[st] = g(st)
        return lead

    _ID_COLS = ("Full Name", "Company Name", "Email/Gmail", "LinkedIn Profile",
                "Phone", "Job Title")

    def _row_has_lead(self, ws, cmap, row) -> bool:
        """True si la fila tiene algún dato de identidad (nombre/empresa/email/
        LinkedIn/teléfono). Permite leads parciales sin nombre."""
        for c in self._ID_COLS:
            i = cmap.idx(c)
            if i and ws.cell(row=row, column=i).value not in (None, ""):
                return True
        return False

    def all_leads(self, sheet) -> list[Lead]:
        ws, cmap = self.wb[sheet], self.maps[sheet]
        out = []
        for row in range(2, ws.max_row + 1):
            if not self._row_has_lead(ws, cmap, row):
                continue
            out.append(self.read_lead(sheet, row))
        return out

    # ---- filtros y orden ------------------------------------------------ #

    def _today(self):
        return dt.date.today()

    def _is_blocked(self, lead: Lead, st: LeadState) -> str | None:
        """Devuelve motivo de bloqueo/pausa, o None si el lead puede recibir envíos."""
        if lead.outcome in OUTCOME_BLOCK:
            return lead.outcome
        if st.responded or lead.outcome == "Respondió":
            return "Respondió"
        if lead.outcome == "Prospectar Después":
            if st.recontact_date and self._today() < dt.date.fromisoformat(st.recontact_date):
                return "Prospectar Después (antes de fecha)"
            if not st.recontact_date:
                return "Prospectar Después (sin fecha)"
        return None

    def _sort_key(self, lead: Lead):
        return (
            lead.seniority_rank,
            LEAD_SCORE_PRIORITY.get(lead.lead_score, 99),
            _norm(lead.company),
            _norm(lead.full_name),
        )

    def build_queue(self, sheet, channel="LinkedIn", step_index=0, limit=30) -> list[Lead]:
        """Cola ordenada de leads elegibles para el paso indicado de un canal."""
        step = CHANNEL_STEPS[channel][step_index]
        elig = []
        for lead in self.all_leads(sheet):
            st = self.state.get(lead.key)
            if self._is_blocked(lead, st):
                continue
            # excluir si el paso ya fue enviado/realizado
            if _is_marked(lead.stage_values.get(step)):
                continue
            # para pasos > 0, exigir que el paso anterior ya esté hecho
            if step_index > 0:
                prev = CHANNEL_STEPS[channel][step_index - 1]
                if not _is_marked(lead.stage_values.get(prev)):
                    continue
            elig.append(lead)
        elig.sort(key=self._sort_key)
        return elig[:limit]

    # ---- escritura (preserva formato y fórmulas) ------------------------ #

    def _set(self, lead: Lead, canon, value):
        ws, cmap = self.wb[lead.sheet], self.maps[lead.sheet]
        i = cmap.idx(canon)
        if i:
            ws.cell(row=lead.row, column=i).value = value

    def mark_sent(self, lead: Lead, channel, step, message="", when=None, result=None):
        """Registra un envío aprobado: marca etapa, sella Last Contact, loguea
        (sidecar + hoja Activity_Log), y programa el siguiente paso del canal.
        'channel' puede ser LinkedIn/Email/Cold Call/Llamada/WhatsApp."""
        when = when or dt.datetime.now()
        mark = DONE if channel in ("Cold Call", "Llamada") else SENT
        self._set(lead, step, mark)
        self._set(lead, "Last Contact", when.date().isoformat())
        if not self._cell(self.wb[lead.sheet], self.maps[lead.sheet], lead.row, "First Contact"):
            self._set(lead, "First Contact", when.date().isoformat())
        res = result or mark
        self.state.add_log(lead.key, channel, step, "sent", result=res,
                           message=message, ts=when)
        self.append_activity(lead, channel, step, message, res, when)
        return mark

    # ---- Activity_Log (hoja nueva del Excel) --------------------------- #

    def _ensure_activity_sheet(self):
        if ACTIVITY_LOG_SHEET not in self.wb.sheetnames:
            ws = self.wb.create_sheet(ACTIVITY_LOG_SHEET)
            ws.append(ACTIVITY_HEADERS)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            ws.freeze_panes = "A2"
        return self.wb[ACTIVITY_LOG_SHEET]

    @staticmethod
    def _fu_number(step):
        m = re.search(r"(\d+)", step or "")
        return int(m.group(1)) if m else ""

    def append_activity(self, lead, channel, step, message, result, when=None):
        """Agrega una fila al historial en la hoja Activity_Log del Excel."""
        when = when or dt.datetime.now()
        ws = self._ensure_activity_sheet()
        ws.append([
            when.isoformat(timespec="seconds"), lead.full_name, lead.company,
            lead.industry, when.date().isoformat(), channel,
            self._fu_number(step), step, message, result,
        ])

    # ---- Notifications (hoja del Excel) -------------------------------- #

    def _ensure_notifications_sheet(self):
        if NOTIFICATIONS_SHEET not in self.wb.sheetnames:
            ws = self.wb.create_sheet(NOTIFICATIONS_SHEET)
            ws.append(NOTIFICATION_HEADERS)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            ws.freeze_panes = "A2"
        return self.wb[NOTIFICATIONS_SHEET]

    def add_notification(self, lead, event_type, summary="", channel="-",
                         status=None, when=None):
        """Crea una notificación (basada en un registro manual de la app) y la
        guarda en la hoja Notifications. No conecta con LinkedIn/Gmail/WhatsApp."""
        when = when or dt.datetime.now()
        ws = self._ensure_notifications_sheet()
        status = status if status is not None else self.lead_status(lead)
        ws.append([
            when.isoformat(timespec="seconds"), when.date().isoformat(),
            lead.full_name, lead.company, channel, event_type, summary,
            status, "Pendiente",
        ])
        return ws.max_row  # fila de la notificación

    def read_notifications(self):
        """Devuelve las notificaciones como lista de dicts (incluye la fila)."""
        if NOTIFICATIONS_SHEET not in self.wb.sheetnames:
            return []
        ws = self.wb[NOTIFICATIONS_SHEET]
        out = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, len(NOTIFICATION_HEADERS) + 1)]
            if not any(v not in (None, "") for v in vals):
                continue
            d = dict(zip(NOTIFICATION_HEADERS, vals))
            d["_row"] = r
            out.append(d)
        return out

    def set_notification_seen(self, row, seen=True):
        """Marca una notificación como 'Vista' o 'Pendiente'."""
        if NOTIFICATIONS_SHEET not in self.wb.sheetnames:
            return
        ws = self.wb[NOTIFICATIONS_SHEET]
        col = NOTIFICATION_HEADERS.index("Visto") + 1
        ws.cell(row=row, column=col).value = "Vista" if seen else "Pendiente"

    def scan_due_followups(self):
        """Genera notificaciones para follow-ups programados vencidos hoy o antes
        (no las duplica en la misma corrida). Retorna cuántas creó."""
        today = dt.date.today()
        existing = {(n["Lead"], n["Mensaje/Resumen"]) for n in self.read_notifications()
                    if n.get("Tipo de evento") == "Follow-up debido"}
        created = 0
        for sheet in self.maps:
            for lead in self.all_leads(sheet):
                st = self.state.get(lead.key)
                if self._is_blocked(lead, st):
                    continue
                for step, iso in st.scheduled.items():
                    d = dt.datetime.fromisoformat(iso).date()
                    if d <= today:
                        summary = f"{lead.full_name} debe recibir {step} hoy"
                        if (lead.full_name, summary) in existing:
                            continue
                        ch = next((c for c, steps in CHANNEL_STEPS.items()
                                   if step in steps), "-")
                        self.add_notification(lead, "Follow-up debido", summary, ch)
                        created += 1
        return created

    # ---- Scheduled_Messages (agenda manual; nunca se envía solo) ------- #

    def _ensure_scheduled_sheet(self):
        if SCHEDULED_SHEET not in self.wb.sheetnames:
            ws = self.wb.create_sheet(SCHEDULED_SHEET)
            ws.append(SCHEDULED_HEADERS)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            ws.freeze_panes = "A2"
        return self.wb[SCHEDULED_SHEET]

    def _next_schedule_id(self, ws):
        return f"SCH-{ws.max_row:05d}" if ws.max_row > 1 else "SCH-00001"

    def schedule_message(self, lead, channel, message, when_iso, follow_up_step="",
                         status="Scheduled", notes=""):
        """Agrega un mensaje a la agenda. NO lo envía; solo lo deja listo para que
        el usuario lo revise y confirme manualmente. Devuelve el Schedule ID."""
        ws = self._ensure_scheduled_sheet()
        sid = self._next_schedule_id(ws)
        ws.append([
            sid, lead.key, lead.full_name, lead.company, lead.seniority_level,
            lead.industry, channel, message, when_iso, status, "",
            follow_up_step, notes,
        ])
        self.state.add_log(lead.key, channel, follow_up_step or "-", "scheduled",
                           result=status, message=message)
        return sid

    def read_scheduled(self, status=None):
        if SCHEDULED_SHEET not in self.wb.sheetnames:
            return []
        ws = self.wb[SCHEDULED_SHEET]
        out = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, len(SCHEDULED_HEADERS) + 1)]
            if not any(v not in (None, "") for v in vals):
                continue
            d = dict(zip(SCHEDULED_HEADERS, vals))
            d["_row"] = r
            if status and d.get("Status") != status:
                continue
            out.append(d)
        return out

    def update_schedule_status(self, row, status, when=None):
        """Cambia el Status de una programación. Si es 'Sent', sella Sent DateTime
        y marca la etapa del lead como enviada (registro manual del usuario)."""
        ws = self.wb[SCHEDULED_SHEET]
        sc = SCHEDULED_HEADERS.index("Status") + 1
        ws.cell(row=row, column=sc).value = status
        if status == "Sent":
            when = when or dt.datetime.now()
            ws.cell(row=row, column=SCHEDULED_HEADERS.index("Sent DateTime") + 1).value = \
                when.isoformat(timespec="seconds")
            # reflejar en el lead: marcar la etapa correspondiente
            lead_id = ws.cell(row=row, column=2).value
            channel = ws.cell(row=row, column=7).value
            step = ws.cell(row=row, column=12).value
            msg = ws.cell(row=row, column=8).value
            lead = self._find_lead_by_key(lead_id)
            if lead and step:
                self.mark_sent(lead, channel or "LinkedIn", step, message=msg or "",
                               when=when)
        return status

    def _find_lead_by_key(self, key):
        for sheet in self.maps:
            for lead in self.all_leads(sheet):
                if lead.key == key:
                    return lead
        return None

    def schedule_next(self, lead: Lead, channel, step_index, gap_days):
        """Programa el siguiente paso del canal a gap_days, si existe."""
        steps = CHANNEL_STEPS[channel]
        if step_index + 1 >= len(steps):
            return None
        nxt = steps[step_index + 1]
        when = dt.datetime.now() + dt.timedelta(days=gap_days)
        st = self.state.get(lead.key)
        st.scheduled[nxt] = when.isoformat(timespec="seconds")
        self.state.put(lead.key, st)
        self._set(lead, "Next Follow-up", when.date().isoformat())
        return nxt, when

    def set_outcome(self, lead: Lead, value, recontact_date=None, reason=None, note=None):
        """Cambia Outcome Status (lo que dispara el color de fila). El usuario es
        quien decide Won/Lost/Blacklist/Prospectar Después/Respondió/etc."""
        assert value in OUTCOME_VALUES, f"Outcome inválido: {value}"
        self._set(lead, "Outcome Status", value)
        st = self.state.get(lead.key)
        if value in ("Won", "Lost", "Blacklist", "Respondió") or value == "Prospectar Después":
            # estos pausan/cierran: limpiar programaciones pendientes
            st.scheduled = {}
        if value == "Prospectar Después":
            st.recontact_date = recontact_date  # SIEMPRE provista por el usuario
        if reason:
            self._set(lead, "Rejection Reason", reason)
        if note:
            self._set(lead, "Notes", note)
        self.state.put(lead.key, st)
        self.state.add_log(lead.key, "-", "Outcome Status", "set", result=value,
                           notes=note or "")

    def mark_stage(self, lead: Lead, stage, tint_row=True, value="Hecho"):
        """Marca Meeting/RFQ/Quote (columna de etapa -> cuenta el dashboard) y,
        opcionalmente, escribe Outcome Status para teñir la fila de verde claro
        sin romper los conteos (el dashboard lee la etapa por posición, no AH)."""
        assert stage in STAGE_COLS
        self._set(lead, stage, value)
        if tint_row and lead.outcome not in OUTCOME_BLOCK:
            self._set(lead, "Outcome Status", stage)
        self.state.add_log(lead.key, "-", stage, "stage", result=value)

    # ---- detección de respuestas --------------------------------------- #

    def find_lead(self, linkedin=None, email=None, name=None, company=None):
        """Localiza el lead que respondió por LinkedIn/email/nombre+empresa."""
        for sheet in self.maps:
            for lead in self.all_leads(sheet):
                if linkedin and _norm(lead.linkedin) == _norm(linkedin):
                    return lead
                if email and _norm(lead.email) == _norm(email) and email:
                    return lead
                if name and company and _norm(lead.full_name) == _norm(name) \
                        and _norm(lead.company) == _norm(company):
                    return lead
        return None

    def register_response(self, lead: Lead, channel, message=""):
        """Un lead respondió: pausa TODOS los canales, marca Respondió + azul,
        y prepara la notificación. NO clasifica (eso lo hace el usuario)."""
        st = self.state.get(lead.key)
        st.responded = True
        st.responded_channel = channel
        st.responded_at = dt.datetime.now().isoformat(timespec="seconds")
        st.last_message = message
        st.scheduled = {}                       # detener todo lo programado
        self.state.put(lead.key, st)
        self._set(lead, "Outcome Status", "Respondió")   # "ambas": notifica y escribe
        self._set(lead, "Last Contact", dt.date.today().isoformat())
        self.state.add_log(lead.key, channel, "-", "response", message=message)
        return self.build_notification(lead, channel, message)

    def build_notification(self, lead: Lead, channel, message):
        return {
            "lead": lead.full_name, "company": lead.company,
            "industry": lead.industry, "job_title": lead.job_title,
            "channel": channel, "message": message,
            "suggested_action": self._suggest_action(message),
            "classes": RESPONSE_CLASSES,
        }

    @staticmethod
    def _suggest_action(message: str) -> str:
        m = _norm(message)
        if any(w in m for w in ["no contact", "remove", "unsubscribe", "stop", "baja", "no escrib"]):
            return "Clasificar: No contactar -> Blacklist"
        if any(w in m for w in ["cotiza", "quote", "precio", "pricing", "rfq"]):
            return "Clasificar: Pide cotización -> RFQ"
        if any(w in m for w in ["reunion", "meeting", "call", "agenda", "demo", "llamada"]):
            return "Clasificar: Quiere reunión -> Meeting"
        if any(w in m for w in ["mas informacion", "info", "detalle", "more information"]):
            return "Clasificar: Pide más información"
        return "Revisar y clasificar respuesta"

    def classify_response(self, lead: Lead, cls, recontact_date=None, reason=None, note=None):
        """Aplica la decisión del usuario sobre una respuesta clasificada."""
        if cls == "No contactar":
            self.set_outcome(lead, "Blacklist", reason=reason, note=note)
        elif cls == "Pide cotización":
            self.mark_stage(lead, "RFQ")
        elif cls == "Quiere reunión":
            self.mark_stage(lead, "Meeting")
        elif cls == "No interesado":
            self.set_outcome(lead, "Lost", reason=reason, note=note)
        elif cls in ("Interesado", "Pide más información", "Otro"):
            # no terminal: queda en pausa (Respondió) hasta nueva instrucción
            if note:
                self._set(lead, "Notes", note)
            if reason:
                self._set(lead, "Interest Reason", reason)
            self.state.add_log(lead.key, "-", "-", "classify", result=cls, notes=note or "")
        else:
            raise ValueError(f"Clase desconocida: {cls}")

    # ---- escalamiento de canal ----------------------------------------- #

    def escalation_suggestion(self, lead: Lead) -> str | None:
        """Sugiere (no ejecuta) el siguiente canal según las reglas del usuario."""
        sv = lead.stage_values
        if _is_marked(sv.get("Cold Call 3")):
            return "Sin respuesta tras Cold Call 3 -> sugerir Lost o Prospectar Después"
        if _is_marked(sv.get("Email 2")) and not _is_marked(sv.get("Cold Call 1")):
            return "Sin respuesta tras Email 2 -> sugerir Cold Call 1"
        if _is_marked(sv.get("Follow Up 3")) and not _is_marked(sv.get("Email 1")):
            return "Sin respuesta tras Follow Up 3 -> sugerir Email 1"
        return None

    # ---- importación de leads ------------------------------------------ #

    def _all_keys(self) -> set:
        """Conjunto de llaves de todos los leads existentes (para deduplicar)."""
        keys = set()
        for sheet in self.maps:
            for lead in self.all_leads(sheet):
                keys.add(lead.key)
        return keys

    def _next_row(self, sheet) -> int:
        """Primera fila libre real: barre toda la fila (cualquier celda con dato)
        para nunca pisar leads parciales, sea cual sea el campo que tengan."""
        ws = self.wb[sheet]
        r = ws.max_row
        while r >= 2 and all(ws.cell(row=r, column=c).value in (None, "")
                             for c in range(1, ws.max_column + 1)):
            r -= 1
        return r + 1

    def _match_industry_sheet(self, value) -> str | None:
        """Mapea un texto de industria a una de las 9 hojas (difuso)."""
        v = _norm(value or "")
        if not v:
            return None
        for sheet in INDUSTRIES:
            sv = _norm(sheet)
            if v == sv or v in sv or sv in v:
                return sheet
        aliases = {"software": "Software IT", "it": "Software IT", "tech": "Software IT",
                   "pharma": "Pharmaceutical", "medical": "Medical Devices",
                   "hvac": "HVAC Appliances", "auto": "Automotive"}
        for a, sheet in aliases.items():
            if a in v:
                return sheet
        return None

    @staticmethod
    def infer_seniority(title):
        """Devuelve (Seniority Level, Seniority Rank) inferido del Job Title."""
        t = _norm(title or "")
        for level, rank, kws in SENIORITY_RULES:
            if any(k in t for k in kws):
                return level, rank
        return "Other", 9

    def detect_import_mapping(self, columns) -> dict:
        """Auto-detecta canon -> nombre de columna del archivo importado.
        Recorre los campos en orden de prioridad para resolver sinónimos
        ambiguos (p. ej. 'nombre' favorece Full Name antes que First Name)."""
        mapping = {}
        used = set()
        for canon in IMPORT_FIELDS:  # orden = prioridad (Full Name primero)
            syns = {_norm(s) for s in CANON_SYNONYMS[canon]}
            for col in columns:
                if col in used:
                    continue
                if _norm(str(col)) in syns:
                    mapping[canon] = col
                    used.add(col)
                    break
        return mapping  # canon -> source column

    def import_leads(self, df, mapping, route_by_industry=True, default_industry="Other"):
        """Agrega leads del DataFrame al Excel principal, respetando columnas,
        deduplicando (LinkedIn / email / nombre+empresa) y tolerando faltantes.
        Devuelve (agregados, duplicados, detalle_por_hoja)."""
        existing = self._all_keys()
        batch = set()
        added = 0
        dups = 0
        per_sheet = {}

        def cell(row, canon):
            col = mapping.get(canon)
            if not col or col not in df.columns:
                return ""
            v = row.get(col)
            try:
                import pandas as pd
                if pd.isna(v):
                    return ""
            except Exception:
                pass
            return str(v).strip()

        for _, r in df.iterrows():
            full = cell(r, "Full Name")
            first = cell(r, "First Name")
            last = cell(r, "Last Name")
            if not full:
                full = (first + " " + last).strip()
            company = cell(r, "Company Name")
            linkedin = cell(r, "LinkedIn Profile")
            email = cell(r, "Email/Gmail")
            phone = cell(r, "Phone")
            title = cell(r, "Job Title")
            location = cell(r, "Location")
            ind_text = cell(r, "Industry")

            # Agregar si hay CUALQUIER dato útil; solo se omite la fila vacía total.
            any_value = any([full, first, last, company, linkedin, email, phone,
                             title, location, ind_text])
            if not any_value:
                continue

            # Deduplicar solo cuando hay alguna señal de identidad. Si el lead no
            # tiene identidad alguna, se agrega siempre (clave sintética única).
            has_identity = bool(linkedin or email or full or company or phone)
            key = StateStore.key(linkedin or None, email or None, full or None,
                                 company or None, phone or None)
            if has_identity:
                if key in existing or key in batch:
                    dups += 1
                    continue
            else:
                key = f"__row{added}__"  # único en el lote

            sheet = (self._match_industry_sheet(ind_text) if route_by_industry else None) \
                or (default_industry if default_industry in self.maps
                    else ("Other" if "Other" in self.maps
                          else next(iter(self.maps))))

            # Crear columnas nuevas necesarias (VP/Pain Point + las de clasificación).
            need_extra = CLASSIFY_COLUMNS + [c for c in ("Value Proposition", "Pain Point")
                                             if mapping.get(c)]
            self.ensure_columns(sheet, need_extra)

            row_i = self._next_row(sheet)
            ws, cmap = self.wb[sheet], self.maps[sheet]

            def put(canon, val):
                i = cmap.idx(canon)
                if i and val not in (None, ""):
                    ws.cell(row=row_i, column=i).value = val

            # Solo se llenan los espacios posibles; lo que falte queda vacío.
            put("Full Name", full)
            put("First Name", first or (full.split(" ")[0] if full else ""))
            put("Last Name", last or (" ".join(full.split(" ")[1:]) if full else ""))
            put("Company Name", company)
            put("Job Title", title)
            put("Location", location)
            put("LinkedIn Profile", linkedin)
            put("Email/Gmail", email)
            put("Phone", phone)
            put("Industry", ind_text or sheet)
            put("Notes", cell(r, "Notes"))
            put("Value Proposition", cell(r, "Value Proposition"))
            put("Pain Point", cell(r, "Pain Point"))
            # Seniority: respeta el del archivo si vino; si no, se infiere del título.
            sen_file = cell(r, "Seniority Level")
            eff_title = title
            if sen_file:
                put("Seniority Level", sen_file)
                _, rank = self.infer_seniority(sen_file)
                put("Seniority Rank", rank)
                eff_title = eff_title or sen_file
            elif title:
                level, rank = self.infer_seniority(title)
                put("Seniority Level", level)
                put("Seniority Rank", rank)
            put("Outcome Status", "Active")

            # Clasificación automática (respeta lo que venga del archivo si vino).
            tmp = Lead(sheet=sheet, row=row_i, full_name=full, company=company,
                       job_title=eff_title, seniority_level=(sen_file or ""),
                       linkedin=linkedin, email=email, phone=phone,
                       industry=ind_text or sheet)
            put("Priority", cell(r, "Priority") or self.lead_priority(tmp))
            put("Lead Warmth Level", cell(r, "Lead Warmth Level") or "Frío")
            put("Recommended Channel", cell(r, "Recommended Channel")
                or self.recommend_channel(tmp))
            put("Current Stage", cell(r, "Current Stage") or "Nuevo lead")

            batch.add(key)
            added += 1
            per_sheet[sheet] = per_sheet.get(sheet, 0) + 1

        self.reanchor_all()  # extender colores y validaciones a las filas nuevas
        return added, dups, per_sheet

    # ---- formato condicional + validaciones (re-anclaje seguro) -------- #

    def reanchor_all(self):
        for sheet in self.maps:
            self._reanchor_sheet(sheet)

    def _reanchor_sheet(self, sheet):
        """Reconstruye las reglas de color y extiende validaciones hasta la última
        fila, conservando las reglas originales y añadiendo las 4 nuevas. Es la
        forma segura: mutar el sqref en sitio corrompe el índice interno."""
        ws, cmap = self.wb[sheet], self.maps[sheet]
        ah = cmap.letter("Outcome Status")
        last_col = get_column_letter(ws.max_column)
        nm = ws.max_row
        if nm < 2 or ah is None:
            return   # hoja sin filas de datos (p. ej. CRM recién creado)
        full = f"A2:{last_col}{nm}"

        # 1) recolectar reglas existentes y sus fórmulas
        rules, formulas = [], set()
        for cf in ws.conditional_formatting:
            for rule in ws.conditional_formatting[cf]:
                rules.append(rule)
                formulas.add(str(rule.formula))

        # 2) reconstruir la lista sobre el rango completo
        ws.conditional_formatting = ConditionalFormattingList()
        for rule in rules:
            ws.conditional_formatting.add(full, rule)

        # 3) sembrar reglas faltantes (las 8: también repara libros nuevos)
        for status in ROW_FILLS:
            f = f'${ah}2="{status}"'
            if any(status in fm for fm in formulas):
                continue
            fill = PatternFill("solid", fgColor=ROW_FILLS[status])
            ws.conditional_formatting.add(
                full, FormulaRule(formula=[f], fill=fill, stopIfTrue=False))

        # 4) extender todas las validaciones a la última fila
        for dv in ws.data_validations.dataValidation:
            rebuilt = [f"{get_column_letter(cr.min_col)}2:{get_column_letter(cr.min_col)}{nm}"
                       for cr in dv.sqref.ranges]
            dv.sqref = " ".join(rebuilt)
        self._extend_outcome_validation(ws, cmap)

    # alias retrocompatible
    def ensure_formatting(self):
        self.reanchor_all()

    def _extend_outcome_validation(self, ws, cmap):
        ah_letter = cmap.letter("Outcome Status")
        if not ah_letter:
            return
        target = f"{ah_letter}2:{ah_letter}{ws.max_row}"
        for dv in list(ws.data_validations.dataValidation):
            if dv.type == "list" and f"{ah_letter}2" in str(dv.sqref):
                ws.data_validations.dataValidation.remove(dv)
        joined = ",".join(OUTCOME_VALUES)
        dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
        dv.add(target)
        ws.data_validations.dataValidation.append(dv)

    # ---- reporte diario ------------------------------------------------- #

    def _read_activity_log(self):
        """Lee las filas de la hoja Activity_Log del CRM (historial persistente)."""
        if ACTIVITY_LOG_SHEET not in self.wb.sheetnames:
            return []
        ws = self.wb[ACTIVITY_LOG_SHEET]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and any(c is not None for c in r):
                rows.append(list(r))
        return rows

    def build_daily_report_data(self, when=None):
        """Construye los datos del reporte del día como hojas ordenadas:
        {nombre_hoja: {"headers": [...], "rows": [[...]]}}. Una sola pasada,
        reutilizable para Excel o CSV. NO modifica el CRM."""
        when = when or dt.datetime.now()
        today = when.date()
        today_s = today.isoformat()
        tomorrow = today + dt.timedelta(days=1)

        leads_rows, fu_rows, email_rows = [], [], []
        responded_rows, blacklist_rows, pending_rows = [], [], []

        total_leads = 0
        fu_steps = CHANNEL_STEPS["LinkedIn"]
        email_steps = CHANNEL_STEPS["Email"]

        for sheet_name in self.maps:
            for lead in self.all_leads(sheet_name):
                total_leads += 1
                st = self.state.get(lead.key)
                last_contact = self._cell(self.wb[sheet_name], self.maps[sheet_name],
                                          lead.row, "Last Contact")
                status = self.lead_status(lead, st)
                leads_rows.append([lead.industry, lead.full_name, lead.company,
                                   lead.job_title, lead.seniority_level,
                                   lead.lead_score, lead.outcome or "Active",
                                   str(last_contact or ""), status])

                # Follow_Up: leads con algún Follow Up enviado
                sent_fu = [s for s in fu_steps if _is_marked(lead.stage_values.get(s))]
                if sent_fu:
                    fu_rows.append([lead.industry, lead.full_name, lead.company,
                                    lead.job_title, sent_fu[-1], ", ".join(sent_fu),
                                    str(last_contact or ""), status])
                # Emails_Enviados: leads con algún Email enviado
                sent_em = [s for s in email_steps if _is_marked(lead.stage_values.get(s))]
                if sent_em:
                    email_rows.append([lead.industry, lead.full_name, lead.company,
                                       lead.email, sent_em[-1], ", ".join(sent_em),
                                       str(last_contact or "")])
                # Respondieron
                if st.responded or lead.outcome == "Respondió":
                    responded_rows.append([lead.industry, lead.full_name, lead.company,
                                           st.responded_channel or "",
                                           (st.last_message or "")[:200],
                                           (st.responded_at or "")[:10]])
                # Blacklist
                if lead.outcome == "Blacklist":
                    reason = self._cell(self.wb[sheet_name], self.maps[sheet_name],
                                        lead.row, "Rejection Reason")
                    blacklist_rows.append([lead.industry, lead.full_name, lead.company,
                                           lead.job_title, str(reason or "")])
                # Pendientes_Mañana
                for step, iso in st.scheduled.items():
                    d = dt.datetime.fromisoformat(iso).date()
                    if d <= tomorrow and not self._is_blocked(lead, st):
                        ch = next((c for c, steps in CHANNEL_STEPS.items()
                                   if step in steps), "-")
                        pending_rows.append([lead.industry, lead.full_name, lead.company,
                                             ch, step, d.isoformat()])

        # Activity_Log (historial persistente) y métricas del día
        act_rows = self._read_activity_log()
        # índices: 4=Fecha, 5=Canal, 6=FU#, 7=Paso, 1=Persona, 2=Empresa
        today_acts = [r for r in act_rows if str(r[4]) == today_s]
        contacted_today = {(r[1], r[2]) for r in today_acts}
        fu_today = [r for r in today_acts if str(r[7] or "").startswith("Follow Up")]
        email_today = [r for r in today_acts if str(r[7] or "").startswith("Email")]
        fu_today_by_step = {}
        for r in fu_today:
            fu_today_by_step[r[7]] = fu_today_by_step.get(r[7], 0) + 1
        email_today_by_step = {}
        for r in email_today:
            email_today_by_step[r[7]] = email_today_by_step.get(r[7], 0) + 1

        # Resumen_Dia (etiqueta, valor)
        resumen = [
            ["Fecha del reporte", today_s],
            ["Total de leads cargados", total_leads],
            ["Total contactados hoy", len(contacted_today)],
            ["Total Follow Ups enviados hoy", len(fu_today)],
        ]
        for s in fu_steps:
            if fu_today_by_step.get(s):
                resumen.append([f"   · {s} enviados hoy", fu_today_by_step[s]])
        resumen.append(["Total correos enviados hoy", len(email_today)])
        for s in email_steps:
            if email_today_by_step.get(s):
                resumen.append([f"   · {s} enviados hoy", email_today_by_step[s]])
        resumen += [
            ["Total no accedieron a recibir mensaje (Blacklist)", len(blacklist_rows)],
            ["Total respondieron", len(responded_rows)],
            ["Total Blacklist", len(blacklist_rows)],
            ["Total pendientes mañana", len(pending_rows)],
        ]

        return {
            "Leads_Actualizados": {
                "headers": ["Industria", "Lead", "Company", "Job Title", "Seniority",
                            "Lead Score", "Outcome Status", "Last Contact", "Estado"],
                "rows": leads_rows},
            "Activity_Log": {"headers": ACTIVITY_HEADERS, "rows": act_rows},
            "Follow_Up": {
                "headers": ["Industria", "Lead", "Company", "Job Title",
                            "Follow Up actual", "Follow Ups enviados",
                            "Último contacto", "Estado"],
                "rows": fu_rows},
            "Emails_Enviados": {
                "headers": ["Industria", "Lead", "Company", "Email",
                            "Email actual", "Emails enviados", "Último contacto"],
                "rows": email_rows},
            "Respondieron": {
                "headers": ["Industria", "Lead", "Company", "Canal", "Mensaje", "Fecha"],
                "rows": responded_rows},
            "Blacklist": {
                "headers": ["Industria", "Lead", "Company", "Job Title", "Razón"],
                "rows": blacklist_rows},
            "Pendientes_Mañana": {
                "headers": ["Industria", "Lead", "Company", "Canal", "Paso", "Programado"],
                "rows": pending_rows},
            "Resumen_Dia": {"headers": ["Métrica", "Valor"], "rows": resumen},
        }

    def export_daily_report(self, out_dir=".", when=None, fmt="xlsx"):
        """Genera el reporte del día. fmt='xlsx' -> un Excel con 8 hojas;
        fmt='csv' -> un .zip con 8 CSV (uno por hoja). NO modifica el CRM."""
        when = when or dt.datetime.now()
        today = when.date().isoformat()
        data = self.build_daily_report_data(when)

        if fmt == "csv":
            import csv, io, zipfile
            out = Path(out_dir) / f"CRM_reporte_{today}.zip"
            with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
                for name, sh in data.items():
                    buf = io.StringIO()
                    w = csv.writer(buf)
                    w.writerow(sh["headers"])
                    w.writerows(sh["rows"])
                    zf.writestr(f"{name}.csv", buf.getvalue())
            return out

        out = Path(out_dir) / f"CRM_reporte_{today}.xlsx"
        wb = Workbook()
        wb.remove(wb.active)
        for name, sh in data.items():
            ws = wb.create_sheet(name[:31])
            ws.append(sh["headers"])
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            ws.freeze_panes = "A2"
            for row in sh["rows"]:
                ws.append(row)
        wb.save(out)
        return out

    def build_dashboard_data(self, when=None):
        """Datos para el dashboard exportable ampliado: hojas por industria,
        seniority, current stage, follow-up step, canal, Won/Lost/Prospectar/
        Blacklist, Scheduled_Messages, Activity_Log, Notifications + Resumen."""
        when = when or dt.datetime.now()
        LEAD_HDR = ["Industria", "Lead", "Company", "Job Title", "Seniority",
                    "Priority", "Warmth", "Canal recomendado", "Current Stage",
                    "Outcome Status", "Último contacto"]
        sheets = {}
        by_ind, by_sen, by_stage, by_fu, by_chan = {}, {}, {}, {}, {}
        won, lost, prosp, black = [], [], [], []
        totals = dict(total=0, prospectados=0, respondieron=0, won=0, lost=0,
                      pend_fu=0)
        t_ind, t_sen, t_chan = {}, {}, {}

        def row_of(lead, st):
            lc = self._cell(self.wb[lead.sheet], self.maps[lead.sheet],
                            lead.row, "Last Contact")
            return [lead.industry, lead.full_name, lead.company, lead.job_title,
                    lead.seniority_level or "—", self.lead_priority(lead),
                    self.classify_warmth(lead, st), self.recommend_channel(lead),
                    self.lead_stage(lead, st), lead.outcome or "Active", str(lc or "")]

        for sheet_name in self.maps:
            for lead in self.all_leads(sheet_name):
                st = self.state.get(lead.key)
                row = row_of(lead, st)
                totals["total"] += 1
                stage = self.lead_stage(lead, st)
                chan = self.lead_channel(lead, st)
                sen = lead.seniority_level or "—"
                by_ind.setdefault(lead.industry, []).append(row)
                by_sen.setdefault(sen, []).append(row)
                by_stage.setdefault(stage, []).append(row)
                by_chan.setdefault(chan, []).append(row)
                t_ind[lead.industry] = t_ind.get(lead.industry, 0) + 1
                t_sen[sen] = t_sen.get(sen, 0) + 1
                t_chan[chan] = t_chan.get(chan, 0) + 1
                fu = [s for s in CHANNEL_STEPS["LinkedIn"]
                      if _is_marked(lead.stage_values.get(s))]
                if fu:
                    by_fu.setdefault(fu[-1], []).append(row)
                    totals["prospectados"] += 1
                if st.responded or lead.outcome == "Respondió":
                    totals["respondieron"] += 1
                if lead.outcome == "Won": won.append(row); totals["won"] += 1
                if lead.outcome == "Lost": lost.append(row); totals["lost"] += 1
                if lead.outcome == "Prospectar Después": prosp.append(row)
                if lead.outcome == "Blacklist": black.append(row)
                if st.scheduled and not self._is_blocked(lead, st):
                    totals["pend_fu"] += 1

        def dim_sheet(prefix, d):
            for k, rows in d.items():
                nm = f"{prefix}_{k}".replace("/", "-")[:31]
                sheets[nm] = {"headers": LEAD_HDR, "rows": rows}

        dim_sheet("Ind", by_ind); dim_sheet("Sen", by_sen)
        dim_sheet("Stage", by_stage); dim_sheet("FU", by_fu)
        dim_sheet("Chan", by_chan)
        sheets["Won"] = {"headers": LEAD_HDR, "rows": won}
        sheets["Lost"] = {"headers": LEAD_HDR, "rows": lost}
        sheets["Prospectar_despues"] = {"headers": LEAD_HDR, "rows": prosp}
        sheets["Blacklist"] = {"headers": LEAD_HDR, "rows": black}
        sheets["Scheduled_Messages"] = {
            "headers": SCHEDULED_HEADERS,
            "rows": [[s[h] for h in SCHEDULED_HEADERS] for s in self.read_scheduled()]}
        sheets["Activity_Log"] = {"headers": ACTIVITY_HEADERS,
                                  "rows": self._read_activity_log()}
        sheets["Notifications"] = {
            "headers": NOTIFICATION_HEADERS,
            "rows": [[n[h] for h in NOTIFICATION_HEADERS] for n in self.read_notifications()]}

        sched = self.read_scheduled()
        nuevos = sum(1 for sheet in self.maps for l in self.all_leads(sheet)
                     if self.lead_stage(l) == "Nuevo lead")
        resumen = [
            ["Total leads", totals["total"]],
            ["Total nuevos importados", nuevos],
            ["Total prospectados", totals["prospectados"]],
            ["Total programados", len(sched)],
            ["Total enviados", len([s for s in sched if s.get("Status") == "Sent"])],
            ["Total respondieron", totals["respondieron"]],
            ["Total won", totals["won"]],
            ["Total lost", totals["lost"]],
            ["Total pendientes follow-up", totals["pend_fu"]],
        ]
        resumen += [[f"Industria · {k}", v] for k, v in sorted(t_ind.items())]
        resumen += [[f"Seniority · {k}", v] for k, v in sorted(t_sen.items())]
        resumen += [[f"Canal · {k}", v] for k, v in sorted(t_chan.items())]
        sheets = {"Resumen": {"headers": ["Métrica", "Valor"], "rows": resumen}, **sheets}
        return sheets

    def export_dashboard(self, out_dir=".", when=None):
        """Exporta el dashboard ampliado (multi-hoja por dimensión). NO modifica
        el CRM. Incluye automáticamente todos los leads (nuevos incluidos)."""
        when = when or dt.datetime.now()
        today = when.date().isoformat()
        data = self.build_dashboard_data(when)
        out = Path(out_dir) / f"CRM_dashboard_{today}.xlsx"
        wb = Workbook()
        wb.remove(wb.active)
        used = set()
        for name, sh in data.items():
            nm = name[:31] or "Hoja"
            base = nm
            i = 1
            while nm in used:  # evitar nombres duplicados de hoja
                nm = f"{base[:28]}_{i}"
                i += 1
            used.add(nm)
            ws = wb.create_sheet(nm)
            ws.append(sh["headers"])
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            ws.freeze_panes = "A2"
            for row in sh["rows"]:
                ws.append(row)
        wb.save(out)
        return out

    def lead_status(self, lead, st=None):
        """Estado unificado para la vista: pendiente/enviado/respondió/won/lost/
        blacklist/prospectar después, derivado de Excel + sidecar."""
        st = st or self.state.get(lead.key)
        if lead.outcome == "Blacklist":
            return "blacklist"
        if lead.outcome == "Prospectar Después":
            return "prospectar después"
        if st.responded or lead.outcome == "Respondió":
            return "respondió"
        if lead.outcome in ("Won", "Lost"):
            return lead.outcome.lower()
        touched = any(_is_marked(lead.stage_values.get(s))
                      for steps in CHANNEL_STEPS.values() for s in steps)
        return "enviado" if touched else "pendiente"

    # ---- columnas faltantes (creación automática al final) ------------- #

    def ensure_columns(self, sheet, names):
        """Agrega columnas que falten AL FINAL de la hoja (no mueve nada que el
        Dashboard lea por posición). Idempotente; actualiza el mapeo."""
        ws, cmap = self.wb[sheet], self.maps[sheet]
        added = False
        for name in names:
            if cmap.idx(name) is not None:
                continue
            col = ws.max_column + 1
            c = ws.cell(row=1, column=col, value=name)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
            added = True
        if added:  # re-leer encabezados para refrescar el mapeo
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            self.maps[sheet] = ColumnMap(header)
        return added

    # ---- plantilla de importación -------------------------------------- #

    @staticmethod
    def make_import_template(out_dir="."):
        """Genera plantilla_importacion_leads.xlsx con las columnas recomendadas
        y dos filas de ejemplo."""
        out = Path(out_dir) / "plantilla_importacion_leads.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(TEMPLATE_COLUMNS)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
        ws.freeze_panes = "A2"
        ws.append(["Jane", "Doe", "Jane Doe", "VP Procurement", "Acme Manufacturing",
                   "Automotive", "VP", "https://linkedin.com/in/janedoe",
                   "jane.doe@acme.com", "+1 555 0100", "Detroit, MI",
                   "Reducimos 15% el costo de tooling", "Lead times largos",
                   "Conocida en feria 2025"])
        ws.append(["", "", "John Smith", "Buyer", "Globex Corp", "Packaging", "",
                   "https://linkedin.com/in/johnsmith", "john@globex.com", "", "",
                   "", "Empaque sustentable", ""])
        for col in ws.columns:
            width = max(len(str(c.value or "")) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(width, 32)
        wb.save(out)
        return out

    # ---- prioridad / paso / canal (calculados, no almacenados) --------- #

    @staticmethod
    def lead_priority(lead):
        """Alta / Media / Baja a partir de Seniority Level y rol."""
        text = _norm((lead.seniority_level or "") + " " + (lead.job_title or ""))
        for label, kws in PRIORITY_RULES:
            if any(k in text for k in kws):
                return label
        return "Media"

    @staticmethod
    def classify_warmth(lead, st=None):
        """Nivel de calentamiento: Caliente si respondió/avanzó, Tibio si contactado,
        Frío si nuevo."""
        outcome = (lead.outcome or "")
        if outcome in ("Respondió", "Meeting", "RFQ", "Quote", "Won"):
            return "Caliente"
        touched = any(_is_marked(lead.stage_values.get(s))
                      for steps in CHANNEL_STEPS.values() for s in steps)
        return "Tibio" if touched else "Frío"

    @staticmethod
    def recommend_channel(lead):
        """Canal recomendado según los datos disponibles del lead."""
        if lead.linkedin:
            return "LinkedIn"
        if lead.email:
            return "Email"
        if lead.phone:
            return "Cold Call"
        return "LinkedIn"

    def classify_stage(self, lead, st=None):
        """Current Stage legible (reutiliza lead_stage)."""
        return self.lead_stage(lead, st)

    def classify_lead(self, lead, st=None):
        """Devuelve dict con las 4 clasificaciones para persistir/mostrar."""
        return {
            "Lead Warmth Level": self.classify_warmth(lead, st),
            "Priority": self.lead_priority(lead),
            "Recommended Channel": self.recommend_channel(lead),
            "Current Stage": self.classify_stage(lead, st),
        }

    def lead_stage(self, lead, st=None):
        """Paso actual legible del embudo."""
        st = st or self.state.get(lead.key)
        if lead.outcome == "Blacklist":
            return "Blacklist"
        if lead.outcome == "Won":
            return "Won"
        if lead.outcome == "Lost":
            return "Lost"
        if lead.outcome == "Prospectar Después":
            return "Prospectar después"
        if lead.outcome in ("Meeting", "RFQ", "Quote"):
            return lead.outcome
        if st.responded or lead.outcome == "Respondió":
            return "Respondió"
        sv = lead.stage_values
        for n in (3, 2, 1):
            if _is_marked(sv.get(f"Follow Up {n}")):
                # Follow Up 1 = mensaje inicial; 2+ = follow ups
                return "Follow Up " + str(n - 1) if n >= 2 else "Mensaje inicial enviado"
        if any(_is_marked(sv.get(s)) for steps in CHANNEL_STEPS.values() for s in steps):
            return "Mensaje inicial enviado"
        return "Nuevo lead"

    def lead_channel(self, lead, st=None):
        """Canal actual (último canal con actividad registrada), legible."""
        st = st or self.state.get(lead.key)
        for e in reversed(st.log):
            ch = e.get("channel")
            if ch and ch != "-":
                return {"Cold Call": "Cold Call", "Llamada": "Cold Call",
                        "Email": "Gmail"}.get(ch, ch)
        # respaldo: deducir por columnas marcadas
        sv = lead.stage_values
        if any(_is_marked(sv.get(s)) for s in CHANNEL_STEPS["Cold Call"]):
            return "Cold Call"
        if any(_is_marked(sv.get(s)) for s in CHANNEL_STEPS["Email"]):
            return "Gmail"
        if any(_is_marked(sv.get(s)) for s in CHANNEL_STEPS["LinkedIn"]):
            return "LinkedIn"
        return "Otro"

    # ---- resultado de respuesta (mapeo a acciones) --------------------- #

    def apply_response_result(self, lead, result, channel="", message="",
                              note=None, recontact_date=None):
        """Registra la respuesta (pausa + Activity_Log) y aplica el resultado
        elegido por el usuario. Decisión humana; la app no decide sola."""
        # 1) registrar la respuesta entrante (pausa secuencias + log)
        self.register_response(lead, channel or "-", message or "")
        # 2) aplicar el resultado
        r = result
        if r == "Won":
            self.set_outcome(lead, "Won", note=note)
        elif r == "Lost":
            self.set_outcome(lead, "Lost", note=note)
        elif r == "Blacklist":
            self.set_outcome(lead, "Blacklist", note=note)
        elif r == "Prospectar después":
            self.set_outcome(lead, "Prospectar Después",
                             recontact_date=recontact_date, note=note)
        elif r == "Reunión agendada":
            self.mark_stage(lead, "Meeting")
            if note:
                self._set(lead, "Notes", note)
        elif r == "Cotización solicitada":
            self.mark_stage(lead, "RFQ")
            if note:
                self._set(lead, "Notes", note)
        elif r in ("Respondió interesado", "Respondió no interesado"):
            # no terminal: queda en Respondió (pausado) con la nota
            if note:
                self._set(lead, "Notes", note)
            self._set(lead, "Interest Reason", r)
            self.state.add_log(lead.key, channel or "-", "-", "classify",
                               result=r, notes=note or "")
        else:
            if note:
                self._set(lead, "Notes", note)
        # 3) notificación de la respuesta
        summ = message[:120] if message else result
        self.add_notification(self.read_lead(lead.sheet, lead.row),
                              "Respuesta registrada", f"{result}: {summ}",
                              channel or "-")
        return result

    # ---- cambio manual de estado (12 opciones) ------------------------- #

    def edit_lead_fields(self, lead, warmth=None, priority=None, channel=None,
                         next_action=None, next_date=None, notes=None, stage=None,
                         fu_step=None):
        """Edita campos extra del lead (crea columnas si faltan) y registra en
        Activity_Log. Solo escribe lo que se pasa (no borra lo demás)."""
        self.ensure_columns(lead.sheet, CLASSIFY_COLUMNS)
        if warmth is not None:
            self._set(lead, "Lead Warmth Level", warmth)
        if priority is not None:
            self._set(lead, "Priority", priority)
        if channel is not None:
            self._set(lead, "Recommended Channel", channel)
        if stage is not None:
            self._set(lead, "Current Stage", stage)
        if next_date is not None:
            self._set(lead, "Next Follow-up", next_date)
        if notes is not None:
            self._set(lead, "Notes", notes)
        summary = "; ".join(f"{k}={v}" for k, v in {
            "warmth": warmth, "priority": priority, "channel": channel,
            "stage": stage, "next_action": next_action, "next_date": next_date,
            "fu_step": fu_step}.items() if v is not None)
        self.state.add_log(lead.key, channel or "-", stage or "-", "edit_fields",
                           result=summary, notes=notes or "")
        return summary

    # ---- reprogramación de follow-ups + historial ----------------------- #

    def _ensure_history_sheet(self):
        if FOLLOWUP_HISTORY_SHEET not in self.wb.sheetnames:
            ws = self.wb.create_sheet(FOLLOWUP_HISTORY_SHEET)
            ws.append(FU_HISTORY_HEADERS)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            ws.freeze_panes = "A2"
        return self.wb[FOLLOWUP_HISTORY_SHEET]

    def log_followup_history(self, lead, prev_date, new_date, prev_channel,
                             new_channel, reason="", notes="", user="",
                             status="Reprogramado", when=None):
        when = when or dt.datetime.now()
        ws = self._ensure_history_sheet()
        ws.append([when.isoformat(timespec="seconds"), lead.key, lead.full_name,
                   lead.company, str(prev_date or ""), str(new_date or ""),
                   str(prev_channel or ""), str(new_channel or ""), reason, notes,
                   user, status])

    def reschedule_followup(self, lead, new_date, new_time="", channel=None,
                            reason="", notes=None, user="", fu_step=None,
                            stage=None, priority=None):
        """Reprograma el próximo follow-up de un lead: actualiza la fila (fecha,
        hora, canal, motivo, owner...), re-sincroniza el scheduler interno,
        registra en Follow_Up_History y Activity_Log, y crea notificación."""
        self.ensure_columns(lead.sheet, RESCHEDULE_COLUMNS + CLASSIFY_COLUMNS)
        g = lambda c: self._cell(self.wb[lead.sheet], self.maps[lead.sheet], lead.row, c)
        prev_date = g("Next Follow-up")
        prev_channel = g("Follow Up Channel") or self.lead_channel(lead)

        # escribir campos en la fila del lead (solo lo provisto)
        self._set(lead, "Next Follow-up", new_date)
        if new_time:
            self._set(lead, "Next Follow Up Time", new_time)
        if channel:
            self._set(lead, "Follow Up Channel", channel)
        if reason:
            self._set(lead, "Follow Up Reason", reason)
        if notes:
            self._set(lead, "Notes", notes)
        if user:
            self._set(lead, "Owner/User", user)
        if fu_step:
            self._set(lead, "Follow Up Step", fu_step)
        if stage:
            self._set(lead, "Current Stage", stage)
        if priority:
            self._set(lead, "Priority", priority)

        # re-sincronizar scheduler interno: mover la programación pendiente
        st = self.state.get(lead.key)
        when_iso = f"{new_date}T{new_time or '09:00'}:00" if len(str(new_date)) == 10 \
            else str(new_date)
        seq = {"Gmail": "Email", "Otro": "LinkedIn"}.get(channel, channel) \
            if channel else None
        target_step = fu_step
        if not target_step:
            if st.scheduled:  # mover la más próxima
                target_step = sorted(st.scheduled.items(), key=lambda kv: kv[1])[0][0]
            elif seq in CHANNEL_STEPS:
                for s in CHANNEL_STEPS[seq]:
                    if not _is_marked(lead.stage_values.get(s)):
                        target_step = s
                        break
        if target_step:
            st.scheduled[target_step] = when_iso
            self.state.put(lead.key, st)

        self.log_followup_history(lead, prev_date, f"{new_date} {new_time}".strip(),
                                  prev_channel, channel or prev_channel,
                                  reason=reason, notes=notes or "", user=user)
        self.state.add_log(lead.key, channel or "-", target_step or "-",
                           "reschedule", result=f"{new_date} {new_time}".strip(),
                           notes=reason or "")
        lead2 = self.read_lead(lead.sheet, lead.row)
        self.add_notification(lead2, "Follow-up reprogramado",
                              f"→ {new_date} {new_time} por {channel or prev_channel}"
                              + (f" · {reason}" if reason else ""),
                              channel or "-")
        return target_step, when_iso

    # ---- registro de cold call ------------------------------------------ #

    def register_cold_call(self, lead, call_date=None, result="No contestó",
                           note="", next_action="", next_date=None, user=""):
        """Registra una llamada realizada (manual): marca la siguiente celda
        Cold Call N, aplica el resultado y agenda el próximo contacto si se dio."""
        when = dt.datetime.now() if not call_date else \
            dt.datetime.fromisoformat(f"{call_date}T12:00:00")
        nxt_cell = None
        for s in CHANNEL_STEPS["Cold Call"]:
            if not _is_marked(lead.stage_values.get(s)):
                nxt_cell = s
                break
        if nxt_cell:
            self.mark_sent(lead, "Llamada", nxt_cell, message=note or result,
                           when=when, result=result)
        # aplicar el resultado
        if result == "Won":
            self.set_outcome(lead, "Won", note=note or None)
        elif result == "Lost":
            self.set_outcome(lead, "Lost", note=note or None)
        elif result == "Blacklist":
            self.set_outcome(lead, "Blacklist", note=note or None)
        elif result == "RFQ":
            self.mark_stage(lead, "RFQ")
        elif result == "Reunión agendada":
            self.mark_stage(lead, "Meeting")
        elif result == "Contactar después":
            self.set_outcome(lead, "Prospectar Después",
                             recontact_date=next_date, note=note or None)
        # próximo contacto (si se dio y el lead sigue vivo)
        if next_date and result in ("No contestó", "Interesado", "Pidió información"):
            self.reschedule_followup(lead, next_date, channel="Cold Call",
                                     reason=f"Tras llamada: {result}",
                                     notes=next_action or None, user=user)
        lead2 = self.read_lead(lead.sheet, lead.row)
        self.add_notification(lead2, "Cold call registrada",
                              f"{result}" + (f" · {note[:80]}" if note else ""),
                              "Cold Call")
        return result

    # ---- configuración de colores por estado ---------------------------- #

    def _ensure_color_sheet(self, seed=True):
        if STATE_COLOR_SHEET not in self.wb.sheetnames:
            ws = self.wb.create_sheet(STATE_COLOR_SHEET)
            ws.append(STATE_COLOR_HEADERS)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            ws.freeze_panes = "A2"
            if seed:
                for row in DEFAULT_STATE_COLORS:
                    ws.append(list(row) + ["Sí"])
        return self.wb[STATE_COLOR_SHEET]

    def read_state_colors(self):
        """Lee la config de colores. Devuelve dict: state -> dict(campos + _row)."""
        ws = self._ensure_color_sheet()
        out = {}
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, len(STATE_COLOR_HEADERS) + 1)]
            if not vals[0]:
                continue
            d = dict(zip(STATE_COLOR_HEADERS, vals))
            d["_row"] = r
            out[str(d["State"])] = d
        return out

    def upsert_state_color(self, state, color, scope="row", category="activo",
                           action="", priority="Media", active=True):
        """Crea o actualiza un estado (incluye estados personalizados)."""
        ws = self._ensure_color_sheet()
        cfg = self.read_state_colors()
        row_vals = [state, str(color).lstrip("#").upper(), scope, category,
                    action, priority, "Sí" if active else "No"]
        if state in cfg:
            r = cfg[state]["_row"]
            for i, v in enumerate(row_vals, 1):
                ws.cell(row=r, column=i).value = v
        else:
            ws.append(row_vals)
        return state

    def resolve_display_state(self, lead, st=None):
        """Estado 'visual' del lead para pintar: usa Current Stage si está en la
        config; si no, deriva del embudo/outcome hacia los estados estándar."""
        cfg = self.read_state_colors()
        cur = self._cell(self.wb[lead.sheet], self.maps[lead.sheet],
                         lead.row, "Current Stage")
        if cur and str(cur) in cfg:
            return str(cur)
        stg = self.lead_stage(lead, st)
        mapping = {
            "Won": "Won", "Lost": "Lost", "Blacklist": "Blacklist",
            "Prospectar después": "Prospectar después",
            "RFQ": "RFQ", "Quote": "RFQ", "Meeting": "Reunión agendada",
            "Respondió": "Respondió interesado",
        }
        if stg in mapping:
            return mapping[stg]
        sst = st or self.state.get(lead.key)
        if sst.scheduled:
            return "Follow Up pendiente"
        touched = any(_is_marked(lead.stage_values.get(s))
                      for steps in CHANNEL_STEPS.values() for s in steps)
        return "Mensaje enviado" if touched else "Nuevo lead"

    def apply_state_colors(self, remove_legacy_cf=False):
        """Pinta filas/celdas del Excel según la config (openpyxl PatternFill).
        'row' = toda la fila; 'cell' = solo la celda Current Stage. Blanco/
        FFFFFF = limpiar. Si remove_legacy_cf, elimina el formato condicional
        heredado para que el pintado estático sea la única fuente visual."""
        cfg = self.read_state_colors()
        painted = 0
        for sheet in self.maps:
            self.ensure_columns(sheet, CLASSIFY_COLUMNS)
            ws, cmap = self.wb[sheet], self.maps[sheet]
            if remove_legacy_cf:
                ws.conditional_formatting = ConditionalFormattingList()
            stage_col = cmap.idx("Current Stage")
            for lead in self.all_leads(sheet):
                state = self.resolve_display_state(lead)
                c = cfg.get(state)
                if not c or str(c.get("Active", "Sí")).lower().startswith("n"):
                    continue
                color = str(c["Color"]).lstrip("#").upper()
                clear = color in ("FFFFFF", "")
                fill = PatternFill() if clear else PatternFill("solid", fgColor=color)
                if str(c.get("Scope", "row")).lower() == "row":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=lead.row, column=col).fill = fill
                else:
                    ws.cell(row=lead.row, column=stage_col).fill = fill
                # reflejar el estado visual en Current Stage si estaba vacío
                if not ws.cell(row=lead.row, column=stage_col).value:
                    ws.cell(row=lead.row, column=stage_col).value = state
                painted += 1
        return painted

    def set_manual_state(self, lead, state, channel=None, note=None,
                         recontact_date=None):
        """Aplica cualquiera de los 12 estados manuales. Los Outcome pintan fila;
        los de embudo marcan la celda de etapa. Registra en Activity_Log y crea
        una notificación. Devuelve el estado aplicado."""
        outcome_map = {
            "Won": "Won", "Lost": "Lost", "Blacklist": "Blacklist",
            "Prospectar después": "Prospectar Después", "Respondió": "Respondió",
            "Reunión agendada": "Meeting", "Cotización solicitada": "RFQ",
        }
        if state in outcome_map:
            val = outcome_map[state]
            if val in ("Meeting", "RFQ"):
                self.mark_stage(lead, val)
                if note:
                    self._set(lead, "Notes", note)
            elif val == "Respondió":
                self.register_response(lead, channel or "-", note or "")
            else:
                self.set_outcome(lead, val, recontact_date=recontact_date, note=note)
        elif state in ("Mensaje inicial enviado", "Follow Up 1", "Follow Up 2",
                       "Follow Up 3"):
            # mapear a la celda de etapa LinkedIn (Follow Up 1 = mensaje inicial)
            step = "Follow Up 1" if state == "Mensaje inicial enviado" else \
                   "Follow Up " + str(int(state.split(" ")[-1]) + 1)
            self.mark_sent(lead, channel or "LinkedIn", step,
                           message=note or "", result="Enviado")
        elif state == "Nuevo lead":
            self._set(lead, "Outcome Status", "Active")
            if note:
                self._set(lead, "Notes", note)
        self.state.add_log(lead.key, channel or "-", state, "manual_state",
                           result=state, notes=note or "")
        lead2 = self.read_lead(lead.sheet, lead.row)
        self.add_notification(lead2, "Cambio de estado",
                              f"Estado → {state}" + (f" · {note}" if note else ""),
                              channel or "-")
        return state

    def next_action(self, lead, st=None):
        """Siguiente acción sugerida legible (no ejecuta nada)."""
        st = st or self.state.get(lead.key)
        blk = self._is_blocked(lead, st)
        if lead.outcome in ("Won", "Lost"):
            return "—"
        if lead.outcome == "Blacklist":
            return "No contactar"
        if lead.outcome == "Prospectar Después":
            return f"Recontactar {st.recontact_date}" if st.recontact_date else \
                   "Definir fecha de recontacto"
        if st.responded or lead.outcome == "Respondió":
            return "Clasificar respuesta"
        if st.scheduled:
            step, iso = sorted(st.scheduled.items(), key=lambda kv: kv[1])[0]
            return f"{step} el {iso[:10]}"
        esc = self.escalation_suggestion(lead)
        if esc:
            return esc
        touched = any(_is_marked(lead.stage_values.get(s))
                      for steps in CHANNEL_STEPS.values() for s in steps)
        return "Continuar secuencia" if touched else "Enviar mensaje inicial"

    # ---- backup + guardado robusto ------------------------------------- #

    def make_backup(self, backup_dir=None):
        """Copia el Excel actual a backups/ con timestamp, antes de modificar."""
        if not self.xlsx_path.exists():
            return None
        bdir = Path(backup_dir) if backup_dir else self.xlsx_path.parent / "backups"
        bdir.mkdir(exist_ok=True)
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = bdir / f"{self.xlsx_path.stem}_{stamp}.xlsx"
        shutil.copy(self.xlsx_path, dest)
        return dest

    def save(self, out_path=None, backup=True):
        """Guarda con backup previo y manejo de archivo abierto (PermissionError).
        Devuelve (ruta, backup, error_msg). Si el Excel está abierto en otra app,
        no truena: devuelve un mensaje claro."""
        out = Path(out_path) if out_path else self.xlsx_path
        bkp = self.make_backup() if (backup and out == self.xlsx_path) else None
        try:
            self.wb.save(out)
            self.state.save()
            return out, bkp, None
        except PermissionError:
            return None, bkp, (
                "No pude guardar: el archivo de Excel parece estar ABIERTO en otra "
                "aplicación (Excel/LibreOffice). Ciérralo y vuelve a intentar. "
                "Tus cambios siguen en memoria; el backup previo está intacto.")
        except OSError as e:
            return None, bkp, f"No pude guardar el archivo: {e}"

    # ===================================================================== #
    # Alta manual de leads y perfil de prospección
    # ===================================================================== #

    def add_lead_manual(self, sheet, **fields):
        """Agrega un lead capturado a mano. fields usa nombres canónicos
        (Full Name, Company Name, Job Title, Industry, LinkedIn Profile,
        Email/Gmail, Phone, Location, Seniority Level, Notes...)."""
        import pandas as pd
        if sheet not in self.maps:
            sheet = next(iter(self.maps))
        df = pd.DataFrame([{k: v for k, v in fields.items() if v}])
        mapping = {k: k for k in df.columns}
        added, dups, per = self.import_leads(df, mapping,
                                             route_by_industry=False,
                                             default_industry=sheet)
        return added, dups

    def build_prospecting_profile(self, lead, st=None):
        """Prospecting Profile calculado con reglas (sin inventar datos que no
        existen). Devuelve dict con los 9 campos + mensaje sugerido."""
        st = st or self.state.get(lead.key)
        pain_col = self._cell(self.wb[lead.sheet], self.maps[lead.sheet],
                              lead.row, "Pain Point")
        vp_col = self._cell(self.wb[lead.sheet], self.maps[lead.sheet],
                            lead.row, "Value Proposition")
        ind_pain, ind_service = INDUSTRY_PAINS.get(
            lead.industry, ("Costos y confiabilidad de proveedores",
                            "Manufactura confiable con costo competitivo"))
        rank = lead.seniority_rank if lead.seniority_rank < 900 else None
        decision = ("Decisor final" if (rank or 9) <= 2 else
                    "Alto poder de decisión" if (rank or 9) <= 4 else
                    "Influenciador / recomendador" if (rank or 9) <= 6 else
                    "Usuario / evaluador técnico")
        prio = self.lead_priority(lead)
        warmth = self.classify_warmth(lead, st)
        angle = ("Directo a valor de negocio (ROI, costo total)"
                 if prio == "Alta" else
                 "Consultivo: dolor específico del rol y caso similar"
                 if prio == "Media" else
                 "Educativo: material útil y contacto de bajo compromiso")
        risk = []
        if not lead.email and not lead.phone:
            risk.append("sin email/teléfono: LinkedIn es el único canal")
        if lead.outcome == "Prospectar Después":
            risk.append("pidió recontacto: respetar la fecha")
        if warmth == "Frío":
            risk.append("sin contacto previo: primer toque, no vender de golpe")
        pain = str(pain_col) if pain_col else ind_pain
        service = str(vp_col) if vp_col else ind_service
        first = (lead.full_name or "").split(" ")[0]
        msg = (f"Hola {first}, vi tu rol de {lead.job_title or 'liderazgo'} en "
               f"{lead.company}. Trabajamos con empresas de {lead.industry} en "
               f"{pain.lower()[:60]} — {service.lower()[:60]}. "
               f"¿Te hace sentido conectar?")
        return {
            "Por qué es relevante": f"{lead.seniority_level or 'Rol'} en "
                                    f"{lead.company} ({lead.industry}); "
                                    f"prioridad {prio}, warmth {warmth}",
            "Industria": lead.industry,
            "Nivel de decisión": decision,
            "Problema probable": pain,
            "Servicio a ofrecer": service,
            "Ángulo de mensaje": angle,
            "Prioridad": prio,
            "Riesgo / nota": "; ".join(risk) or "Sin riesgos detectados",
            "Mensaje sugerido (LinkedIn)": msg,
        }

    # ===================================================================== #
    # Perfil de Cliente Ideal (Buyer Persona) + evaluación de fit
    # ===================================================================== #

    def read_icp(self):
        """Lee el Perfil de Cliente Ideal desde la hoja (con defaults)."""
        return self.read_kv(ICP_SHEET, ICP_DEFAULTS)

    def save_icp(self, profile: dict):
        """Guarda el perfil completo en la hoja Ideal_Customer_Profile."""
        for k, _label in ICP_FIELDS:
            self.set_kv(ICP_SHEET, k, profile.get(k, ""), ICP_DEFAULTS)

    @staticmethod
    def _terms(csv_text):
        return [t.strip() for t in str(csv_text or "").split(",") if t.strip()]

    @staticmethod
    def _match_any(text, terms):
        blob = _norm(str(text or ""))
        return [t for t in terms if _norm(t) and _norm(t) in blob]

    def evaluate_lead_fit(self, lead, icp=None, st=None):
        """Evalúa un lead contra el Cliente Ideal. Reglas transparentes, sin
        inventar datos: los criterios sin dato o sin configurar se omiten.
        Devuelve dict: score 0-100, level, pros, cons, canal, mensaje, pain,
        beneficio, accion."""
        icp = icp or self.read_icp()
        st = st or self.state.get(lead.key)
        loc = self._cell(self.wb[lead.sheet], self.maps[lead.sheet],
                         lead.row, "Location") or ""
        notes = self._cell(self.wb[lead.sheet], self.maps[lead.sheet],
                           lead.row, "Notes") or ""
        pain_col = self._cell(self.wb[lead.sheet], self.maps[lead.sheet],
                              lead.row, "Pain Point") or ""
        vp_col = self._cell(self.wb[lead.sheet], self.maps[lead.sheet],
                            lead.row, "Value Proposition") or ""
        pros, cons = [], []
        score = 50

        # bloqueados: score 0 directo
        if self._is_blocked(lead, st):
            estado = lead.outcome or ("Respondió" if st.responded else "bloqueado")
            return {"score": 0, "level": "Malo",
                    "pros": [], "cons": [f"Lead bloqueado/cerrado: {estado}"],
                    "canal": "—", "mensaje": "", "pain": pain_col or "—",
                    "beneficio": vp_col or "—", "accion": self.next_action(lead, st)}

        def crit(terms_key, value, pts_hit, pts_miss, pro_txt, con_txt):
            nonlocal score
            terms = self._terms(icp.get(terms_key))
            if not terms or value in (None, ""):
                return
            hits = self._match_any(value, terms)
            if hits:
                score += pts_hit
                pros.append(pro_txt.format(", ".join(hits[:2])))
            elif pts_miss:
                score += pts_miss
                cons.append(con_txt)

        crit("industrias_objetivo", lead.industry, 15, -10,
             "Industria objetivo ({})", "Industria fuera del objetivo")
        crit("puestos_objetivo", lead.job_title, 15, -5,
             "Puesto objetivo ({})", "Puesto no coincide con el objetivo")
        crit("seniority_ideal", (lead.seniority_level or "") + " " +
             (lead.job_title or ""), 10, 0, "Seniority ideal ({})", "")
        crit("ubicacion_ideal", loc, 8, -3,
             "Ubicación ideal ({})", "Fuera de la ubicación ideal")
        crit("empresas_objetivo", lead.company, 8, 0,
             "Empresa objetivo ({})", "")

        blob = " ".join([lead.job_title or "", lead.company or "", str(notes),
                         str(pain_col), str(vp_col)])
        pos = self._match_any(blob, self._terms(icp.get("keywords_positivas")))
        if pos:
            score += 10
            pros.append(f"Keywords positivas: {', '.join(pos[:3])}")
        neg = self._match_any(blob, self._terms(icp.get("keywords_negativas")))
        if neg:
            score -= 25
            cons.append(f"Keywords negativas: {', '.join(neg[:3])}")
        evitar = self._match_any(
            " ".join([lead.full_name or "", lead.company or "",
                      lead.job_title or ""]),
            self._terms(icp.get("leads_a_evitar")))
        cap = 100
        if evitar:
            cap = 20   # lista explícita de evitar: el fit no puede ser bueno
            cons.append(f"Coincide con 'leads a evitar': {', '.join(evitar[:2])}")

        # canal preferido vs datos de contacto disponibles
        prefs = [c.strip() for c in
                 str(icp.get("canales_preferidos") or "").split(",") if c.strip()]
        has = {"LinkedIn": bool(lead.linkedin), "Email": bool(lead.email),
               "Cold Call": bool(lead.phone)}
        avail = [c for c in prefs if has.get(c)]
        if prefs:
            if avail:
                score += 8
                pros.append(f"Datos para canal preferido: {avail[0]}")
            else:
                score -= 10
                cons.append("Sin datos de contacto para los canales preferidos")
        canal = avail[0] if avail else self.recommend_channel(lead)

        # prioridad mínima aceptada
        order = {"Alta": 3, "Media": 2, "Baja": 1}
        prio = self.lead_priority(lead)
        minp = str(icp.get("prioridad_minima") or "Baja").strip().title()
        if order.get(prio, 2) < order.get(minp, 1):
            score -= 15
            cons.append(f"Prioridad {prio} debajo del mínimo ({minp})")
        else:
            pros.append(f"Prioridad {prio}")

        # historial / estado
        if st.responded or lead.outcome == "Respondió":
            score += 10
            pros.append("Ya respondió antes (lead tibio/caliente)")
        if self.classify_warmth(lead, st) == "Caliente":
            score += 5
        if lead.outcome == "Prospectar Después":
            cons.append("Pidió recontacto: respetar la fecha")

        score = max(0, min(cap, min(100, score)))
        level = ("Excelente" if score >= 80 else "Bueno" if score >= 60 else
                 "Medio" if score >= 40 else "Bajo" if score >= 20 else "Malo")

        pain = str(pain_col) or (self._terms(icp.get("problemas_que_resuelvo"))
                                 or [INDUSTRY_PAINS.get(lead.industry,
                                     ("Costos y confiabilidad", ""))[0]])[0]
        beneficio = str(vp_col) or (self._terms(icp.get("beneficios"))
                                    or ["Valor claro para su operación"])[0]
        tmpl = icp.get("mensaje_base") or ICP_DEFAULTS["mensaje_base"]
        tmpl = tmpl.replace("{{problema}}", pain).replace("{{beneficio}}",
                                                          beneficio)
        mensaje = self.personalize(tmpl, lead)
        return {"score": score, "level": level, "pros": pros, "cons": cons,
                "canal": canal, "mensaje": mensaje, "pain": pain,
                "beneficio": beneficio, "accion": self.next_action(lead, st)}

    def assisted_mark_sent(self, lead, channel="LinkedIn", message="", user=""):
        """Marca enviado el siguiente paso del canal (asistido, fuera de
        campaña): celda de etapa + Current Stage + Follow Up Step + próxima
        fecha según Workflow_Config + Activity_Log. Devuelve (paso, próxima)."""
        seq = {"Gmail": "Email", "Email": "Email",
               "Cold Call": "Cold Call"}.get(channel, "LinkedIn")
        steps = CHANNEL_STEPS[seq]
        idx = next((i for i, s in enumerate(steps)
                    if not _is_marked(lead.stage_values.get(s))), None)
        if idx is None:
            return None, None
        step = steps[idx]
        self.mark_sent(lead, channel, step, message=message)
        self.ensure_columns(lead.sheet, CLASSIFY_COLUMNS + RESCHEDULE_COLUMNS)
        stage_label = "Mensaje inicial enviado" if idx == 0 else f"Follow Up {idx}"
        self._set(lead, "Current Stage", stage_label)
        self._set(lead, "Follow Up Step", step)
        if user:
            self._set(lead, "Owner/User", user)
        nxt = None
        if idx + 1 < len(steps):
            gap_h = self.fu_gap_hours(idx)
            nxt = dt.datetime.now() + dt.timedelta(hours=gap_h)
            sst = self.state.get(lead.key)
            sst.scheduled[steps[idx + 1]] = nxt.isoformat(timespec="seconds")
            self.state.put(lead.key, sst)
            self._set(lead, "Next Follow-up", nxt.date().isoformat())
        return step, nxt

    # ===================================================================== #
    # Configuración clave-valor (Notification_Settings / Workflow_Config)
    # ===================================================================== #

    def _ensure_kv_sheet(self, name, defaults):
        if name not in self.wb.sheetnames:
            ws = self.wb.create_sheet(name)
            ws.append(KV_HEADERS)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            for k, v in defaults.items():
                ws.append([k, v])
        return self.wb[name]

    def read_kv(self, name, defaults):
        ws = self._ensure_kv_sheet(name, defaults)
        out = dict(defaults)
        for r in range(2, ws.max_row + 1):
            k, v = ws.cell(r, 1).value, ws.cell(r, 2).value
            if k:
                out[str(k)] = str(v) if v is not None else ""
        return out

    def set_kv(self, name, key, value, defaults=None):
        ws = self._ensure_kv_sheet(name, defaults or {})
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value) == key:
                ws.cell(r, 2).value = str(value)
                return
        ws.append([key, str(value)])

    def notif_settings(self):
        return self.read_kv(NOTIF_SETTINGS_SHEET, NOTIF_DEFAULTS)

    def workflow_config(self):
        return self.read_kv(WORKFLOW_SHEET, WORKFLOW_DEFAULTS)

    def fu_gap_hours(self, step_index):
        """Horas configuradas entre el paso step_index y el siguiente."""
        cfg = self.workflow_config()
        keys = ["fu1_to_fu2_hours", "fu2_to_fu3_hours", "fu3_to_fu4_hours"]
        k = keys[min(step_index, len(keys) - 1)]
        try:
            return float(cfg.get(k, 72))
        except ValueError:
            return 72.0

    # ===================================================================== #
    # Alertas configurables de no-respuesta
    # ===================================================================== #

    def scan_no_response(self):
        """Genera alertas para leads contactados que no respondieron tras las
        horas configuradas, follow-ups vencidos y leads a reprogramar/descartar.
        No duplica alertas del mismo tipo por lead. Devuelve cuántas creó."""
        hours = float(self.notif_settings().get("no_response_hours", 48))
        now = dt.datetime.now()
        existing = {(n.get("Lead"), n.get("Tipo de evento"))
                    for n in self.read_notifications()}
        created = 0
        for sheet in self.maps:
            for lead in self.all_leads(sheet):
                st = self.state.get(lead.key)
                if self._is_blocked(lead, st):
                    continue
                lc = self._cell(self.wb[sheet], self.maps[sheet], lead.row,
                                "Last Contact")
                if not lc:
                    continue
                try:
                    last = dt.datetime.fromisoformat(str(lc)[:10])
                except ValueError:
                    continue
                elapsed_h = (now - last).total_seconds() / 3600
                if elapsed_h < hours:
                    continue
                # ¿qué alerta corresponde?
                fu_sent = sum(_is_marked(lead.stage_values.get(s))
                              for s in CHANNEL_STEPS["LinkedIn"])
                maxfu = int(float(self.workflow_config().get("max_followups", 3)))
                if fu_sent > maxfu:
                    ev, summ = "Lead debe descartarse", \
                        f"{lead.full_name}: {fu_sent} toques sin respuesta"
                elif st.scheduled and any(
                        dt.datetime.fromisoformat(v).date() < now.date()
                        for v in st.scheduled.values()):
                    ev, summ = "Follow-up vencido", \
                        f"{lead.full_name}: follow-up programado ya venció"
                elif fu_sent >= 1:
                    ev = f"Campaña lista para Follow Up {fu_sent + 1}"
                    summ = f"{lead.full_name}: {elapsed_h:.0f}h sin respuesta " \
                           f"tras Follow Up {fu_sent}"
                else:
                    ev, summ = "Lead sin respuesta", \
                        f"{lead.full_name}: {elapsed_h:.0f}h sin respuesta"
                if (lead.full_name, ev) in existing:
                    continue
                self.add_notification(lead, ev, summ, self.lead_channel(lead, st))
                created += 1
        return created

    # ===================================================================== #
    # Campañas (workflow asistido LinkedIn / multicanal)
    # ===================================================================== #

    def _ensure_campaigns_sheet(self):
        if CAMPAIGNS_SHEET not in self.wb.sheetnames:
            ws = self.wb.create_sheet(CAMPAIGNS_SHEET)
            ws.append(CAMPAIGN_HEADERS)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            ws.freeze_panes = "A2"
        return self.wb[CAMPAIGNS_SHEET]

    def create_campaign(self, name, channel, industry, leads, step_index=0,
                        status="Activa", scheduled_date=None, notes=""):
        """Registra una campaña con su lista de leads (por key)."""
        ws = self._ensure_campaigns_sheet()
        cid = f"CMP-{ws.max_row:04d}"
        keys = "|".join(l.key for l in leads)
        ws.append([cid, name, channel, industry, step_index, status, len(leads), 0,
                   scheduled_date or dt.date.today().isoformat(),
                   dt.datetime.now().isoformat(timespec="seconds"), notes, keys])
        self.state.add_log("__campaigns__", channel, f"step{step_index}",
                           "campaign_created", result=cid, notes=name)
        return cid

    def read_campaigns(self, status=None):
        if CAMPAIGNS_SHEET not in self.wb.sheetnames:
            return []
        ws = self.wb[CAMPAIGNS_SHEET]
        out = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, len(CAMPAIGN_HEADERS) + 1)]
            if not vals[0]:
                continue
            d = dict(zip(CAMPAIGN_HEADERS, vals))
            d["_row"] = r
            if status and d.get("Status") != status:
                continue
            out.append(d)
        return out

    def _campaign_step_name(self, camp):
        idx = int(camp.get("Step") or 0)
        ch = str(camp.get("Channel") or "LinkedIn")
        seq = {"Gmail": "Email", "Email": "Email", "Cold Call": "Cold Call"} \
            .get(ch, "LinkedIn")
        steps = CHANNEL_STEPS[seq]
        return steps[min(idx, len(steps) - 1)], seq

    def campaign_pending_leads(self, camp):
        """Leads de la campaña que aún no reciben el paso actual y siguen vivos
        (excluye respondidos, descartados, blacklist, Won...)."""
        step, _seq = self._campaign_step_name(camp)
        keys = str(camp.get("Lead Keys") or "").split("|")
        pend = []
        for k in keys:
            lead = self._find_lead_by_key(k)
            if lead is None:
                continue
            st = self.state.get(lead.key)
            if self._is_blocked(lead, st):
                continue
            if _is_marked(lead.stage_values.get(step)):
                continue
            pend.append(lead)
        return pend

    def next_pending_lead(self, camp):
        pend = self.campaign_pending_leads(camp)
        return pend[0] if pend else None

    def mark_campaign_sent(self, camp, lead, message="", user=""):
        """Confirmación manual: marca el paso actual como enviado, calcula el
        siguiente follow-up con la config del workflow, actualiza contadores y,
        si la campaña terminó, crea automáticamente la de Follow Up N+1."""
        step, seq = self._campaign_step_name(camp)
        self.mark_sent(lead, camp.get("Channel") or seq, step, message=message)
        # reflejar en columnas: Current Stage + Follow Up Step
        self.ensure_columns(lead.sheet, CLASSIFY_COLUMNS + RESCHEDULE_COLUMNS)
        idx0 = int(camp.get("Step") or 0)
        stage_label = "Mensaje inicial enviado" if idx0 == 0 else f"Follow Up {idx0}"
        self._set(lead, "Current Stage", stage_label)
        self._set(lead, "Follow Up Step", step)
        # siguiente follow-up con las horas configuradas
        idx = int(camp.get("Step") or 0)
        gap_h = self.fu_gap_hours(idx)
        when = dt.datetime.now() + dt.timedelta(hours=gap_h)
        steps = CHANNEL_STEPS[seq]
        if idx + 1 < len(steps):
            st = self.state.get(lead.key)
            st.scheduled[steps[idx + 1]] = when.isoformat(timespec="seconds")
            self.state.put(lead.key, st)
            self._set(lead, "Next Follow-up", when.date().isoformat())
        if user:
            self._set(lead, "Owner/User", user)
        # contador de la campaña
        ws = self.wb[CAMPAIGNS_SHEET]
        r = camp["_row"]
        sent = int(ws.cell(r, 8).value or 0) + 1
        ws.cell(r, 8).value = sent
        # ¿terminó? (campaign_pending_leads ya excluye a este lead: su celda
        # de etapa acaba de marcarse en mark_sent)
        remaining = len(self.campaign_pending_leads(camp))
        if remaining <= 0 and str(ws.cell(r, 6).value) != "Terminada":
            ws.cell(r, 6).value = "Terminada"
            self._auto_create_followup_campaign(camp, when)
        return sent, when

    def _auto_create_followup_campaign(self, camp, scheduled_dt):
        """Al terminar una campaña, crea 'Follow Up N+1 - <nombre>' con los
        mismos leads menos respondidos/descartados/blacklist/Won."""
        idx = int(camp.get("Step") or 0)
        _step, seq = self._campaign_step_name(camp)
        steps = CHANNEL_STEPS[seq]
        maxfu = int(float(self.workflow_config().get("max_followups", 3)))
        if idx + 1 >= len(steps) or idx + 1 > maxfu:
            return None
        keys = str(camp.get("Lead Keys") or "").split("|")
        survivors = []
        for k in keys:
            lead = self._find_lead_by_key(k)
            if lead is None:
                continue
            st = self.state.get(lead.key)
            if self._is_blocked(lead, st):     # respondió/blacklist/won/lost/DNC
                continue
            survivors.append(lead)
        if not survivors:
            return None
        base = str(camp.get("Name") or "").split(" - ", 1)[-1] \
            if str(camp.get("Name") or "").startswith("Follow Up") \
            else str(camp.get("Name") or "")
        name = f"Follow Up {idx + 2} - {base}"
        if any(str(c.get("Name")) == name for c in self.read_campaigns()):
            return None  # ya existe: no duplicar
        cid = self.create_campaign(name, camp.get("Channel"), camp.get("Industry"),
                                   survivors, step_index=idx + 1, status="Pendiente",
                                   scheduled_date=scheduled_dt.date().isoformat())
        # notificación-recordatorio
        fake = survivors[0]
        self.add_notification(fake, "Campaña lista",
                              f"Tienes una campaña lista: {name} · "
                              f"{len(survivors)} leads pendientes · "
                              f"{camp.get('Channel')}", camp.get("Channel") or "-")
        return cid

    def reschedule_campaign(self, camp, new_date):
        """Reprograma la fecha de una campaña Pendiente/Pausada."""
        ws = self.wb[CAMPAIGNS_SHEET]
        ws.cell(camp["_row"], 9).value = str(new_date)
        ws.cell(camp["_row"], 6).value = "Pendiente"
        self.state.add_log("__campaigns__", camp.get("Channel") or "-", "-",
                           "campaign_rescheduled", result=str(new_date),
                           notes=camp.get("Name") or "")

    def scan_campaign_reminders(self):
        """Notifica campañas Pendientes cuya fecha ya llegó."""
        today = dt.date.today()
        existing = {n.get("Mensaje/Resumen") for n in self.read_notifications()
                    if n.get("Tipo de evento") == "Campaña lista"}
        created = 0
        for camp in self.read_campaigns(status="Pendiente"):
            try:
                d = dt.date.fromisoformat(str(camp.get("Scheduled Date"))[:10])
            except ValueError:
                continue
            if d > today:
                continue
            pend = self.campaign_pending_leads(camp)
            if not pend:
                continue
            summ = (f"Tienes una campaña lista: {camp.get('Name')} · "
                    f"{len(pend)} leads pendientes · {camp.get('Channel')}")
            if summ in existing:
                continue
            self.add_notification(pend[0], "Campaña lista", summ,
                                  camp.get("Channel") or "-")
            created += 1
        return created

    # ===================================================================== #
    # Email: clasificación de respuestas y procesamiento de cola
    # ===================================================================== #

    @staticmethod
    def classify_email_reply(from_email, subject, snippet):
        """Clasifica una respuesta de Gmail por palabras clave. Devuelve una de:
        bounce, blacklist, not_interested, rfq, meeting, later, interested."""
        blob = _norm(f"{from_email} {subject} {snippet}")
        for cat in ("bounce", "blacklist", "not_interested", "rfq", "meeting",
                    "later"):
            if any(_norm(k) in blob for k in REPLY_KEYWORDS[cat]):
                return cat
        return "interested"

    def handle_email_event(self, lead, category, message="", recontact_date=None):
        """Aplica el efecto de una respuesta/rebote clasificado y notifica.
        Siempre cancela los mensajes programados del lead cuando corresponde."""
        label = {
            "bounce": "Email Bounced", "blacklist": "Blacklist",
            "not_interested": "Lost", "later": "Prospectar Después",
        }.get(category)
        if category == "bounce":
            self._set(lead, "Outcome Status", "Email Bounced")
            self.cancel_scheduled_for_lead(lead)
            self.state.add_log(lead.key, "Email", "-", "bounce", result="Email Bounced")
        elif category in ("blacklist", "not_interested"):
            self.set_outcome(lead, label, note=message[:120] or None)
            self.cancel_scheduled_for_lead(lead)
        elif category == "later":
            self.set_outcome(lead, "Prospectar Después",
                             recontact_date=recontact_date, note=message[:120] or None)
            self.cancel_scheduled_for_lead(lead)
        elif category == "rfq":
            self.register_response(lead, "Email", message)
            self.mark_stage(lead, "RFQ")
        elif category == "meeting":
            self.register_response(lead, "Email", message)
            self.mark_stage(lead, "Meeting")
        else:  # interested
            self.register_response(lead, "Email", message)
        lead2 = self.read_lead(lead.sheet, lead.row)
        self.add_notification(lead2, f"Email: {category}",
                              (message or category)[:120], "Email")
        return category

    def cancel_scheduled_for_lead(self, lead):
        """Cancela todos los mensajes Scheduled/Draft del lead y limpia su
        scheduler interno (detiene follow-ups automáticos)."""
        n = 0
        for s in self.read_scheduled():
            if s.get("Lead ID") == lead.key and s.get("Status") in ("Scheduled",
                                                                    "Draft",
                                                                    "Ready to send"):
                ws = self.wb[SCHEDULED_SHEET]
                ws.cell(s["_row"], SCHEDULED_HEADERS.index("Status") + 1).value = \
                    "Cancelled"
                n += 1
        st = self.state.get(lead.key)
        st.scheduled = {}
        self.state.put(lead.key, st)
        return n

    def emails_sent_today(self):
        today = dt.date.today().isoformat()
        return sum(1 for r in self._read_activity_log()
                   if str(r[4]) == today and str(r[5]) in ("Email", "Gmail"))

    def process_due_emails(self, send_fn, daily_cap=None, now=None):
        """Procesa la cola de emails vencidos (Status=Scheduled, canal Email/
        Gmail, fecha <= ahora). send_fn(to, subject, body) hace el envío real
        (inyectable: en la app es Gmail API; en pruebas, un mock). Respeta el
        tope diario, salta leads bloqueados (cancelándolos) y evita duplicados
        marcando Sent antes de continuar. Devuelve (enviados, saltados, motivo)."""
        now = now or dt.datetime.now()
        cap = daily_cap if daily_cap is not None else \
            int(float(self.workflow_config().get("email_daily_cap", 50)))
        already = self.emails_sent_today()
        sent = skipped = 0
        for s in self.read_scheduled(status="Scheduled"):
            if str(s.get("Channel")) not in ("Email", "Gmail"):
                continue
            try:
                when = dt.datetime.fromisoformat(str(s.get("Scheduled DateTime"))[:19])
            except ValueError:
                continue
            if when > now:
                continue
            if already + sent >= cap:
                return sent, skipped, f"Tope diario alcanzado ({cap})"
            lead = self._find_lead_by_key(s.get("Lead ID"))
            if lead is None:
                skipped += 1
                continue
            st = self.state.get(lead.key)
            if self._is_blocked(lead, st):      # respondió / blacklist / DNC...
                self.cancel_scheduled_for_lead(lead)
                skipped += 1
                continue
            if not lead.email:
                skipped += 1
                continue
            subject = (s.get("Notes") or "").split("|SUBJ|")[-1] \
                if "|SUBJ|" in str(s.get("Notes") or "") else \
                f"Seguimiento — {lead.company or 'contacto'}"
            # marcar Sent ANTES del envío real evita dobles si algo re-entra
            self.update_schedule_status(s["_row"], "Sent", when=now)
            try:
                send_fn(lead.email, subject, s.get("Message") or "")
            except Exception as e:
                # revertir a Scheduled si el envío falló
                ws = self.wb[SCHEDULED_SHEET]
                ws.cell(s["_row"], SCHEDULED_HEADERS.index("Status") + 1).value = \
                    "Scheduled"
                ws.cell(s["_row"],
                        SCHEDULED_HEADERS.index("Sent DateTime") + 1).value = ""
                skipped += 1
                self.state.add_log(lead.key, "Email", s.get("Follow Up Step") or "-",
                                   "send_error", result=str(e)[:80])
                continue
            stage = "Follow Up sent" if str(s.get("Follow Up Step") or "") \
                .startswith(("Email 2", "Email 3", "Email 4", "Email 5")) \
                else "Email sent"
            self._set(lead, "Current Stage", stage)
            sent += 1
        return sent, skipped, None

    # ===================================================================== #
    # Gmail Campaigns (hojas, cuentas, personalización, Follow Up 1)
    # ===================================================================== #

    def _ensure_headers_sheet(self, name, headers):
        if name not in self.wb.sheetnames:
            ws = self.wb.create_sheet(name)
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F4E78")
            ws.freeze_panes = "A2"
        return self.wb[name]

    def ensure_gmail_sheets(self):
        """Crea si no existen todas las hojas del módulo Gmail (y las comunes)."""
        self._ensure_headers_sheet(GMAIL_ACCOUNTS_SHEET, GMAIL_ACCOUNTS_HEADERS)
        self._ensure_headers_sheet(GMAIL_CAMPAIGNS_SHEET, GMAIL_CAMPAIGNS_HEADERS)
        self._ensure_headers_sheet(GMAIL_CAMPAIGN_LEADS_SHEET,
                                   GMAIL_CAMPAIGN_LEADS_HEADERS)
        self._ensure_headers_sheet(GMAIL_FOLLOWUPS_SHEET, GMAIL_FOLLOWUPS_HEADERS)
        self._ensure_scheduled_sheet(); self._ensure_history_sheet()
        self._ensure_notifications_sheet(); self._ensure_activity_sheet()

    def save_gmail_account(self, email, status="Conectada"):
        """Upsert de la cuenta conectada en Gmail_Accounts."""
        ws = self._ensure_headers_sheet(GMAIL_ACCOUNTS_SHEET, GMAIL_ACCOUNTS_HEADERS)
        now = dt.datetime.now().isoformat(timespec="seconds")
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value).lower() == str(email).lower():
                ws.cell(r, 3).value = now
                ws.cell(r, 4).value = status
                return r
        ws.append([email, now, now, status])
        return ws.max_row

    def read_gmail_accounts(self):
        if GMAIL_ACCOUNTS_SHEET not in self.wb.sheetnames:
            return []
        ws = self.wb[GMAIL_ACCOUNTS_SHEET]
        return [dict(zip(GMAIL_ACCOUNTS_HEADERS,
                         [ws.cell(r, c).value for c in range(1, 5)]))
                for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value]

    @staticmethod
    def personalize(template, lead):
        """Variables dinámicas {{first_name}}, {{full_name}}, {{company}},
        {{job_title}}, {{industry}}, {{seniority_level}} (y estilo {first})."""
        first = (lead.full_name or "").split(" ")[0]
        pairs = {"first_name": first, "full_name": lead.full_name or "",
                 "company": lead.company or "", "job_title": lead.job_title or "",
                 "industry": lead.industry or "",
                 "seniority_level": lead.seniority_level or "",
                 "first": first, "name": lead.full_name or "",
                 "title": lead.job_title or ""}
        out = str(template or "")
        for k, v in pairs.items():
            out = out.replace("{{" + k + "}}", v).replace("{" + k + "}", v)
        return out

    def create_gmail_campaign(self, name, sender, user="", subject="",
                              total=0, status="Draft"):
        ws = self._ensure_headers_sheet(GMAIL_CAMPAIGNS_SHEET,
                                        GMAIL_CAMPAIGNS_HEADERS)
        cid = f"GM-{ws.max_row:04d}"
        ws.append([cid, name, sender, "Gmail",
                   dt.datetime.now().isoformat(timespec="seconds"), user, status,
                   total, subject])
        self.state.add_log("__campaigns__", "Gmail", "-", "gmail_campaign",
                           result=cid, notes=name)
        return cid

    def add_gmail_campaign_lead(self, cid, lead, subject, message,
                                status="Draft", when_iso=""):
        ws = self._ensure_headers_sheet(GMAIL_CAMPAIGN_LEADS_SHEET,
                                        GMAIL_CAMPAIGN_LEADS_HEADERS)
        ws.append([cid, lead.key, lead.full_name, lead.email, lead.company,
                   subject, message, status, when_iso, ""])
        return ws.max_row

    def log_gmail_followup(self, lead, step, subject, message, status,
                           when_iso="", note=""):
        ws = self._ensure_headers_sheet(GMAIL_FOLLOWUPS_SHEET,
                                        GMAIL_FOLLOWUPS_HEADERS)
        ws.append([dt.datetime.now().isoformat(timespec="seconds"), lead.key,
                   lead.full_name, lead.email, step, subject, message, status,
                   when_iso, note])

    def gmail_followup1_candidates(self, min_hours=None):
        """Leads que recibieron Email 1, no respondieron, no están bloqueados y
        ya cumplieron el tiempo configurado. Devuelve (lead, info_dict)."""
        gap = min_hours if min_hours is not None else self.fu_gap_hours(0)
        now = dt.datetime.now()
        out = []
        for sheet in self.maps:
            for lead in self.all_leads(sheet):
                st = self.state.get(lead.key)
                if self._is_blocked(lead, st):
                    continue
                if not _is_marked(lead.stage_values.get("Email 1")):
                    continue
                if _is_marked(lead.stage_values.get("Email 2")):
                    continue
                lc = self._cell(self.wb[sheet], self.maps[sheet], lead.row,
                                "Last Contact")
                sent_date = None
                for e in st.log:
                    if e.get("action") == "sent" and                             str(e.get("step", "")).startswith("Email 1"):
                        sent_date = e.get("ts")
                        break
                base = sent_date or (str(lc) + "T09:00:00" if lc else None)
                if not base:
                    continue
                try:
                    t0 = dt.datetime.fromisoformat(str(base)[:19])
                except ValueError:
                    continue
                hours = (now - t0).total_seconds() / 3600
                if hours < gap:
                    continue
                last_msg = subj = ""
                for e in reversed(st.log):
                    if e.get("action") == "sent":
                        last_msg = (e.get("message") or "")[:80]
                        break
                for s in self.read_scheduled():
                    if s.get("Lead ID") == lead.key and "|SUBJ|" in str(
                            s.get("Notes") or ""):
                        subj = str(s["Notes"]).split("|SUBJ|")[-1]
                out.append((lead, {
                    "email_inicial": str(base)[:16].replace("T", " "),
                    "horas_sin_respuesta": round(hours),
                    "ultimo_mensaje": last_msg, "asunto_anterior": subj,
                    "proxima_accion": self.next_action(lead, st),
                }))
        return out


def create_blank_crm(path, industries=None):
    """Crea un CRM vacío y funcional: hoja(s) de leads con el encabezado
    canónico + Activity_Log, Notifications, Campaigns, Follow_Ups y
    Scheduled_Messages, con colores y validaciones sembrados."""
    industries = industries or ["Leads"]
    wb = Workbook()
    wb.remove(wb.active)
    for name in industries:
        ws = wb.create_sheet(name)
        ws.append(CANONICAL_HEADER)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
        ws.freeze_panes = "A2"
    wb.save(path)
    crm = CRM(path)
    crm._ensure_activity_sheet()
    crm._ensure_notifications_sheet()
    crm._ensure_campaigns_sheet()
    crm._ensure_history_sheet()      # Follow_Ups (Follow_Up_History)
    crm._ensure_scheduled_sheet()
    crm._ensure_color_sheet()
    crm.reanchor_all()               # siembra las 8 reglas de color + dropdowns
    crm.save(backup=False)
    return path


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _num(v, default):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default

def _is_marked(v) -> bool:
    """True si una celda de etapa cuenta como 'tocada' (no vacía / no Pendiente)."""
    if v is None:
        return False
    s = str(v).strip()
    return s != "" and s.lower() not in ("pendiente", "no", "no respuesta")
