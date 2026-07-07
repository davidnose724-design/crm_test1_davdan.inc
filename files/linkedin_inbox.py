"""
linkedin_inbox.py — Inbox LinkedIn en modo cumplimiento.

POSTURA (innegociable, y regla propia del proyecto desde el inicio):
- SIN scraping, SIN automatización del navegador (Playwright/Selenium), SIN
  lectura automática de mensajes privados contra la sesión de LinkedIn.
  El Acuerdo de Usuario de LinkedIn lo prohíbe y arriesga la cuenta.
- Fuentes de datos legítimas HOY:
  (a) Registro MANUAL de conversaciones (el usuario pega el mensaje).
  (b) La EXPORTACIÓN OFICIAL de datos de LinkedIn (Configuración → "Obtener
      una copia de tus datos" → messages.csv), que LinkedIn provee.
- Capa de adaptadores (LinkedInInboxProvider) lista para sustituir por una
  API oficial o proveedor autorizado sin reescribir el CRM.

Todo lo que escribe al CRM pasa por crm_core (Activity_Log, Notifications,
outcomes, cancelación de follow-ups).
"""

from __future__ import annotations
import csv
import datetime as dt
import hashlib
import io

from openpyxl.styles import Font, PatternFill

from crm_core import _norm, _is_marked

# --------------------------------------------------------------------------- #
# Hojas
# --------------------------------------------------------------------------- #
LI_INBOX_SHEET = "LinkedIn_Inbox"
LI_INBOX_HEADERS = ["Message ID", "Conversation ID", "Date", "Sender",
                    "Sender Profile", "Text", "Lead ID", "Category",
                    "Confidence", "Requires Review", "Processed At"]

LI_NOTIF_SHEET = "LinkedIn_Notifications"
LI_NOTIF_HEADERS = ["id", "user_id", "lead_id", "campaign_id", "type", "title",
                    "message", "created_at", "read_at", "priority",
                    "action_url_or_reference"]

NOTIF_TYPES = ["NEW_REPLY", "INTERESTED_LEAD", "QUESTION_RECEIVED",
               "OBJECTION_RECEIVED", "NOT_INTERESTED", "DO_NOT_CONTACT",
               "POSSIBLE_BLOCK", "SESSION_EXPIRED", "VERIFICATION_REQUIRED",
               "BROWSER_ERROR", "SYNC_ERROR"]

# --------------------------------------------------------------------------- #
# Máquina de estados LinkedIn (punto 14)
# --------------------------------------------------------------------------- #
LI_STATES = ["NOT_CONTACTED", "REVIEW_REQUIRED", "READY_TO_CONNECT",
             "CONNECTION_PENDING", "CONNECTED", "MESSAGE_PREPARED",
             "CONTACTED", "REPLIED", "INTERESTED", "NOT_INTERESTED",
             "DO_NOT_CONTACT", "BLOCKED", "PROSPECT_LATER", "BLACKLISTED",
             "MANUAL_REVIEW", "ERROR"]

_TERMINALISH = {"DO_NOT_CONTACT", "BLOCKED", "BLACKLISTED"}
TRANSITIONS = {
    "NOT_CONTACTED": {"REVIEW_REQUIRED", "READY_TO_CONNECT", "MANUAL_REVIEW"},
    "REVIEW_REQUIRED": {"READY_TO_CONNECT", "MANUAL_REVIEW", "NOT_CONTACTED"},
    "READY_TO_CONNECT": {"CONNECTION_PENDING", "MESSAGE_PREPARED",
                         "MANUAL_REVIEW"},
    "CONNECTION_PENDING": {"CONNECTED", "READY_TO_CONNECT", "MANUAL_REVIEW",
                           "ERROR"},
    "CONNECTED": {"MESSAGE_PREPARED", "CONTACTED", "MANUAL_REVIEW"},
    "MESSAGE_PREPARED": {"CONTACTED", "CONNECTED", "MANUAL_REVIEW"},
    "CONTACTED": {"REPLIED", "PROSPECT_LATER", "MANUAL_REVIEW", "ERROR"},
    "REPLIED": {"INTERESTED", "NOT_INTERESTED", "MANUAL_REVIEW",
                "PROSPECT_LATER"},
    "INTERESTED": {"PROSPECT_LATER", "MANUAL_REVIEW"},
    "NOT_INTERESTED": {"PROSPECT_LATER", "MANUAL_REVIEW"},
    "PROSPECT_LATER": {"READY_TO_CONNECT", "CONTACTED", "MANUAL_REVIEW"},
    "MANUAL_REVIEW": set(LI_STATES) - {"MANUAL_REVIEW"},
    "ERROR": {"MANUAL_REVIEW", "READY_TO_CONNECT"},
    "DO_NOT_CONTACT": set(), "BLOCKED": set(), "BLACKLISTED": set(),
}
# Desde cualquier estado no-terminal se puede ir a los de bloqueo/respuesta:
_ALWAYS_ALLOWED = {"DO_NOT_CONTACT", "BLOCKED", "BLACKLISTED", "REPLIED"}


def can_transition(current_state, new_state):
    cur = (current_state or "NOT_CONTACTED").strip().upper()
    new = (new_state or "").strip().upper()
    if new not in LI_STATES:
        return False
    if cur not in LI_STATES:
        cur = "NOT_CONTACTED"
    if cur in _TERMINALISH:
        return False          # los terminales no salen (protege DNC/BLOCKED)
    if new in _ALWAYS_ALLOWED:
        return True
    return new in TRANSITIONS.get(cur, set())


def get_lead_state(crm, lead):
    crm.ensure_columns(lead.sheet, ["LinkedIn State"])
    v = crm._cell(crm.wb[lead.sheet], crm.maps[lead.sheet], lead.row,
                  "LinkedIn State")
    return str(v).strip().upper() if v else "NOT_CONTACTED"


def transition_lead_state(crm, lead, new_state, reason="", actor="app"):
    """Aplica una transición válida; registra en Activity_Log. Devuelve
    (ok, estado_resultante)."""
    cur = get_lead_state(crm, lead)
    new = (new_state or "").strip().upper()
    if not can_transition(cur, new):
        crm.state.add_log(lead.key, "LinkedIn", "-", "li_state_denied",
                          result=f"{cur}→{new}", notes=reason)
        return False, cur
    crm._set(lead, "LinkedIn State", new)
    crm.state.add_log(lead.key, "LinkedIn", "-", "li_state",
                      result=f"{cur}→{new}", notes=f"{actor}: {reason}")
    return True, new


# --------------------------------------------------------------------------- #
# Clasificador de respuestas (punto 13) — reglas locales, listo para IA futura
# --------------------------------------------------------------------------- #
_RULES = [
    # (categoría, confianza, señales) — orden = prioridad
    ("BLOCKED", 0.9, ["te he bloqueado", "i blocked you", "reported your",
                      "te reporté", "he reportado", "stop messaging me or"]),
    ("DO_NOT_CONTACT", 0.9, ["do not contact", "no me contactes",
                             "no vuelvas a escribir", "remove me",
                             "unsubscribe", "déjame en paz", "stop contacting"]),
    ("NOT_INTERESTED", 0.8, ["not interested", "no me interesa",
                             "no estamos interesados", "no gracias",
                             "not a fit", "no necesitamos", "no lo ocupamos"]),
    ("OUT_OF_OFFICE", 0.85, ["out of office", "fuera de la oficina",
                             "on vacation", "de vacaciones", "maternity leave",
                             "regreso el", "estaré de vuelta"]),
    ("WRONG_PERSON", 0.8, ["not the right person", "no soy la persona",
                           "ya no trabajo", "no longer work", "wrong person",
                           "te paso con", "contacta a mi colega"]),
    ("OBJECTION", 0.7, ["muy caro", "too expensive", "ya tenemos proveedor",
                        "already have a supplier", "no tenemos presupuesto",
                        "no budget", "estamos bajo contrato"]),
    ("QUESTION", 0.7, ["?", "¿", "cuánto cuesta", "how much", "qué incluye",
                       "más información", "more info", "podrías explicar"]),
    ("INTERESTED", 0.8, ["me interesa", "interested", "agendemos", "let's talk",
                         "cotiza", "quote", "manda info", "send me",
                         "suena bien", "sounds good", "sí, hablemos"]),
    ("POSITIVE", 0.6, ["gracias por conectar", "thanks for connecting",
                       "mucho gusto", "nice to meet", "encantado", "hola,",
                       "buen día"]),
]

_SUGGESTED_ACTION = {
    "POSITIVE": "Registrar respuesta y continuar conversación",
    "INTERESTED": "Registrar respuesta como interesado y agendar siguiente paso",
    "QUESTION": "Responder la pregunta; registrar respuesta",
    "OBJECTION": "Responder objeción; registrar respuesta",
    "NOT_INTERESTED": "Cancelar follow-ups automáticos; marcar no interesado",
    "DO_NOT_CONTACT": "Detener TODO contacto futuro (Do Not Contact)",
    "BLOCKED": "Marcar Blocked; detener contacto; revisar cuenta",
    "OUT_OF_OFFICE": "Reprogramar follow-up para su regreso",
    "WRONG_PERSON": "Actualizar contacto correcto; revisión humana",
    "UNKNOWN": "Revisión humana requerida",
}


def classify_linkedin_response(message_text):
    """Clasifica un mensaje por reglas locales. BLOCKED solo con señal
    explícita en el TEXTO (nunca por ausencia de respuesta). UNKNOWN →
    revisión humana. Arquitectura lista para sustituir por IA."""
    raw = str(message_text or "")
    blob = _norm(raw)
    if not blob.strip():
        return {"category": "UNKNOWN", "confidence": 0.0,
                "reason": "Mensaje vacío",
                "suggested_crm_action": _SUGGESTED_ACTION["UNKNOWN"],
                "requires_human_review": True}
    for cat, conf, keys in _RULES:
        hits = [k for k in keys if _norm(k) and _norm(k) in blob]
        if cat == "QUESTION" and not hits and ("?" in raw or "¿" in raw):
            hits = ["?"]
        if hits:
            return {"category": cat, "confidence": conf,
                    "reason": f"Coincidencia: '{hits[0]}'",
                    "suggested_crm_action": _SUGGESTED_ACTION[cat],
                    "requires_human_review": cat in ("WRONG_PERSON",
                                                     "OBJECTION")}
    return {"category": "UNKNOWN", "confidence": 0.2,
            "reason": "Sin coincidencias con reglas conocidas",
            "suggested_crm_action": _SUGGESTED_ACTION["UNKNOWN"],
            "requires_human_review": True}


# --------------------------------------------------------------------------- #
# Providers (punto 11) — capa sustituible
# --------------------------------------------------------------------------- #
class LinkedInInboxProvider:
    """Interfaz. Implementaciones actuales: manual y exportación oficial.
    NO existe implementación Playwright/scraping en este proyecto; el slot
    para API oficial / proveedor autorizado queda listo (FutureAPIProvider)."""
    name = "base"

    def list_conversations(self):
        raise NotImplementedError

    def get_conversation(self, conversation_id):
        return [m for m in self.list_messages()
                if m["conversation_id"] == conversation_id]

    def get_recent_messages(self, conversation_id, n=10):
        return self.get_conversation(conversation_id)[-n:]

    def list_messages(self):
        raise NotImplementedError

    def get_unread_conversations(self, processed_ids):
        return sorted({m["conversation_id"] for m in self.list_messages()
                       if m["message_id"] not in processed_ids})


def _msg_id(conversation_id, date, sender, text):
    """Identificador estable del mensaje (idempotencia)."""
    raw = f"{conversation_id}|{date}|{sender}|{str(text)[:80]}"
    return hashlib.sha1(raw.encode("utf-8", "ignore")).hexdigest()[:16]


class ManualProvider(LinkedInInboxProvider):
    """El usuario pega una conversación/mensaje a mano."""
    name = "Registro manual"

    def __init__(self, entries=None):
        self.entries = entries or []

    def add(self, sender, text, profile_url="", conversation_id="",
            date=None):
        date = date or dt.datetime.now().isoformat(timespec="minutes")
        cid = conversation_id or f"manual-{_norm(sender)[:24]}"
        self.entries.append({
            "message_id": _msg_id(cid, date, sender, text),
            "conversation_id": cid, "date": str(date), "sender": sender,
            "sender_profile": profile_url, "text": text,
        })

    def list_messages(self):
        return list(self.entries)

    def list_conversations(self):
        return sorted({m["conversation_id"] for m in self.entries})


class DataExportProvider(LinkedInInboxProvider):
    """Lee messages.csv de la EXPORTACIÓN OFICIAL de datos de LinkedIn
    (Configuración → Obtener una copia de tus datos). Columnas típicas:
    CONVERSATION ID, CONVERSATION TITLE, FROM, SENDER PROFILE URL, TO,
    DATE, SUBJECT, CONTENT."""
    name = "Exportación oficial (messages.csv)"

    def __init__(self, csv_bytes, own_name=""):
        self.own = _norm(own_name)
        text = csv_bytes.decode("utf-8-sig", errors="replace")
        self.rows = list(csv.DictReader(io.StringIO(text)))
        self._msgs = None

    @staticmethod
    def _col(row, *names):
        for n in names:
            for k in row:
                if _norm(k) == _norm(n):
                    return row[k] or ""
        return ""

    def list_messages(self):
        if self._msgs is None:
            out = []
            for r in self.rows:
                sender = self._col(r, "FROM", "From")
                if self.own and _norm(sender) == self.own:
                    continue          # solo mensajes RECIBIDOS
                cid = self._col(r, "CONVERSATION ID") or \
                    self._col(r, "CONVERSATION TITLE") or "export"
                date = self._col(r, "DATE")
                text = self._col(r, "CONTENT", "Message", "BODY")
                if not str(text).strip():
                    continue
                out.append({
                    "message_id": _msg_id(cid, date, sender, text),
                    "conversation_id": cid, "date": date, "sender": sender,
                    "sender_profile": self._col(r, "SENDER PROFILE URL"),
                    "text": text,
                })
            self._msgs = out
        return self._msgs

    def list_conversations(self):
        return sorted({m["conversation_id"] for m in self.list_messages()})


class FutureAPIProvider(LinkedInInboxProvider):
    """Slot para una API oficial de LinkedIn o un proveedor autorizado.
    Sustituir esta clase NO requiere tocar el resto del CRM."""
    name = "API oficial (no configurada)"

    def list_messages(self):
        raise RuntimeError("No hay API oficial configurada. Usa registro "
                           "manual o la exportación oficial de LinkedIn.")

    def list_conversations(self):
        return []


# --------------------------------------------------------------------------- #
# Hojas + notificaciones LinkedIn (punto 15)
# --------------------------------------------------------------------------- #
def _ensure_sheet(crm, name, headers):
    if name not in crm.wb.sheetnames:
        ws = crm.wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
        ws.freeze_panes = "A2"
    return crm.wb[name]


def processed_message_ids(crm):
    if LI_INBOX_SHEET not in crm.wb.sheetnames:
        return set()
    ws = crm.wb[LI_INBOX_SHEET]
    return {str(ws.cell(r, 1).value) for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value}


def add_li_notification(crm, ntype, title, message, lead=None, campaign_id="",
                        priority="Media", ref=""):
    ws = _ensure_sheet(crm, LI_NOTIF_SHEET, LI_NOTIF_HEADERS)
    nid = f"LN-{ws.max_row:05d}"
    ws.append([nid, "user", lead.key if lead else "", campaign_id, ntype,
               title, message, dt.datetime.now().isoformat(timespec="seconds"),
               "", priority, ref or (lead.linkedin if lead else "")])
    return nid


def read_li_notifications(crm, unread_only=False):
    if LI_NOTIF_SHEET not in crm.wb.sheetnames:
        return []
    ws = crm.wb[LI_NOTIF_SHEET]
    out = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(LI_NOTIF_HEADERS) + 1)]
        if not vals[0]:
            continue
        d = dict(zip(LI_NOTIF_HEADERS, vals))
        d["_row"] = r
        if unread_only and d.get("read_at"):
            continue
        out.append(d)
    return out


def mark_li_notification_read(crm, row, read=True):
    ws = crm.wb[LI_NOTIF_SHEET]
    col = LI_NOTIF_HEADERS.index("read_at") + 1
    ws.cell(row=row, column=col).value = \
        dt.datetime.now().isoformat(timespec="seconds") if read else ""


# --------------------------------------------------------------------------- #
# Match, procesamiento e integración CRM (punto 12)
# --------------------------------------------------------------------------- #
def match_conversation_to_lead(crm, sender_name, profile_url=""):
    """Asocia por LinkedIn URL primero; luego por nombre exacto normalizado."""
    purl = _norm(profile_url).rstrip("/")
    for sheet in crm.maps:
        for lead in crm.all_leads(sheet):
            if purl and lead.linkedin and \
                    _norm(lead.linkedin).rstrip("/") == purl:
                return lead
    if sender_name:
        target = _norm(sender_name)
        for sheet in crm.maps:
            for lead in crm.all_leads(sheet):
                if _norm(lead.full_name) == target:
                    return lead
    return None


_CATEGORY_NOTIF = {
    "POSITIVE": ("NEW_REPLY", "Media"), "INTERESTED": ("INTERESTED_LEAD", "Alta"),
    "QUESTION": ("QUESTION_RECEIVED", "Alta"),
    "OBJECTION": ("OBJECTION_RECEIVED", "Media"),
    "NOT_INTERESTED": ("NOT_INTERESTED", "Media"),
    "DO_NOT_CONTACT": ("DO_NOT_CONTACT", "Alta"),
    "BLOCKED": ("POSSIBLE_BLOCK", "Alta"),
    "OUT_OF_OFFICE": ("NEW_REPLY", "Baja"),
    "WRONG_PERSON": ("NEW_REPLY", "Media"), "UNKNOWN": ("NEW_REPLY", "Media"),
}


def process_new_linkedin_message(crm, msg):
    """Pipeline del punto 12: idempotencia → match → guardar evento →
    clasificar → actualizar CRM → notificar. Devuelve dict resumen."""
    ws = _ensure_sheet(crm, LI_INBOX_SHEET, LI_INBOX_HEADERS)
    mid = msg.get("message_id") or _msg_id(
        msg.get("conversation_id", ""), msg.get("date", ""),
        msg.get("sender", ""), msg.get("text", ""))
    if mid in processed_message_ids(crm):
        return {"status": "duplicate", "message_id": mid}

    lead = match_conversation_to_lead(crm, msg.get("sender", ""),
                                      msg.get("sender_profile", ""))
    cls = classify_linkedin_response(msg.get("text", ""))
    cat = cls["category"]

    # guardar evento (antes de efectos: idempotencia primero)
    ws.append([mid, msg.get("conversation_id", ""), msg.get("date", ""),
               msg.get("sender", ""), msg.get("sender_profile", ""),
               str(msg.get("text", ""))[:500],
               lead.key if lead else "", cat, cls["confidence"],
               "Sí" if cls["requires_human_review"] else "No",
               dt.datetime.now().isoformat(timespec="seconds")])

    # efectos en CRM (solo si hay lead asociado)
    if lead is not None:
        if cat == "DO_NOT_CONTACT":
            crm.set_outcome(lead, "Do Not Contact",
                            note=str(msg.get("text", ""))[:120])
            crm.cancel_scheduled_for_lead(lead)
            transition_lead_state(crm, lead, "DO_NOT_CONTACT",
                                  cls["reason"], "inbox")
        elif cat == "BLOCKED":
            crm._set(lead, "Outcome Status", "Blocked")
            crm.cancel_scheduled_for_lead(lead)
            transition_lead_state(crm, lead, "BLOCKED", cls["reason"], "inbox")
        elif cat == "NOT_INTERESTED":
            crm.register_response(lead, "LinkedIn", msg.get("text", ""))
            crm.cancel_scheduled_for_lead(lead)   # cancela follow-ups autos
            transition_lead_state(crm, lead, "REPLIED", cls["reason"], "inbox")
            transition_lead_state(crm, lead, "NOT_INTERESTED",
                                  cls["reason"], "inbox")
        elif cat in ("INTERESTED", "POSITIVE", "QUESTION", "OBJECTION"):
            crm.register_response(lead, "LinkedIn", msg.get("text", ""))
            transition_lead_state(crm, lead, "REPLIED", cls["reason"], "inbox")
            if cat == "INTERESTED":
                transition_lead_state(crm, lead, "INTERESTED",
                                      cls["reason"], "inbox")
        elif cat == "OUT_OF_OFFICE":
            crm.state.add_log(lead.key, "LinkedIn", "-", "ooo",
                              result="OUT_OF_OFFICE",
                              notes=str(msg.get("text", ""))[:80])
        else:   # WRONG_PERSON / UNKNOWN → sin cambios de estado automáticos
            transition_lead_state(crm, lead, "MANUAL_REVIEW",
                                  cls["reason"], "inbox")

    ntype, prio = _CATEGORY_NOTIF.get(cat, ("NEW_REPLY", "Media"))
    who = lead.full_name if lead else msg.get("sender", "desconocido")
    add_li_notification(crm, ntype, f"{ntype}: {who}",
                        str(msg.get("text", ""))[:140], lead=lead,
                        priority=prio)
    return {"status": "processed", "message_id": mid, "lead":
            (lead.full_name if lead else None), "category": cat,
            "requires_human_review": cls["requires_human_review"]}


def sync_inbox(crm, provider):
    """Sincronización MANUAL (botón) o por intervalo conservador — nunca un
    loop dentro de Streamlit. Procesa solo mensajes nuevos (idempotente)."""
    done = processed_message_ids(crm)
    results = {"processed": 0, "duplicates": 0, "matched": 0, "review": 0}
    try:
        msgs = provider.list_messages()
    except Exception as e:
        add_li_notification(crm, "SYNC_ERROR", "Error de sincronización",
                            str(e)[:140], priority="Alta")
        return {"error": str(e), **results}
    for m in msgs:
        if m["message_id"] in done:
            results["duplicates"] += 1
            continue
        r = process_new_linkedin_message(crm, m)
        if r["status"] == "processed":
            results["processed"] += 1
            done.add(r["message_id"])
            if r.get("lead"):
                results["matched"] += 1
            if r.get("requires_human_review"):
                results["review"] += 1
        else:
            results["duplicates"] += 1
    return results


# --------------------------------------------------------------------------- #
# Preparación de mensaje (punto 10) — NUNCA envía; aprobar ≠ enviar
# --------------------------------------------------------------------------- #
def prepare_linkedin_message(crm, lead, campaign=None, icp=None):
    """Genera un borrador personalizado con preview, conteo y warnings.
    El envío es SIEMPRE manual del usuario; aprobar solo marca el estado
    MESSAGE_PREPARED."""
    icp = icp or crm.read_icp()
    fit = crm.evaluate_lead_fit(lead, icp)
    msg = fit["mensaje"]
    warnings = []
    if len(msg) > 300:
        warnings.append(f"{len(msg)} caracteres: las notas de conexión de "
                        "LinkedIn permiten máx. 300")
    if "{{" in msg or "{" in msg and "}" in msg and "{{" in msg:
        warnings.append("Quedan variables sin resolver en el mensaje")
    if not (lead.full_name or "").strip():
        warnings.append("El lead no tiene nombre: personalización incompleta")
    if not lead.linkedin:
        warnings.append("Sin LinkedIn URL: no podrás abrir el perfil")
    st_ = crm.state.get(lead.key)
    if crm._is_blocked(lead, st_):
        warnings.append(f"Lead bloqueado/cerrado ({lead.outcome}): NO enviar")
    return {
        "lead": lead.full_name, "company": lead.company,
        "job_title": lead.job_title,
        "campaign": (campaign or {}).get("Name", "") if campaign else "",
        "message": msg, "char_count": len(msg), "warnings": warnings,
        "fit": fit,
    }
