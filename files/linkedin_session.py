"""
linkedin_session.py — Sesión LinkedIn + ICP + evaluación + prepare_connection.

POSTURA (innegociable, regla del proyecto desde el inicio):
- NO se automatiza el navegador contra la sesión de LinkedIn (sin Playwright/
  Selenium abriendo login, guardando contexto autenticado o ejecutando
  conexiones). El Acuerdo de Usuario de LinkedIn lo prohíbe y arriesga la cuenta.
- La arquitectura de "launchers" queda pluggable para el futuro, pero solo el
  launcher MANUAL está activo: "Open LinkedIn" abre el sitio en TU navegador y
  el login es tuyo. Los launchers Local/Remote/Server están declarados y
  DESHABILITADOS con un mensaje claro (slot para API oficial/proveedor
  autorizado, sin acoplar el CRM a una sola forma).
- Nunca se marca "Connected" sin verificación real (aquí: confirmación humana
  de sesión iniciada). CAPTCHA/2FA/checkpoint => verificación manual, jamás
  resolución automática.
- prepare_connection valida y prepara; la EJECUCIÓN de la conexión la hace el
  usuario manualmente en su navegador. No hay envíos masivos ni loops.
"""

from __future__ import annotations
import datetime as dt
import json
import os
import uuid

from openpyxl.styles import Font, PatternFill

from crm_core import _norm, ICP_FIELDS

# --------------------------------------------------------------------------- #
# Estados de sesión y hojas
# --------------------------------------------------------------------------- #
SESSION_STATES = ["Disconnected", "Connected", "Verification Required",
                  "Session Expired", "Error"]

ICP_STORE_SHEET = "ICP_Profiles"
ICP_STORE_HEADERS = ["id", "name", "created_at", "updated_at", "created_by",
                     "version", "status", "criteria_json"]

CONNECTION_LOG_SHEET = "LinkedIn_Connection_Log"
CONNECTION_LOG_HEADERS = ["lead_id", "user_id", "campaign_id", "action_type",
                          "timestamp", "result", "error", "browser_session_id"]

COOLDOWN_DAYS = 14   # no volver a preparar conexión al mismo lead tan seguido


# --------------------------------------------------------------------------- #
# Launchers (pluggable). Solo el manual está activo.
# --------------------------------------------------------------------------- #
class SessionLauncher:
    """Interfaz. La app NO se acopla a una sola forma de 'lanzar' sesión."""
    name = "base"
    enabled = False
    reason_disabled = ""

    def open_linkedin_url(self):
        return "https://www.linkedin.com/feed/"


class ManualSessionLauncher(SessionLauncher):
    """Único activo: el usuario abre LinkedIn en SU navegador e inicia sesión.
    No hay contexto automatizado ni credenciales gestionadas por la app."""
    name = "Manual (tu navegador)"
    enabled = True


class LocalBrowserLauncher(SessionLauncher):
    name = "LocalBrowserLauncher"
    enabled = False
    reason_disabled = ("Automatizar el navegador contra LinkedIn viola su "
                       "Acuerdo de Usuario y arriesga la cuenta. Deshabilitado "
                       "por política del proyecto.")


class RemoteBrowserLauncher(SessionLauncher):
    name = "RemoteBrowserLauncher"
    enabled = False
    reason_disabled = ("Requiere un proveedor de navegador remoto autorizado. "
                       "No se automatiza LinkedIn desde el CRM.")


class BrowserServerLauncher(SessionLauncher):
    name = "BrowserServerLauncher"
    enabled = False
    reason_disabled = ("Slot para una API oficial de LinkedIn / proveedor "
                       "autorizado. No configurado.")


LAUNCHERS = {
    "manual": ManualSessionLauncher(),
    "local": LocalBrowserLauncher(),
    "remote": RemoteBrowserLauncher(),
    "server": BrowserServerLauncher(),
}


def gui_available():
    """¿Hay entorno gráfico para un navegador local? En Codespaces/servidores
    headless normalmente NO. Se usa solo para informar al usuario."""
    if os.name == "nt" or os.sys.platform == "darwin":
        return True
    return bool(os.environ.get("DISPLAY") or os.environ.get("WAYLAND_DISPLAY"))


def environment_note():
    if not gui_available():
        return ("Entorno sin interfaz gráfica detectado (p. ej. Codespaces / "
                "servidor headless): no es posible lanzar un navegador local. "
                "Usa el modo manual: abre LinkedIn en tu equipo e inicia sesión "
                "tú mismo.")
    return ("Modo manual: abre LinkedIn en tu navegador e inicia sesión tú. "
            "La app no automatiza el navegador.")


# --------------------------------------------------------------------------- #
# Estado de sesión (en session_state de Streamlit; sin credenciales)
# --------------------------------------------------------------------------- #
def new_session_id():
    return f"sess-{uuid.uuid4().hex[:12]}"


def blank_session():
    return {"status": "Disconnected", "session_id": "", "confirmed_at": "",
            "note": ""}


def set_session_status(sess, status, note=""):
    if status not in SESSION_STATES:
        status = "Error"
    sess["status"] = status
    sess["note"] = note
    if status == "Connected":
        sess["confirmed_at"] = dt.datetime.now().isoformat(timespec="seconds")
        if not sess.get("session_id"):
            sess["session_id"] = new_session_id()
    return sess


# --------------------------------------------------------------------------- #
# ICP versionado y persistido (punto 6)
# --------------------------------------------------------------------------- #
def _ensure_sheet(crm, name, headers):
    if name not in crm.wb.sheetnames:
        ws = crm.wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
        ws.freeze_panes = "A2"
    return crm.wb[name]


def suggest_icp_options(crm, industries=None):
    """Sugiere valores para el ICP a partir de los leads cargados. Si se pasan
    industrias, restringe empresas/puestos/seniority a esas industrias."""
    inds, comps, titles, sens, countries = set(), set(), set(), set(), set()
    want = {_norm(i) for i in (industries or [])}
    for sheet in crm.maps:
        for l in crm.all_leads(sheet):
            if l.industry:
                inds.add(l.industry)
            in_scope = (not want) or (_norm(l.industry) in want)
            if not in_scope:
                continue
            if l.company:
                comps.add(l.company)
            if l.job_title:
                titles.add(l.job_title)
            if l.seniority_level:
                sens.add(l.seniority_level)
            loc = crm._cell(crm.wb[l.sheet], crm.maps[l.sheet], l.row,
                            "Location") or ""
            if loc:
                # último token como país aproximado (dato tal cual, sin inventar)
                countries.add(str(loc).split(",")[-1].strip())
    def top(s, n=40):
        return sorted(x for x in s if x)[:n]
    return {"industries": top(inds), "companies": top(comps, 60),
            "job_titles": top(titles, 60), "seniorities": top(sens),
            "countries": top(countries)}


def save_icp_profile(crm, name, criteria: dict, created_by="user",
                     status="confirmed", existing_id=None):
    """Guarda un ICP versionado en ICP_Profiles (criteria como JSON). Si
    existing_id, incrementa versión y updated_at."""
    ws = _ensure_sheet(crm, ICP_STORE_SHEET, ICP_STORE_HEADERS)
    now = dt.datetime.now().isoformat(timespec="seconds")
    cjson = json.dumps(criteria, ensure_ascii=False)
    if existing_id:
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value) == existing_id:
                ver = int(ws.cell(r, 6).value or 1) + 1
                ws.cell(r, 4).value = now
                ws.cell(r, 6).value = ver
                ws.cell(r, 7).value = status
                ws.cell(r, 8).value = cjson
                return existing_id, ver
    icp_id = f"ICP-{ws.max_row:04d}"
    ws.append([icp_id, name, now, now, created_by, 1, status, cjson])
    return icp_id, 1


def read_icp_profiles(crm):
    if ICP_STORE_SHEET not in crm.wb.sheetnames:
        return []
    ws = crm.wb[ICP_STORE_SHEET]
    out = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(ICP_STORE_HEADERS) + 1)]
        if not vals[0]:
            continue
        d = dict(zip(ICP_STORE_HEADERS, vals))
        d["_row"] = r
        try:
            d["criteria"] = json.loads(d.get("criteria_json") or "{}")
        except Exception:
            d["criteria"] = {}
        out.append(d)
    return out


def confirmed_icp(crm):
    """Devuelve el ICP confirmado más reciente, o None (gate de campaña)."""
    profs = [p for p in read_icp_profiles(crm)
             if str(p.get("status")) == "confirmed"]
    return profs[-1] if profs else None


# --------------------------------------------------------------------------- #
# Evaluación de lead vs ICP (punto 7) — contrato exacto
# --------------------------------------------------------------------------- #
def _terms(criteria, key):
    v = criteria.get(key)
    if isinstance(v, list):
        return [str(x) for x in v if str(x).strip()]
    return [t.strip() for t in str(v or "").split(",") if t.strip()]


def evaluate_lead_against_icp(crm, lead, icp):
    """Evalúa un lead contra un ICP (dict con 'criteria'). Devuelve:
    eligible, score, matched_criteria, failed_criteria, warnings,
    recommended_action, reason, fit_class."""
    crit = icp.get("criteria", icp) if isinstance(icp, dict) else {}
    st = crm.state.get(lead.key)
    loc = crm._cell(crm.wb[lead.sheet], crm.maps[lead.sheet], lead.row,
                    "Location") or ""
    notes = crm._cell(crm.wb[lead.sheet], crm.maps[lead.sheet], lead.row,
                      "Notes") or ""
    matched, failed, warnings = [], [], []

    def hit(field, value, label, positive=True):
        terms = _terms(crit, field)
        if not terms:
            return None
        blob = _norm(str(value or ""))
        found = [t for t in terms if _norm(t) and _norm(t) in blob]
        if found:
            (matched if positive else failed).append(
                f"{label}: {', '.join(found[:2])}")
            return True
        return False

    # exclusiones duras -> EXCLUDED
    excl = False
    for f, val, lbl in [("excluded_companies", lead.company, "Empresa excluida"),
                        ("excluded_job_titles", lead.job_title,
                         "Puesto excluido"),
                        ("excluded_industries", lead.industry,
                         "Industria excluida"),
                        ("leads_to_avoid",
                         f"{lead.full_name} {lead.company} {lead.job_title}",
                         "Lead a evitar")]:
        terms = _terms(crit, f)
        if terms and any(_norm(t) in _norm(str(val)) for t in terms if _norm(t)):
            failed.append(lbl)
            excl = True
    if crm._is_blocked(lead, st):
        failed.append(f"Estado CRM bloqueado/cerrado: {lead.outcome}")
        excl = True
    if excl:
        return {"eligible": False, "score": 0, "matched_criteria": matched,
                "failed_criteria": failed, "warnings": warnings,
                "recommended_action": "EXCLUDE",
                "reason": "Coincide con exclusiones o estado cerrado",
                "fit_class": "EXCLUDED"}

    score = 0
    weights = [("target_industries", lead.industry, "Industria", 25),
               ("job_titles", lead.job_title, "Puesto", 20),
               ("seniorities", (lead.seniority_level or "") + " " +
                (lead.job_title or ""), "Seniority", 15),
               ("target_companies", lead.company, "Empresa", 10),
               ("countries", loc, "País/Ubicación", 10)]
    any_criteria = False
    for field, value, label, w in weights:
        terms = _terms(crit, field)
        if not terms:
            continue
        any_criteria = True
        r = hit(field, value, label, positive=True)
        if r:
            score += w
        else:
            failed.append(f"{label} fuera de objetivo")

    kw = _terms(crit, "keywords")
    if kw:
        blob = _norm(f"{lead.job_title} {lead.company} {notes}")
        found = [k for k in kw if _norm(k) and _norm(k) in blob]
        if found:
            score += 10
            matched.append(f"Keywords: {', '.join(found[:3])}")

    # sin datos de contacto para prospección LinkedIn
    if not lead.linkedin:
        warnings.append("Sin LinkedIn URL: no se puede prospectar por LinkedIn")
    if not any_criteria:
        warnings.append("El ICP no define criterios positivos: score limitado")

    score = max(0, min(100, score))
    if score >= 70:
        fit, elig, act = "HIGH_FIT", True, "PROSPECT"
    elif score >= 45:
        fit, elig, act = "MEDIUM_FIT", True, "REVIEW"
    elif score >= 20:
        fit, elig, act = "LOW_FIT", False, "REVIEW"
    else:
        fit, elig, act = "LOW_FIT", False, "SKIP"
    reason = (f"{len(matched)} criterios cumplidos, {len(failed)} fallidos; "
              f"score {score}")
    return {"eligible": elig and bool(lead.linkedin), "score": score,
            "matched_criteria": matched, "failed_criteria": failed,
            "warnings": warnings, "recommended_action": act, "reason": reason,
            "fit_class": fit}


# --------------------------------------------------------------------------- #
# prepare_connection (punto 9) — valida y prepara; NO ejecuta
# --------------------------------------------------------------------------- #
def _last_connection_action(crm, lead_key):
    if CONNECTION_LOG_SHEET not in crm.wb.sheetnames:
        return None
    ws = crm.wb[CONNECTION_LOG_SHEET]
    last = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == lead_key:
            last = ws.cell(r, 5).value
    return last


def prepare_connection(crm, lead, icp, campaign_id=""):
    """Ejecuta las 10 verificaciones y devuelve un preview + checklist. NO abre
    navegador ni envía nada: la conexión la hace el usuario manualmente tras
    pulsar 'Confirm Connection Action'."""
    checks = []
    ok = True

    def step(name, passed, detail=""):
        nonlocal ok
        checks.append({"paso": name, "ok": bool(passed), "detalle": detail})
        if not passed:
            ok = False

    # 1) cargar lead
    step("1. Cargar lead", True, lead.full_name)
    # 2) verificar ICP
    ev = evaluate_lead_against_icp(crm, lead, icp)
    step("2. Cumple ICP", ev["fit_class"] in ("HIGH_FIT", "MEDIUM_FIT"),
         f"{ev['fit_class']} (score {ev['score']})")
    # 3) blacklist / bloqueos
    stt = crm.state.get(lead.key)
    step("3. No bloqueado", not crm._is_blocked(lead, stt),
         lead.outcome or "activo")
    # 4) historial
    touched = any(e.get("action") == "sent" for e in stt.log)
    step("4. Historial revisado", True,
         "con contacto previo" if touched else "sin contacto previo")
    # 5) cooldown
    last = _last_connection_action(crm, lead.key)
    cooldown_ok = True
    if last:
        try:
            days = (dt.datetime.now() -
                    dt.datetime.fromisoformat(str(last)[:19])).days
            cooldown_ok = days >= COOLDOWN_DAYS
            step("5. Cooldown", cooldown_ok,
                 f"última acción hace {days} días (mín {COOLDOWN_DAYS})")
        except ValueError:
            step("5. Cooldown", True, "sin fecha previa válida")
    else:
        step("5. Cooldown", True, "sin acciones previas")
    # 6) duplicados
    step("6. Sin duplicado de conexión", cooldown_ok,
         "ya había una acción reciente" if not cooldown_ok else "ok")
    # 7) abrir perfil (manual)
    step("7. Perfil disponible", bool(lead.linkedin),
         lead.linkedin or "sin LinkedIn URL")
    # 8) preparar acción
    from crm_core import CRM
    step("8. Acción preparada", True, "Solicitud de conexión (manual)")
    # 9) preview
    preview = (f"Conectar con {lead.full_name} — {lead.job_title or ''} @ "
               f"{lead.company}. Abre su perfil y envía la solicitud tú mismo.")
    step("9. Preview generado", True, preview)
    # 10) confirmación (la da el usuario en la UI)
    step("10. Requiere confirmación del usuario", True,
         "pulsa 'Confirm Connection Action'")

    return {"ready": ok and bool(lead.linkedin), "checks": checks,
            "evaluation": ev, "preview": preview,
            "profile_url": lead.linkedin or ""}


def log_connection_action(crm, lead, action_type, result, error="",
                          user_id="user", campaign_id="", session_id=""):
    """Registra la acción de conexión (punto 9). action_type p.ej.
    'connection_confirmed' | 'connection_skipped'."""
    ws = _ensure_sheet(crm, CONNECTION_LOG_SHEET, CONNECTION_LOG_HEADERS)
    ws.append([lead.key, user_id, campaign_id, action_type,
               dt.datetime.now().isoformat(timespec="seconds"), result, error,
               session_id])
    crm.state.add_log(lead.key, "LinkedIn", "-", "connection",
                      result=f"{action_type}:{result}", notes=error)
